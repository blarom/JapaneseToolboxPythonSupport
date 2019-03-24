#!/usr/bin/python -tt

import re
import openpyxl

import Constants
from JapaneseToolboxConverter import Converter

current_entry = ''

# region Preparations
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji.xlsx', data_only=True)
wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalMeaningsFR = localWordsWorkbook["MeaningsFR"]
wsLocalMeaningsES = localWordsWorkbook["MeaningsES"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]


class Word:
    def __init__(self, kanji, hiragana, romaji, french_meanings, spanish_meanings):
        self.kanji = kanji
        self.hiragana = hiragana
        self.romaji = romaji
        self.uniqueID = romaji + "zzz" + kanji
        self.french_meanings = french_meanings
        self.spanish_meanings = spanish_meanings


converter = Converter()
# endregion

# region Filling the foreign words list
words = []
reached_first_entry = False
with open("JMDict", encoding='utf-8') as infile:
    for line in infile:

        if "<entry>" in line:
            reached_first_entry = True
        if not reached_first_entry:
            continue

        current_entry += line

        if "</entry>" in line:
            match = re.search("<keb>(\w+)</keb>", current_entry)
            kanji = ''
            if match:
                kanji = match.group(1)

            match = re.search("<reb>(\w+)</reb>", current_entry)
            hiragana = ''
            if match:
                hiragana = match.group(1)
                if kanji == '': kanji = hiragana

            senses = re.findall(r"<sense>(.*?)</sense>", current_entry, re.MULTILINE | re.DOTALL)

            french_meanings = []
            spanish_meanings = []
            for sense in senses:
                french_meanings_in_sense = re.findall(r"<gloss xml:lang=\"fre\">(.+?)</gloss>", sense, re.MULTILINE | re.DOTALL)
                if len(french_meanings_in_sense) > 0: french_meanings.append(", ".join(french_meanings_in_sense))
                spanish_meanings_in_sense = re.findall(r"<gloss xml:lang=\"spa\">(.+?)</gloss>", sense, re.MULTILINE | re.DOTALL)
                if len(spanish_meanings_in_sense) > 0: spanish_meanings.append(", ".join(spanish_meanings_in_sense))

            if not kanji == '' and not hiragana == '' \
                    and not (len(french_meanings) == 0 and len(spanish_meanings) == 0):
                word = Word(kanji,
                            hiragana,
                            converter.get_latin_hiragana_katakana(hiragana)[converter.TYPE_LATIN],
                            french_meanings,
                            spanish_meanings)
                words.append(word)

            current_entry = ''


# endregion

# region Updating the Types and Meaning sheets with the values
def binary_search(words_list, text, lo=0, hi=None):
    # https://stackoverflow.com/questions/4161629/search-a-list-of-objects-in-python
    if hi is None:
        hi = len(words_list)
    while lo < hi:
        mid = (lo + hi) // 2

        if text > words_list[mid].uniqueID:
            lo = mid + 1
        elif text < words_list[mid].uniqueID:
            hi = mid
        elif text == words_list[mid].uniqueID:
            return mid
        else:
            return -1

    return -1


localWordsList = []
jmDictWordsList = list(words)
jmDictWordsList.sort(key=lambda x: x.uniqueID)
meaningFRindex = 2
meaningESindex = 2
for i in range(0, 3):

    if i == 0:
        typesIndex = 2
        workingWorksheet = wsLocalTypes
    elif i == 1:
        typesIndex = 2
        workingWorksheet = wsLocalGrammar
    else:
        typesIndex = 4
        workingWorksheet = wsLocalVerbs

    while True:
        if i==0 or i==1:
            romaji = workingWorksheet.cell(row=typesIndex, column=Constants.TYPES_COL_ROMAJI).value
            kanji = workingWorksheet.cell(row=typesIndex, column=Constants.TYPES_COL_KANJI).value

            if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
                break
        else:
            romaji = workingWorksheet.cell(row=typesIndex, column=Constants.VERBS_COL_ROMAJI).value
            kanji = workingWorksheet.cell(row=typesIndex, column=Constants.VERBS_COL_KANJI).value

            if (workingWorksheet.cell(row=typesIndex, column=1).value == '-'\
                    and workingWorksheet.cell(row=typesIndex, column=2).value == '-'\
                    and workingWorksheet.cell(row=typesIndex, column=3).value == '-'):
                break

            if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
                typesIndex += 1
                continue


        localWordsList.append(Word(kanji, "", romaji, "", ""))

        uniqueID = romaji + "zzz" + kanji
        # print(uniqueID + "\n")

        index_of_first_hit = binary_search(jmDictWordsList, uniqueID)
        if index_of_first_hit != -1:

            word = jmDictWordsList[index_of_first_hit]

            # print(word.romaji + "-" + word.kanji + "\n")

            meaningIndexesFR = []
            for j in range(len(word.french_meanings)):
                wsLocalMeaningsFR.cell(row=meaningFRindex, column=Constants.MEANINGS_COL_INDEX).value = meaningFRindex
                wsLocalMeaningsFR.cell(row=meaningFRindex, column=Constants.MEANINGS_COL_MEANING).value = word.french_meanings[j].replace('œ','oe')
                wsLocalMeaningsFR.cell(row=meaningFRindex, column=Constants.MEANINGS_COL_TYPE).value = ""
                meaningIndexesFR.append(str(meaningFRindex))
                meaningFRindex += 1
            workingWorksheet.cell(row=typesIndex,
                                  column=Constants.TYPES_COL_MEANINGS_FR if i < 2 else Constants.VERBS_COL_MEANINGS_FR
                                  ).value = ";".join(meaningIndexesFR)

            meaningIndexesES = []
            for j in range(len(word.spanish_meanings)):
                wsLocalMeaningsES.cell(row=meaningESindex, column=Constants.MEANINGS_COL_INDEX).value = meaningESindex
                wsLocalMeaningsES.cell(row=meaningESindex, column=Constants.MEANINGS_COL_MEANING).value = word.spanish_meanings[j]
                wsLocalMeaningsES.cell(row=meaningESindex, column=Constants.MEANINGS_COL_TYPE).value = ""
                meaningIndexesES.append(str(meaningESindex))
                meaningESindex += 1
            workingWorksheet.cell(row=typesIndex,
                                  column=Constants.TYPES_COL_MEANINGS_ES if i < 2 else Constants.VERBS_COL_MEANINGS_ES
                                  ).value = ";".join(meaningIndexesES)

        typesIndex += 1
# endregion

# region Creating a list of words that are not in the local database
newWords = []
localWordsList.sort(key=lambda x: x.uniqueID)
for word in jmDictWordsList:
    index_of_first_hit = binary_search(localWordsList, word.uniqueID)
    if index_of_first_hit == -1:
        newWords.append(word)

print("Number of new words: " + str(len(newWords)))

# endregion

# Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with JMDict_foreign.xlsx')
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with JMDict_foreign.xlsx')
