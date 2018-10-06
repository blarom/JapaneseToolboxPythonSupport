# Japanese Toolbox database creator

import openpyxl


#region Getting the Excel sheets and their sizes
# Preparing the excel sheet for writing
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with firebase results.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with firebase results.xlsx', data_only=True)

# assign_sheet=wb.active
wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]

# Getting the size of the Local Types list
lastLocalTypesIndex = 1
while True:
    value = wsLocalTypes.cell(row=lastLocalTypesIndex, column=3).value
    if not value:
        lastLocalTypesIndex -= 1
        break
    lastLocalTypesIndex += 1

# Getting the size of the Local Verbs list
lastLocalVerbsIndex = 1
while True:
    value = wsLocalVerbs.cell(row=lastLocalVerbsIndex, column=3).value
    value2 = wsLocalVerbs.cell(row=lastLocalVerbsIndex+1, column=3).value
    value3 = wsLocalVerbs.cell(row=lastLocalVerbsIndex+2, column=3).value
    if (not value) & (not value2) & (not value3):
        lastLocalVerbsIndex -= 1
        break
    lastLocalVerbsIndex += 1
#endregion

#region Moving the suru verb results from the Grammar/Types to the Verbs/Verbs sheet
typesSheetIndex = lastLocalTypesIndex
verbSheetIndex = lastLocalVerbsIndex
while " suru" in wsLocalTypes.cell(row=typesSheetIndex, column=3).value:
    romaji = str(wsLocalTypes.cell(row=typesSheetIndex, column=3).value)
    kanji = str(wsLocalTypes.cell(row=typesSheetIndex, column=4).value)
    meanings = str(wsLocalTypes.cell(row=typesSheetIndex, column=5).value)
    altSpellings = str(wsLocalTypes.cell(row=typesSheetIndex, column=6).value)
    if not altSpellings or altSpellings == "None": altSpellings = ""

    wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = "suru verb"
    wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = "I"
    wsLocalVerbs.cell(row=verbSheetIndex, column=7).value = romaji
    wsLocalVerbs.cell(row=verbSheetIndex, column=6).value = kanji
    wsLocalVerbs.cell(row=verbSheetIndex, column=12).value = meanings
    wsLocalVerbs.cell(row=verbSheetIndex, column=11).value = altSpellings
    wsLocalVerbs.cell(row=verbSheetIndex, column=9).value = romaji[:-5]
    wsLocalVerbs.cell(row=verbSheetIndex, column=8).value = kanji[:-2]
    wsLocalVerbs.cell(row=verbSheetIndex, column=13).value = kanji[:-2]
    wsLocalVerbs.cell(row=verbSheetIndex, column=10).value = 68

    meaningIndexStrings = meanings.split(";")
    combinedMeanings = ""
    currentType = "I"
    for index in range(len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        currentMeaning = str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value).strip().replace("\u200b", "")
        currentType = str(wsLocalMeanings.cell(row=meaningIndexValue, column=3).value).strip().replace("\u200b", "")
        if index>0: combinedMeanings += ", "
        combinedMeanings += currentMeaning

    wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = combinedMeanings
    wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = currentType[-1]
    if currentType[-1] == "T": wsLocalVerbs.cell(row=verbSheetIndex, column=4).value = "を"

    wsLocalTypes.delete_rows(typesSheetIndex, 1)

    typesSheetIndex -= 1
    verbSheetIndex += 1

#wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = "-"
#wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = "-"
#wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = "-"
#endregion

#region Moving the regular verbs from the Grammar/Types to the Verbs/Verbs sheet
while not wsLocalTypes.cell(row=typesSheetIndex, column=2).value:

    romaji = str(wsLocalTypes.cell(row=typesSheetIndex, column=3).value)
    kanji = str(wsLocalTypes.cell(row=typesSheetIndex, column=4).value)
    meanings = str(wsLocalTypes.cell(row=typesSheetIndex, column=5).value)
    altSpellings = str(wsLocalTypes.cell(row=typesSheetIndex, column=6).value)
    if not altSpellings or altSpellings == "None": altSpellings = ""

    meaningIndexStrings = meanings.split(";")
    combinedMeanings = ""
    currentType = "I"
    for index in range(len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        currentMeaning = str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value).strip().replace("\u200b", "")
        currentType = str(wsLocalMeanings.cell(row=meaningIndexValue, column=3).value).strip().replace("\u200b", "")
        if index>0: combinedMeanings += ", "
        combinedMeanings += currentMeaning

    family = ""
    romajiRoot = ""
    if kanji[-1] == "す":
        family = "su godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "く":
        family = "ku godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "ぐ":
        family = "gu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "ぶ":
        family = "bu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "む":
        family = "mu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "ぬ":
        family = "nu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "る" and "rui" in currentType:
        family = "ru godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "る" and "rug" in currentType:
        family = "ru ichidan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "つ":
        family = "tsu godan"
        romajiRoot = romaji[:-3]
    elif kanji[-1] == "う":
        family = "u godan"
        romajiRoot = romaji[:-1]

    if currentType[0] == "V" and currentType != "VC":
        wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = family
        wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = combinedMeanings
        wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = currentType[-1]
        wsLocalVerbs.cell(row=verbSheetIndex, column=7).value = romaji
        wsLocalVerbs.cell(row=verbSheetIndex, column=6).value = kanji
        wsLocalVerbs.cell(row=verbSheetIndex, column=8).value = kanji[:-1]
        wsLocalVerbs.cell(row=verbSheetIndex, column=12).value = meanings
        wsLocalVerbs.cell(row=verbSheetIndex, column=11).value = altSpellings
        if currentType[-1] == "T": wsLocalVerbs.cell(row=verbSheetIndex, column=4).value = "を"

        wsLocalTypes.delete_rows(typesSheetIndex, 1)

        verbSheetIndex += 1
    typesSheetIndex -= 1

wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = "-"
wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = "-"
wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = "-"
#endregion


#region Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with firebase results.xlsx')
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with firebase results.xlsx')
#endregion