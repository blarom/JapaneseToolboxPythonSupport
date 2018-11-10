# Japanese Toolbox database creator

import openpyxl
import re

# region Getting the Excel sheets and their sizes
# Preparing the excel sheet for writing
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with firebase results.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with firebase results.xlsx', data_only=True)

# assign_sheet=wb.active
wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]




# Getting the size of the Local Types list
lastLocalTypesIndex = 1
while True:
    value = wsLocalTypes.cell(row=lastLocalTypesIndex, column=3).value
    if not value:
        lastLocalTypesIndex -= 1
        break
    lastLocalTypesIndex += 1

# Getting the size of the Local Verbs list
lastLocalVerbsIndex = 1
while True:
    value = wsLocalVerbs.cell(row=lastLocalVerbsIndex, column=3).value
    value2 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 1, column=3).value
    value3 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 2, column=3).value
    if (not value) & (not value2) & (not value3):
        lastLocalVerbsIndex -= 1
        break
    lastLocalVerbsIndex += 1
# endregion

# region Removing unused meaning indexes in the types sheet
typesSheetIndex = 2
while not wsLocalTypes.cell(row=typesSheetIndex, column=4).value == "" and wsLocalTypes.cell(row=typesSheetIndex, column=4).value is not None:
    meanings = str(wsLocalTypes.cell(row=typesSheetIndex, column=5).value)
    meaningIndexStrings = meanings.split(";")

    for index in range(len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        currentMeaning = str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value)

        if currentMeaning == 'None':
            meanings = meanings.replace(str(meaningIndexValue) + ";", "")
            meanings = meanings.replace(";" + str(meaningIndexValue), "")

    wsLocalTypes.cell(row=typesSheetIndex, column=5).value = meanings
    typesSheetIndex = typesSheetIndex + 1
# endregion

# region Moving the suru verb results from the Grammar/Types to the Verbs/Verbs sheet
typesSheetIndex = lastLocalTypesIndex
verbSheetIndex = lastLocalVerbsIndex
while " suru" in wsLocalTypes.cell(row=typesSheetIndex, column=3).value:
    romaji = str(wsLocalTypes.cell(row=typesSheetIndex, column=3).value)
    kanji = str(wsLocalTypes.cell(row=typesSheetIndex, column=4).value)
    meanings = str(wsLocalTypes.cell(row=typesSheetIndex, column=5).value)
    altSpellings = str(wsLocalTypes.cell(row=typesSheetIndex, column=6).value)
    if not altSpellings or altSpellings == "None": altSpellings = ""

    wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = "suru verb"
    wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = "I"
    wsLocalVerbs.cell(row=verbSheetIndex, column=7).value = romaji
    wsLocalVerbs.cell(row=verbSheetIndex, column=6).value = kanji
    wsLocalVerbs.cell(row=verbSheetIndex, column=12).value = meanings
    wsLocalVerbs.cell(row=verbSheetIndex, column=11).value = altSpellings
    wsLocalVerbs.cell(row=verbSheetIndex, column=9).value = romaji[:-4]
    wsLocalVerbs.cell(row=verbSheetIndex, column=8).value = kanji[:-2]
    wsLocalVerbs.cell(row=verbSheetIndex, column=13).value = kanji[:-2]
    wsLocalVerbs.cell(row=verbSheetIndex, column=10).value = 68

    meaningIndexStrings = meanings.split(";")
    combinedMeanings = ""
    currentType = "I"
    for index in range(len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        currentMeaning = str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value).strip().replace("\u200b", "")
        currentType = str(wsLocalMeanings.cell(row=meaningIndexValue, column=3).value).strip().replace("\u200b", "")

        # Adding the meaning to the lot
        if index > 0: combinedMeanings += ", "
        combinedMeanings += currentMeaning

    wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = combinedMeanings
    wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = currentType[-1]
    if currentType[-1] == "T": wsLocalVerbs.cell(row=verbSheetIndex, column=4).value = "を"

    wsLocalTypes.delete_rows(typesSheetIndex, 1)

    typesSheetIndex -= 1
    verbSheetIndex += 1

# wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = "-"
# wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = "-"
# wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = "-"
# endregion

# region Moving the regular verbs from the Grammar/Types to the Verbs/Verbs sheet and removing unused meaning indexes
while not wsLocalTypes.cell(row=typesSheetIndex, column=2).value:

    romaji = str(wsLocalTypes.cell(row=typesSheetIndex, column=3).value)
    kanji = str(wsLocalTypes.cell(row=typesSheetIndex, column=4).value)
    meanings = str(wsLocalTypes.cell(row=typesSheetIndex, column=5).value)
    altSpellings = str(wsLocalTypes.cell(row=typesSheetIndex, column=6).value)
    if not altSpellings or altSpellings == "None": altSpellings = ""

    meaningIndexStrings = meanings.split(";")
    combinedMeanings = ""
    currentType = "I"
    for index in range(len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        currentMeaning = str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value).strip().replace("\u200b", "")
        currentType = str(wsLocalMeanings.cell(row=meaningIndexValue, column=3).value).strip().replace("\u200b", "")

        # Adding the meaning to the lot
        if index > 0: combinedMeanings += ", "
        combinedMeanings += currentMeaning

    # If indexes were removed from the meanings, then update the value in the sheet
    wsLocalTypes.cell(row=typesSheetIndex, column=5).value = meanings

    family = ""
    romajiRoot = ""
    if kanji[-1] == "す":
        family = "su godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "く":
        family = "ku godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "ぐ":
        family = "gu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "ぶ":
        family = "bu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "む":
        family = "mu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "ぬ":
        family = "nu godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "る" and "rug" in currentType:
        family = "ru godan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "る" and "rui" in currentType:
        family = "ru ichidan"
        romajiRoot = romaji[:-2]
    elif kanji[-1] == "つ":
        family = "tsu godan"
        romajiRoot = romaji[:-3]
    elif kanji[-1] == "う":
        family = "u godan"
        romajiRoot = romaji[:-1]

    if currentType[0] == "V" and currentType != "VC":
        wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = family
        wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = combinedMeanings
        wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = currentType[-1]
        wsLocalVerbs.cell(row=verbSheetIndex, column=7).value = romaji
        wsLocalVerbs.cell(row=verbSheetIndex, column=6).value = kanji
        wsLocalVerbs.cell(row=verbSheetIndex, column=8).value = kanji[:-1]
        wsLocalVerbs.cell(row=verbSheetIndex, column=9).value = romajiRoot
        wsLocalVerbs.cell(row=verbSheetIndex, column=12).value = meanings
        wsLocalVerbs.cell(row=verbSheetIndex, column=11).value = altSpellings
        if currentType[-1] == "T": wsLocalVerbs.cell(row=verbSheetIndex, column=4).value = "を"

        wsLocalTypes.delete_rows(typesSheetIndex, 1)

        verbSheetIndex += 1
    typesSheetIndex -= 1

wsLocalVerbs.cell(row=verbSheetIndex, column=1).value = "-"
wsLocalVerbs.cell(row=verbSheetIndex, column=2).value = "-"
wsLocalVerbs.cell(row=verbSheetIndex, column=3).value = "-"
# endregion

# region Minimizing older meanings in the types sheet
typesSheetIndex = 2
while (not wsLocalTypes.cell(row=typesSheetIndex, column=4).value == "") and (wsLocalTypes.cell(row=typesSheetIndex, column=4).value is not None):

    meanings = str(wsLocalTypes.cell(row=typesSheetIndex, column=5).value)
    if meanings == 'None' or meanings == '' or meanings is None:
        typesSheetIndex += 1
        continue

    meaningIndexStrings = meanings.split(";")
    olderMeaningIndexValue = int(meaningIndexStrings[0])
    olderMeaningSource = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=9).value)
    if olderMeaningSource == 'J':
        typesSheetIndex += 1
        continue

    meaningType = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=3).value)
    if meaningType == 'PC':
        typesSheetIndex += 1
        continue

    olderMeaning = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=2).value)
    olderMeaningElements = re.split(r',\s*(?![^()]*\))', olderMeaning)
    olderMeaningElementsWithCancellations = list.copy(olderMeaningElements)
    olderMeaningElementsWithCancellations = list(map(str.strip, olderMeaningElementsWithCancellations))

    newMeanings = []
    for index in range(1, len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        newMeanings.append(str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value))

    newMeaningsAsSingleString = ', '.join(newMeanings)
    newMeaningElements = re.split(r',\s*(?![^()]*\))', newMeaningsAsSingleString)
    newMeaningElements = list(map(str.strip, newMeaningElements))


    for i in range(len(olderMeaningElements)):
        for newMeaningElement in newMeaningElements:
            if olderMeaningElementsWithCancellations[i] in newMeaningElement:
                olderMeaningElementsWithCancellations[i] = ''
                break

    finalOlderMeaningElements = []
    for element in olderMeaningElementsWithCancellations:
        if element != '':
            finalOlderMeaningElements.append(element)

    finalOlderMeaningElementsAsString = ', '.join(finalOlderMeaningElements)
    wsLocalMeanings.cell(row=olderMeaningIndexValue, column=2).value = finalOlderMeaningElementsAsString
    if finalOlderMeaningElementsAsString == '':
        wsLocalMeanings.cell(row=olderMeaningIndexValue, column=3).value = '-'
        meanings = meanings.replace(meaningIndexStrings[0] + ";", "")
        wsLocalTypes.cell(row=typesSheetIndex, column=5).value = meanings

        # Transferring the details from the older meaning being cancelled to the first available new meaning (ie. skipping the older ones), unless it already has details
        explanations = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=4).value)
        rules = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=5).value)
        example = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=6).value)
        opposite = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=7).value)
        synonym = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=8).value)

        if ((explanations == '' or explanations is None or explanations == 'None')
                and (rules == '' or rules is None or rules == 'None')
                and (example == '' or example is None or example == 'None')
                and (opposite == '' or opposite is None or opposite == 'None')
                and (synonym == '' or synonym is None or synonym == 'None')):
            pass
        else:
            for i in range(1, len(meaningIndexStrings)):
                # If the next meaning index is consecutive to the older meaning, then it's also an old meaning, so skip it
                if int(meaningIndexStrings[i]) == int(meaningIndexStrings[0])+i:
                    continue
                else:
                    explanationsNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=4).value)
                    rulesNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=5).value)
                    exampleNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=6).value)
                    oppositeNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=7).value)
                    synonymNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=8).value)
                    if ((explanationsNew == '' or explanationsNew is None or explanationsNew == 'None')
                            and (rulesNew == '' or rulesNew is None or rulesNew == 'None')
                            and (exampleNew == '' or exampleNew is None or exampleNew == 'None')
                            and (oppositeNew == '' or oppositeNew is None or oppositeNew == 'None')
                            and (synonymNew == '' or synonymNew is None or synonymNew == 'None')):
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=4).value = explanationsNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=5).value = rulesNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=6).value = exampleNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=7).value = oppositeNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=8).value = synonymNew
                    break

    typesSheetIndex += 1
# endregion

# region Minimizing older meanings in the verbs sheet
verbsSheetIndex = 2
while wsLocalVerbs.cell(row=verbsSheetIndex, column=1).value != "-"\
        and wsLocalVerbs.cell(row=verbsSheetIndex, column=2).value != "-"\
        and wsLocalVerbs.cell(row=verbsSheetIndex, column=3).value != "-":

    meanings = str(wsLocalVerbs.cell(row=verbsSheetIndex, column=12).value)
    if meanings == 'None' or meanings == '' or meanings is None:
        verbsSheetIndex += 1
        continue

    meaningIndexStrings = meanings.split(";")
    olderMeaningIndexValue = int(meaningIndexStrings[0])
    source = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=9).value)
    if source == 'J':
        verbsSheetIndex += 1
        continue

    olderMeaning = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=2).value)
    olderMeaningElements = re.split(r',\s*(?![^()]*\))', olderMeaning)
    olderMeaningElementsWithCancellations = list.copy(olderMeaningElements)
    olderMeaningElementsWithCancellations = list(map(str.strip, olderMeaningElementsWithCancellations))

    newMeanings = []
    for index in range(1, len(meaningIndexStrings)):
        meaningIndexValue = int(meaningIndexStrings[index])
        newMeanings.append(str(wsLocalMeanings.cell(row=meaningIndexValue, column=2).value))

    newMeaningsAsSingleString = ', '.join(newMeanings)
    newMeaningElements = re.split(r',\s*(?![^()]*\))', newMeaningsAsSingleString)
    newMeaningElements = list(map(str.strip, newMeaningElements))


    for i in range(len(olderMeaningElements)):
        for newMeaningElement in newMeaningElements:
            if olderMeaningElementsWithCancellations[i] in newMeaningElement:
                olderMeaningElementsWithCancellations[i] = ''
                break

    finalOlderMeaningElements = []
    for element in olderMeaningElementsWithCancellations:
        if element != '':
            finalOlderMeaningElements.append(element)

    finalOlderMeaningElementsAsString = ', '.join(finalOlderMeaningElements)
    wsLocalMeanings.cell(row=olderMeaningIndexValue, column=2).value = finalOlderMeaningElementsAsString
    if finalOlderMeaningElementsAsString == '':
        wsLocalMeanings.cell(row=olderMeaningIndexValue, column=3).value = '-'
        meanings = meanings.replace(meaningIndexStrings[0] + ";", "")
        wsLocalVerbs.cell(row=verbsSheetIndex, column=12).value = meanings

        # Transferring the details from the older meaning being cancelled to the first available new meaning (ie. skipping the older ones), unless it already has details
        explanations = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=4).value)
        rules = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=5).value)
        example = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=6).value)
        opposite = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=7).value)
        synonym = str(wsLocalMeanings.cell(row=olderMeaningIndexValue, column=8).value)

        if ((explanations == '' or explanations is None or explanations == 'None')
                and (rules == '' or rules is None or rules == 'None')
                and (example == '' or example is None or example == 'None')
                and (opposite == '' or opposite is None or opposite == 'None')
                and (synonym == '' or synonym is None or synonym == 'None')):
            pass
        else:
            for i in range(1, len(meaningIndexStrings)):
                # If the next meaning index is consecutive to the older meaning, then it's also an old meaning, so skip it
                if int(meaningIndexStrings[i]) == int(meaningIndexStrings[0])+i:
                    continue
                else:
                    explanationsNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=4).value)
                    rulesNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=5).value)
                    exampleNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=6).value)
                    oppositeNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=7).value)
                    synonymNew = str(wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=8).value)
                    if ((explanationsNew == '' or explanationsNew is None or explanationsNew == 'None')
                            and (rulesNew == '' or rulesNew is None or rulesNew == 'None')
                            and (exampleNew == '' or exampleNew is None or exampleNew == 'None')
                            and (oppositeNew == '' or oppositeNew is None or oppositeNew == 'None')
                            and (synonymNew == '' or synonymNew is None or synonymNew == 'None')):
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=4).value = explanationsNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=5).value = rulesNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=6).value = exampleNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=7).value = oppositeNew
                        wsLocalMeanings.cell(row=int(meaningIndexStrings[i]), column=8).value = synonymNew
                    break

    verbsSheetIndex += 1
# endregion

# region Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with firebase results.xlsx')
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with firebase results.xlsx')
# endregion
