# Japanese Toolbox database creator
import Globals
import pyrebase
import json
import openpyxl
import re
from keys import apiKey
from keys import serviceAccount

# region Initializing the Database
config = {
    "apiKey": apiKey,
    "authDomain": "japanese-toolbox.firebaseapp.com",
    "databaseURL": "https://japanese-toolbox.firebaseio.com",
    "storageBucket": "japanese-toolbox.appspot.com",
    "serviceAccount": serviceAccount
}
firebase = pyrebase.initialize_app(config)
db = firebase.database()

jishoWordsListFromFirebase = db.child("wordsList").get()


# Defining the word class that maps the relevant json input to values in the class
class Word(object):
    romaji = ""
    kanji = ""
    altSpellings = ""
    isCommon = False
    meaningsEN = []
    meaningsFR = []
    meaningsES = []

    def __init__(self, json_content):
        data = json.loads(json_content)
        for key, value in data.items():
            self.__dict__[key] = value

    class Meaning(object):
        def __init__(self, json_content):
            data = json.loads(json_content)
            for key, value in data.items():
                self.__dict__[key] = value

        class Explanation(object):
            def __init__(self, json_content):
                data = json.loads(json_content)
                for key, value in data.items():
                    self.__dict__[key] = value


# endregion

# region Getting the Excel sheets and their sizes
# Preparing the excel sheet for writing
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - before foreign.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - before foreign.xlsx', data_only=True)

# assign_sheet=wb.active
wsLocalMeaningsEN = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]

# Getting the size of the Local Types list
lastLocalTypesIndex = 1
while True:
    value = wsLocalTypes.cell(row=lastLocalTypesIndex, column=Globals.TYPES_COL_ROMAJI).value
    if not value:
        lastLocalTypesIndex -= 1
        break
    lastLocalTypesIndex += 1

# Getting the size of the Local Grammar list
lastLocalGrammarIndex = 1
while True:
    value = wsLocalGrammar.cell(row=lastLocalGrammarIndex, column=Globals.TYPES_COL_ROMAJI).value
    if not value:
        lastLocalGrammarIndex -= 1
        break
    lastLocalGrammarIndex += 1

# Getting the size of the Local Meanings list
lastLocalMeaningsIndex = 1
while True:
    value = wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex, column=Globals.MEANINGS_COL_INDEX).value
    if not value:
        lastLocalMeaningsIndex -= 1
        break
    lastLocalMeaningsIndex += 1

# Getting the size of the Local Verbs list
lastLocalVerbsIndex = 1
while True:
    value = wsLocalVerbs.cell(row=lastLocalVerbsIndex, column=Globals.VERBS_COL_KANJI).value
    value2 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 1, column=Globals.VERBS_COL_KANJI).value
    value3 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 2, column=Globals.VERBS_COL_KANJI).value
    if (not value) & (not value2) & (not value3):
        lastLocalVerbsIndex -= 1
        break
    lastLocalVerbsIndex += 1

# endregion

# region Defining the helper classes & methods for the database integration

JISHO_TO_JT_LEGEND = {
    "Adverb": "A",
    "Abbreviation": "Abr",
    "Noun or verb acting prenominally": "Af",
    "I-adjective": "Ai",
    "I-adjective (yoi/ii class)": "Ai",
    "Kari adjective (archaic)": "Akari;ar",
    "Ku-adjective (archaic)": "Aku;ar",
    "Shiku adjective (archaic)": "Ashiku;ar",
    "Adjective": "Aj",
    "Na-adjective": "Ana",
    "No-adjective": "Ano",
    "Pre-noun adjectival": "Apn",
    "Adverb taking the 'to' particle": "Ato",
    "Taru-adjective": "Atr",
    "Archaism": "ar",
    "Auxiliary adjective": "Ax",
    "Counter": "C",
    "Expression": "CE",
    "Conjunction": "CO",
    "Food term": "Cu",
    "Derogatory": "Dr",
    "Proverb": "idp",
    "Idiomatic expression": "idp",
    "Rude or X-rated term": "IES",
    "Colloquialism": "LF",
    "Familiar language": "LF",
    "Female term or language": "LFt",
    "Humble (kenjougo)": "LHm",
    "Honorific or respectful (sonkeigo)": "LHn",
    "Polite (teineigo)": "LHn",
    "Male term or language": "LMt",
    "Music term": "MAC",
    "Medicine, etc. term": "Md",
    "Noun": "N",
    "Adverbial noun": "NAdv",
    "Proper noun": "Ne",
    "Family or surname": "Ne",
    "Male given name": "Ne",
    "Female given name": "Ne",
    "Full name": "Ne",
    "Given name, gender not specified": "Ne",
    "Numeric": "num",
    "Out-dated kanji": "Obs",
    "Obsolete term": "Obs",
    "Obscure term": "Obs",
    "Old or irregular kana form": "Obs",
    "Out-dated or obsolete kana usage": "Obs",
    "Interjection": "OI",
    "Onomatopoeic or mimetic word": "OI",
    "Pronoun": "P",
    "Place": "Pl",
    "Particle": "PP",
    "Prefix": "Px",
    "Noun - used as a prefix": "N;Px",
    "Noun - used as a suffix": "N;Sx",
    "Male slang": "Sl",
    "Slang": "Sl", "Suffix": "Sx",
    "Temporal noun": "N;T",
    "Manga slang": "ZMg",
    "Mahjong term": "ZMj",
    "Military": "ZMl",
    "Mathematics": "ZMt",
    "Poetical term": "ZP",
    "Physics terminology": "ZPh",
    "Shogi term": "ZSg",
    "Sumo term": "ZSm",
    "Sports term": "ZSp",
    "Shinto term": "ZSt",
    "Zoology term": "ZZ",
    "Auxiliary verb": "Vx",
    "Architecture term": "ZAc",
    "Anatomical term": "ZAn",
    "Astronomy, etc. term": "ZAs",
    "Buddhist term": "ZB",
    "Baseball term": "ZBb",
    "Biology term": "ZBi",
    "Business term": "ZBs",
    "Botany term": "ZBt",
    "Chemistry term": "ZC",
    "Children's language": "ZCL",
    "Economics term": "ZEc",
    "Engineering term": "ZEg",
    "Finance term": "ZFn",
    "Geometry term": "ZG",
    "Geology, etc. term": "ZGg",
    "Jocular, humorous term": "ZH",
    "Computer terminology": "ZI",
    "linguistics terminology": "ZL",
    "Law, etc. term": "ZLw",
    "Martial arts term": "ZM",
    "Vulgar": "vul",
    "Godan verb - aru special class": "VaruI",
    "Godan verb - aru special class, intransitive verb": "VaruI",
    "Godan verb - aru special class, Transitive verb": "VaruT",
    "Godan verb with bu ending": "VbuI",
    "Godan verb with bu ending, intransitive verb": "VbuI",
    "Godan verb with bu ending, Transitive verb": "VbuT",
    "Godan verb with gu ending": "VguI",
    "Godan verb with gu ending, intransitive verb": "VguI",
    "Godan verb with gu ending, Transitive verb": "VguT",
    "Godan verb - Iku/Yuku special class": "VikuI",
    "Godan verb - Iku/Yuku special class, intransitive verb": "VikuI",
    "Godan verb - Iku/Yuku special class, Transitive verb": "VikuT",
    "Godan verb with ku ending": "VkuI",
    "Godan verb with ku ending, intransitive verb": "VkuI",
    "Godan verb with ku ending, Transitive verb": "VkuT",
    "Godan verb with mu ending": "VmuI",
    "Godan verb with mu ending, intransitive verb": "VmuI",
    "Godan verb with mu ending, Transitive verb": "VmuT",
    "Godan verb with nu ending": "VnuI",
    "Godan verb with nu ending, intransitive verb": "VnuI",
    "Godan verb with nu ending, Transitive verb": "VnuT",
    "Godan verb with ru ending": "VrugI",
    "Godan verb with ru ending, intransitive verb": "VrugI",
    "Godan verb with ru ending, Transitive verb": "VrugT",
    "Godan verb with ru ending (irregular verb)": "VrugI",
    "Godan verb with ru ending (irregular verb), intransitive verb": "VrugI",
    "Godan verb with ru ending (irregular verb), Transitive verb": "VrugT",
    "Ichidan verb": "VruiI",
    "Ichidan verb, intransitive verb": "VruiI",
    "Ichidan verb, Transitive verb": "VruiT",
    "Godan verb with su ending": "VsuI",
    "Godan verb with su ending, intransitive verb": "VsuI",
    "Godan verb with su ending, Transitive verb": "VsuT",
    "Godan verb with tsu ending": "VtsuI",
    "Godan verb with tsu ending, intransitive verb": "VtsuI",
    "Godan verb with tsu ending, Transitive verb": "VtsuT",
    "Godan verb with u ending (special class)": "VusI",
    "Godan verb with u ending (special class), intransitive verb": "VusI",
    "Godan verb with u ending (special class), Transitive verb": "VusT",
    "Godan verb with u ending": "VuI",
    "Godan verb with u ending, intransitive verb": "VuI",
    "Godan verb with u ending, Transitive verb": "VuT",
    "Suru verb": "VsuruI",
    "Suru verb, intransitive verb": "VsuruI",
    "Suru verb, Transitive verb": "VsuruT",
    "Suru verb - irregular": "VsuruI",
    "Suru verb - irregular, intransitive verb": "VsuruI",
    "Suru verb - irregular, Transitive verb": "VsuruT",
    "Suru verb - special class": "VsuruI",
    "Suru verb - special class, intransitive verb": "VsuruI",
    "Suru verb - special class, Transitive verb": "VsuruT",
    "Kuru verb - special class": "VkuruI",
    "Kuru verb - special class, intransitive verb": "VkuruI",
    "Kuru verb - special class, Transitive verb": "VkuruT"

}

LEGEND_SORTED_KEYS = list(JISHO_TO_JT_LEGEND.keys())
LEGEND_SORTED_KEYS.sort(key=len, reverse=True)

LEGEND_SORTED_VALUES = list(JISHO_TO_JT_LEGEND.values())
LEGEND_SORTED_VALUES.sort(key=len, reverse=True)

def getJTType(jishoType):
    if jishoType == "Adverb":
        return "A"
    elif jishoType == "Noun":
        return "N"
    elif jishoType == "Place":
        return "Pl"
    elif jishoType == "Temporal Noun":
        return "T"
    elif jishoType == "Proper Noun":
        return "Ne"
    elif jishoType == "Numeric":
        return "num"
    elif "Suru verb" in jishoType:
        if "Transitive" in jishoType:
            return "VsuruT"
        else:
            return "VsuruI"
    elif "Kuru verb" in jishoType:
        if "Transitive" in jishoType:
            return "VkuruT"
        else:
            return "VkuruI"
    elif "Ichidan verb" in jishoType:
        if "Transitive" in jishoType:
            return "VruiT"
        else:
            return "VruiI"
    elif "Godan verb with ru" in jishoType:
        if "Transitive" in jishoType:
            return "VruT"
        else:
            return "VrugI"
    elif "Godan verb with bu" in jishoType:
        if "Transitive" in jishoType:
            return "VbuT"
        else:
            return "VbuI"
    elif "Godan verb with gu" in jishoType:
        if "Transitive" in jishoType:
            return "VguT"
        else:
            return "VguI"
    elif "Godan verb with ku" in jishoType:
        if "Transitive" in jishoType:
            return "VkuT"
        else:
            return "VkuI"
    elif "Godan verb - Iku/Yuku" in jishoType:
        if "Transitive" in jishoType:
            return "VikuT"
        else:
            return "VikuI"
    elif "Godan verb - aru" in jishoType:
        if "Transitive" in jishoType:
            return "VaruT"
        else:
            return "VaruI"
    elif "Godan verb with u ending (special class)" in jishoType:
        if "Transitive" in jishoType:
            return "VusT"
        else:
            return "VusI"
    elif "Godan verb with mu" in jishoType:
        if "Transitive" in jishoType:
            return "VmuT"
        else:
            return "VmuI"
    elif "Godan verb with nu" in jishoType:
        if "Transitive" in jishoType:
            return "VnuT"
        else:
            return "VnuI"
    elif "Godan verb with su" in jishoType:
        if "Transitive" in jishoType:
            return "VsuT"
        else:
            return "VsuI"
    elif "Godan verb with tsu" in jishoType:
        if "Transitive" in jishoType:
            return "VtsuT"
        else:
            return "VtsuI"
    elif "Godan verb with u" in jishoType:
        if "Transitive" in jishoType:
            return "VuT"
        else:
            return "VuI"
    elif "Suffix, Counter" in jishoType:
        return "C"
    elif ("Suffix" in jishoType) | ("suffix" in jishoType):
        return "Sx"
    elif ("Prefix" in jishoType) | ("prefix" in jishoType):
        return "Px"
    elif ("I-adjective" in jishoType) | ("i-adjective" in jishoType):
        return "Ai"
    elif ("Na-adjective" in jishoType) | ("na-adjective" in jishoType):
        return "Ana"
    elif "adjective" in jishoType:
        return "Aj"
    elif "Pre-noun adjectival" in jishoType or "Pronoun" in jishoType:
        return "P"
    elif "Expression" in jishoType:
        return "CE"
    elif ("Auxiliary verb" in jishoType) | ("Auxiliary adjective" in jishoType):
        return ""
    elif ("Conjunction" in jishoType) | ("Auxiliary" in jishoType) | ("Particle" in jishoType):
        return "PP"
    elif "Counter" in jishoType:
        return "C"

    elif jishoType == "A":
        return jishoType
    elif jishoType == "ADc":
        return jishoType
    elif jishoType == "ADg":
        return jishoType
    elif jishoType == "Ai":
        return jishoType
    elif jishoType == "Aj":
        return jishoType
    elif jishoType == "AM":
        return jishoType
    elif jishoType == "Ana":
        return jishoType
    elif jishoType == "AO":
        return jishoType
    elif jishoType == "AP":
        return jishoType
    elif jishoType == "AT":
        return jishoType
    elif jishoType == "C":
        return jishoType
    elif jishoType == "CE":
        return jishoType
    elif jishoType == "iAC":
        return jishoType
    elif jishoType == "IES":
        return jishoType
    elif jishoType == "naAC":
        return jishoType
    elif jishoType == "Ac":
        return jishoType
    elif jishoType == "Abs":
        return jishoType
    elif jishoType == "An":
        return jishoType
    elif jishoType == "At":
        return jishoType
    elif jishoType == "B":
        return jishoType
    elif jishoType == "Col":
        return jishoType
    elif jishoType == "Cu":
        return jishoType
    elif jishoType == "DM":
        return jishoType
    elif jishoType == "DW":
        return jishoType
    elif jishoType == "Fa":
        return jishoType
    elif jishoType == "Fl":
        return jishoType
    elif jishoType == "Fy":
        return jishoType
    elif jishoType == "GO":
        return jishoType
    elif jishoType == "JEP":
        return jishoType
    elif jishoType == "MAC":
        return jishoType
    elif jishoType == "Md":
        return jishoType
    elif jishoType == "Mo":
        return jishoType
    elif jishoType == "MSE":
        return jishoType
    elif jishoType == "N":
        return jishoType
    elif jishoType == "Ne":
        return jishoType
    elif jishoType == "NE":
        return jishoType
    elif jishoType == "Nn":
        return jishoType
    elif jishoType == "org":
        return jishoType
    elif jishoType == "Pe":
        return jishoType
    elif jishoType == "Pl":
        return jishoType
    elif jishoType == "Sp":
        return jishoType
    elif jishoType == "T":
        return jishoType
    elif jishoType == "V":
        return jishoType
    elif jishoType == "NW":
        return jishoType
    elif jishoType == "NY":
        return jishoType
    elif jishoType == "OI":
        return jishoType
    elif jishoType == "P":
        return jishoType
    elif jishoType == "PP":
        return jishoType
    elif jishoType == "Px":
        return jishoType
    elif jishoType == "Sa":
        return jishoType
    elif jishoType == "SI":
        return jishoType
    elif jishoType == "CO":
        return jishoType
    elif jishoType == "PC":
        return jishoType
    elif jishoType == "Sx":
        return jishoType
    elif jishoType == "UNC":
        return jishoType
    elif jishoType == "VbuI":
        return jishoType
    elif jishoType == "VbuT":
        return jishoType
    elif jishoType == "VC":
        return jishoType
    elif jishoType == "VdaI":
        return jishoType
    elif jishoType == "VdaT":
        return jishoType
    elif jishoType == "VguI":
        return jishoType
    elif jishoType == "VguT":
        return jishoType
    elif jishoType == "VkuI":
        return jishoType
    elif jishoType == "VkuruI":
        return jishoType
    elif jishoType == "VkuruT":
        return jishoType
    elif jishoType == "VkuT":
        return jishoType
    elif jishoType == "VmuI":
        return jishoType
    elif jishoType == "VmuT":
        return jishoType
    elif jishoType == "VnuI":
        return jishoType
    elif jishoType == "VnuT":
        return jishoType
    elif jishoType == "VrugI":
        return jishoType
    elif jishoType == "VrugT":
        return jishoType
    elif jishoType == "VruiI":
        return jishoType
    elif jishoType == "VruiT":
        return jishoType
    elif jishoType == "VsuI":
        return jishoType
    elif jishoType == "VsuruI":
        return jishoType
    elif jishoType == "VsuruT":
        return jishoType
    elif jishoType == "VsuT":
        return jishoType
    elif jishoType == "VtsuI":
        return jishoType
    elif jishoType == "VtsuT":
        return jishoType
    elif jishoType == "VuI":
        return jishoType
    elif jishoType == "VuT":
        return jishoType
    elif jishoType == "VikuI":
        return jishoType
    elif jishoType == "VikuT":
        return jishoType
    elif jishoType == "VaruI":
        return jishoType
    elif jishoType == "VaruT":
        return jishoType
    elif jishoType == "VusI":
        return jishoType
    elif jishoType == "VusT":
        return jishoType

    elif jishoType == "Invalid":
        return ""
    else:
        return "UNC"


class JishoWord(object):
    romaji = ""
    kanji = ""
    fixedJishoWordAltSpellings = ""
    meanings = []
    common = 0

    def __init__(self, romaji, kanji, altSpellings, meanings, common):
        self.romaji = romaji
        self.kanji = kanji
        self.altSpellings = altSpellings
        self.meanings = meanings
        self.common = 1 if common else 0

    class Meaning(object):
        meaning = ""
        grammarType = ""
        grammarTypeForSheets = ""

        def __init__(self, meaning, grammarType, grammarTypeForSheets):
            self.meaning = meaning
            self.grammarType = grammarType
            self.grammarTypeForSheets = grammarTypeForSheets


# endregion

# region Creating the JishoWord list to be used in the integrator

jishoWords = []
for wordKey in jishoWordsListFromFirebase.val():

    # Getting the Jisho word's characteristics
    word = jishoWordsListFromFirebase.val()[wordKey]
    wordObject = Word(json.dumps(word))

    meanings = []
    wordMeanings = []
    try:
        wordMeanings = wordObject.meaningsEN
    except:
        try:
            wordMeanings = wordObject.meanings
        except:
            continue

    for jishoWordMeaningString in wordMeanings:
        meaningObject = Word.Meaning(json.dumps(jishoWordMeaningString))
        grammarType = meaningObject.type
        meaning = meaningObject.meaning
        grammarTypeForSheets = "O"

        jishoMeaning = JishoWord.Meaning(meaning, grammarType, grammarTypeForSheets)
        meanings.append(jishoMeaning)

    # Creating the local word object
    if not [wordObject.romaji, wordObject.kanji] in Globals.EDICT_EXCEPTIONS and not ["*", wordObject.kanji] in Globals.EDICT_EXCEPTIONS:
        jishoWord = JishoWord(wordObject.romaji, wordObject.kanji, wordObject.altSpellings, meanings, wordObject.isCommon)
        jishoWords.append(jishoWord)
# endregion

# region Iterating on the JishoWord list to integrate the words in the local sheets
for jishoWord in jishoWords:

    # region Getting the Jisho word's characteristics
    jishoWordRomaji = jishoWord.romaji
    jishoWordKanji = jishoWord.kanji
    jishoWordAltSpellings = str(jishoWord.altSpellings)
    jishoWordCommon = jishoWord.common
    if jishoWordAltSpellings.strip() != '':
        jishoWordAltSpellings = ", ".join(set([element.strip() for element in jishoWordAltSpellings.split(",")]))
        suruJishoWordAltSpellings = ", ".join([element.strip() + "する" if not re.match("^[a-zA-Z]+$", element.strip()) else element.strip() + "suru" for element in jishoWordAltSpellings.split(",")])
    else:
        suruJishoWordAltSpellings = jishoWordAltSpellings.strip()

    jishoWordMeaningObjects = jishoWord.meanings
    jishoWordMeaningsOriginalList = []
    jishoWordMeaningTypesOriginalList = []
    jishoWordTypesForSheetsOriginalList = []
    for jishoWordMeaningObject in jishoWordMeaningObjects:
        jishoWordMeaningsOriginalList.append(jishoWordMeaningObject.meaning)
        jishoWordMeaningTypesOriginalList.append(jishoWordMeaningObject.grammarType)
        jishoWordTypesForSheetsOriginalList.append(jishoWordMeaningObject.grammarTypeForSheets)
    # endregion

    # region Cycling over the jishoType elements and creating the final jisho word's meanings array accordingly
    # Also: updating or creating the suru verbs as necessary

    # Skipping duplicate meanings while uniting their types
    jishoWordMeaningsCondensed = []
    jishoWordMeaningTypesCondensed = []
    jishoWordTypesForSheetsCondensed = []
    for jishoWordMeaningIndex in range(len(jishoWordMeaningsOriginalList)):

        current_meaning = jishoWordMeaningsOriginalList[jishoWordMeaningIndex]
        current_type = jishoWordMeaningTypesOriginalList[jishoWordMeaningIndex]
        current_typeForSheet = jishoWordTypesForSheetsOriginalList[jishoWordMeaningIndex]

        if current_meaning in jishoWordMeaningsCondensed:
            index = jishoWordMeaningsCondensed.index(current_meaning)
            jishoWordMeaningTypesCondensed[index] += ";" + current_type
            jishoWordTypesForSheetsCondensed[index] += ";" + current_typeForSheet
        else:
            jishoWordMeaningsCondensed.append(current_meaning)
            jishoWordMeaningTypesCondensed.append(current_type)
            jishoWordTypesForSheetsCondensed.append(current_typeForSheet)


    # Looping through the meanings to create the meanings/types/typesForSheet lists, and updating altSpellings for suru verbs
    jishoWordMeanings = []
    jishoWordMeaningTypes = []
    jishoWordTypesForSheets = []
    for jishoWordMeaningIndex in range(len(jishoWordMeaningsCondensed)):

        # Getting the object meaning parameters
        jishoWordMeaningString = jishoWordMeaningsCondensed[jishoWordMeaningIndex]
        jishoTypesString = jishoWordMeaningTypesCondensed[jishoWordMeaningIndex]
        jishoWordTypeForSheet = jishoWordTypesForSheetsCondensed[jishoWordMeaningIndex]

        # Converting to only JTTypes while keeping the order
        typeElements = []
        for item in jishoTypesString.split(";"):
            if item in JISHO_TO_JT_LEGEND.keys():
                typeElements.append(JISHO_TO_JT_LEGEND[item])
            elif item in JISHO_TO_JT_LEGEND.values():
                typeElements.append(item)

        # Adding meanings if necessary
        typeElements_not_used_yet = typeElements
        if 'Ana' in typeElements:
            jishoWordMeanings.append(jishoWordMeaningString)
            jishoWordMeaningTypes.append('Ana')
            typeElements_not_used_yet.remove('Ana')
            jishoWordTypesForSheets.append(jishoWordTypeForSheet)

            if 'A' in typeElements:
                jishoWordMeanings.append(jishoWordMeaningString)
                jishoWordMeaningTypes.append('A')
                typeElements_not_used_yet.remove('A')
                jishoWordTypesForSheets.append(jishoWordTypeForSheet)

                if len(typeElements_not_used_yet) > 0:
                    jishoWordMeanings.append(jishoWordMeaningString)
                    jishoWordMeaningTypes.append(';'.join(typeElements_not_used_yet))
                    jishoWordTypesForSheets.append(jishoWordTypeForSheet)
            else:
                if len(typeElements_not_used_yet) > 0:
                    jishoWordMeanings.append(jishoWordMeaningString)
                    jishoWordMeaningTypes.append(';'.join(typeElements_not_used_yet))
                    jishoWordTypesForSheets.append(jishoWordTypeForSheet)

        elif 'A' in typeElements:
            jishoWordMeanings.append(jishoWordMeaningString)
            jishoWordMeaningTypes.append('A')
            typeElements_not_used_yet.remove('A')
            jishoWordTypesForSheets.append(jishoWordTypeForSheet)

            if len(typeElements_not_used_yet) > 0:
                jishoWordMeanings.append(jishoWordMeaningString)
                jishoWordMeaningTypes.append(';'.join(typeElements_not_used_yet))
                jishoWordTypesForSheets.append(jishoWordTypeForSheet)

        elif "V" in [item[0] for item in typeElements if item != 'VC']:
            for jTType in typeElements:
                if jTType == "": continue

                if ((jTType == "VsuruT") | (jTType == "VsuruI")) \
                        and (" suru" not in jishoWordRomaji) \
                        and ("ssuru" not in jishoWordRomaji):

                    newMeanings = [JishoWord.Meaning(jishoWordMeaningString, jTType, "V")]

                    # Check if the suru verb is already in the list of verbs, and if not then add it
                    tempVerbSheetIndex = lastLocalVerbsIndex - 1
                    isInSuruList = False

                    while "suru" in str(wsLocalVerbs.cell(row=tempVerbSheetIndex, column=Globals.VERBS_COL_ROMAJI).value):
                        currentLocalRomaji = str(wsLocalVerbs.cell(row=tempVerbSheetIndex, column=Globals.VERBS_COL_ROMAJI).value)
                        currentLocalKanji = str(wsLocalVerbs.cell(row=tempVerbSheetIndex, column=Globals.VERBS_COL_KANJI).value)
                        currentLocalAltSpellings = str(wsLocalVerbs.cell(row=tempVerbSheetIndex, column=Globals.VERBS_COL_ALTS).value)

                        if jishoWordKanji in currentLocalKanji and jishoWordRomaji in currentLocalRomaji:
                            isInSuruList = True
                            break

                        if currentLocalRomaji == "" or currentLocalRomaji == "None":
                            break

                        tempVerbSheetIndex -= 1

                    if not isInSuruList:
                        newJishoWord = JishoWord(jishoWordRomaji + " suru", jishoWordKanji + "する", suruJishoWordAltSpellings, newMeanings, jishoWordCommon)
                        jishoWords.append(newJishoWord)

                elif ((jTType == "VsuruT") | (jTType == "VsuruI")) & ("ssuru" not in jishoWordRomaji):
                    jishoWordTypesForSheets.append("V")
                    jishoWordMeaningTypes.append(jTType)
                    typeElements_not_used_yet.remove(jTType)
                    jishoWordMeanings.append(jishoWordMeaningString)
                elif (jTType[0] == "V") & ((jTType[-1] == "I") | (jTType[-1] == "T")):
                    # If the word is a verb, remove the "to "
                    jishoWordMeaningString = jishoWordMeaningString.replace(", to ", ", ")
                    jishoWordMeaningString = jishoWordMeaningString[3:]
                    jishoWordTypesForSheets.append("V")
                    typeElements_not_used_yet.remove(jTType)
                    jishoWordMeaningTypes.append(jTType)
                    jishoWordMeanings.append(jishoWordMeaningString)

            if len(typeElements_not_used_yet) > 0:
                jishoWordTypesForSheets.append("O")
                jishoWordMeaningTypes.append(";".join(typeElements_not_used_yet))
                jishoWordMeanings.append(jishoWordMeaningString)

        else:
            has_at_least_one_type = False
            for jTType in typeElements:
                if (jTType == "VC") | (jTType == "CO") | (jTType == "PC") | (jTType == "PP") | (jTType == "iAC") | (jTType == "naAC"):
                    jishoWordTypesForSheets.append("G")
                    jishoWordMeaningTypes.append(jTType)
                    typeElements_not_used_yet.remove(jTType)
                    jishoWordMeanings.append(jishoWordMeaningString)
                    has_at_least_one_type = True

            if len(typeElements_not_used_yet) > 0:
                jishoWordTypesForSheets.append("O")
                jishoWordMeaningTypes.append(";".join(typeElements_not_used_yet))
                jishoWordMeanings.append(jishoWordMeaningString)
            elif not has_at_least_one_type:
                jishoWordTypesForSheets.append("O")
                jishoWordMeaningTypes.append("CE")
                jishoWordMeanings.append(jishoWordMeaningString)

    # endregion

    wordAlreadyExists = False

    # region Finding the identical local entry in the Grammar sheet
    rowIndexInLocalWordSheet = 2
    while True:
        value = wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ROMAJI).value
        if not value:
            break

        currentLocalRomaji = str(wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ROMAJI).value)
        currentLocalKanji = str(wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_KANJI).value)
        currentLocalAltSpellings = str(wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ALTS).value)
        currentLocalCommon = str(wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_COMMON).value)
        currentLocalMeaningsIndexesEN = str(wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value)

        # Once the identical local entry is found, find the meanings that are unique to the Jisho word, and add them to the database
        if (currentLocalRomaji == jishoWordRomaji) and (currentLocalKanji == jishoWordKanji):

            wordAlreadyExists = True

            # Updating the altSpellings
            if currentLocalAltSpellings == "" or currentLocalAltSpellings == "None":
                wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ALTS).value = jishoWordAltSpellings
            else:
                currentLocalAltSpellingsList = [element.strip() for element in currentLocalAltSpellings.split(",")]
                for jishoWordAltSpelling in jishoWordAltSpellings.split(","):
                    if jishoWordAltSpelling not in currentLocalAltSpellings:
                        currentLocalAltSpellingsList.append(jishoWordAltSpelling)
                wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ALTS).value = ', '.join(currentLocalAltSpellingsList)

            # Updating the common status
            if currentLocalCommon == 0 or currentLocalCommon is None or currentLocalCommon == "":
                wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_COMMON).value = jishoWordCommon

            # Updating the meanings
            localMeaningIndexes = currentLocalMeaningsIndexesEN.split(";")
            localMeaningLoopIndex = 0
            while localMeaningLoopIndex < len(localMeaningIndexes):

                localMeaningIndex = localMeaningIndexes[localMeaningLoopIndex]

                # Getting the current local meaning
                rowIndexInLocalMeanings = int(localMeaningIndex)
                currentMeaning = str(wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_MEANING).value).strip().replace(
                    "\u200b", "")  # Fixes Zero Width Space bug

                # Checking if the Jisho meaning is already equal to one of the Local word's meanings
                foundMeaning = False
                jishoMeaningLoopIndex = 0
                for jishoWordMeaningString in jishoWordMeanings:

                    if jishoWordTypesForSheets[jishoMeaningLoopIndex] == "G":
                        if jishoWordMeaningString == currentMeaning:
                            foundMeaning = True
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "J"
                            break
                        elif jishoWordMeaningString in currentMeaning:
                            foundMeaning = True
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "LOC"
                            break
                        elif currentMeaning in jishoWordMeaningString:
                            foundMeaning = True
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "J"
                            break
                    jishoMeaningLoopIndex += 1

                if foundMeaning:
                    del jishoWordMeanings[jishoMeaningLoopIndex]
                    del jishoWordTypesForSheets[jishoMeaningLoopIndex]
                    del jishoWordMeaningTypes[jishoMeaningLoopIndex]

                localMeaningLoopIndex += 1

            # The leftover meanings are not present in the local database, so add them to it and update the Local meanings index list
            for i in range(len(jishoWordMeanings)):

                if jishoWordTypesForSheets[i] == "G":
                    jishoWordMeaningString = jishoWordMeanings[i]
                    jishoWordMeaningType = jishoWordMeaningTypes[i]

                    # Adding the results to the meanings excel sheet
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_INDEX).value = lastLocalMeaningsIndex + 1
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_TYPE).value = jishoWordMeaningType
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_SOURCE).value = "J"

                    # Updating the meaning indexes in the Local Grammar
                    wsLocalGrammar.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value \
                        = str(wsLocalGrammar.cell(
                        row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value) + ";" + str(lastLocalMeaningsIndex + 1)

                    # Incrementing the meaning index for the next iteration
                    lastLocalMeaningsIndex += 1

            # Once the meanings have been handled, break the while loop in order to process the next Jisho word
            break

        rowIndexInLocalWordSheet += 1
    # endregion

    # region Finding the identical local entry in the Verbs sheet
    rowIndexInLocalWordSheet = 4
    while True:
        value = wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ROMAJI).value
        value2 = wsLocalVerbs.cell(row=rowIndexInLocalWordSheet + 1, column=Globals.VERBS_COL_ROMAJI).value
        value3 = wsLocalVerbs.cell(row=rowIndexInLocalWordSheet + 2, column=Globals.VERBS_COL_ROMAJI).value
        if (not value) and (not value2) and (not value3):
            break

        currentLocalRomaji = str(wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ROMAJI).value)
        currentLocalKanji = str(wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_KANJI).value)
        currentLocalMeaningsIndexesEN = str(wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_MEANINGS_EN).value)
        currentLocalAltSpellings = str(wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ALTS).value)
        currentLocalCommon = str(wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_COMMON).value)

        # Once the identical local entry is found, find the meanings/altSpellings that are unique to the Jisho word, and add them to the database
        if ((currentLocalRomaji == jishoWordRomaji) or (currentLocalRomaji == (jishoWordRomaji + " suru"))) \
                and ((currentLocalKanji == jishoWordKanji) or (currentLocalKanji == (jishoWordKanji + "する"))):

            wordAlreadyExists = True
            #
            # if "mitsukaru" in currentLocalRomaji:
            #     print("currentLocalRomaji: " + currentLocalRomaji
            #           + "\njishoWordRomaji: " + jishoWordRomaji
            #           + "\ncurrentLocalKanji: " + currentLocalKanji
            #           + "\njishoWordKanji: " + jishoWordKanji
            #           + "\ncurrentLocalAltSpellings: " + str(currentLocalAltSpellings)
            #           + "\nsuruJishoWordAltSpellings: " + str(suruJishoWordAltSpellings))

            suruJishoWordAltSpellings = ', '.join([element.strip() for element in set(suruJishoWordAltSpellings.split(','))])

            # Updating the altSpellings
            currentLocalAltSpellingsList = [element.strip() for element in currentLocalAltSpellings.split(",")]
            if currentLocalRomaji == (jishoWordRomaji + " suru") or currentLocalKanji == (jishoWordKanji + "する"):
                if currentLocalAltSpellings == "" or currentLocalAltSpellings == "None":
                    wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ALTS).value = suruJishoWordAltSpellings
                else:
                    for jishoWordAltSpelling in suruJishoWordAltSpellings.split(","):
                        if jishoWordAltSpelling not in currentLocalAltSpellings:
                            currentLocalAltSpellingsList.append(jishoWordAltSpelling)
                    wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ALTS).value = ', '.join(currentLocalAltSpellingsList)
            else:
                if currentLocalAltSpellings == "" or currentLocalAltSpellings == "None":
                    wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ALTS).value = jishoWordAltSpellings
                else:
                    for jishoWordAltSpelling in jishoWordAltSpellings.split(","):
                        if jishoWordAltSpelling not in currentLocalAltSpellings:
                            currentLocalAltSpellingsList.append(jishoWordAltSpelling)
                    wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_ALTS).value = ', '.join(currentLocalAltSpellingsList)

            # Updating the common status
            if currentLocalCommon == 0 or currentLocalCommon is None or currentLocalCommon == "":
                wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_COMMON).value = jishoWordCommon

            # Updating the meanings
            localMeaningIndexes = currentLocalMeaningsIndexesEN.split(";")
            localMeaningLoopIndex = 0
            while localMeaningLoopIndex < len(localMeaningIndexes):

                localMeaningIndex = localMeaningIndexes[localMeaningLoopIndex]

                # Getting the current local meaning
                rowIndexInLocalMeanings = int(localMeaningIndex)
                currentMeaning = str(wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_MEANING).value).strip().replace(
                    "\u200b", "")  # Fixes Zero Width Space bug

                # Checking is the Local meaning is already equal to one of the Jisho word's meanings
                foundMeaning = False
                jishoMeaningLoopIndex = 0
                for jishoWordMeaningString in jishoWordMeanings:

                    if jishoWordTypesForSheets[jishoMeaningLoopIndex] == "V":
                        if jishoWordMeaningString == currentMeaning:
                            foundMeaning = True
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "J"
                            break
                        elif jishoWordMeaningString in currentMeaning:
                            foundMeaning = True
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "LOC"
                            break
                        elif currentMeaning in jishoWordMeaningString:
                            foundMeaning = True
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
                            wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "J"
                            break
                    jishoMeaningLoopIndex += 1

                if foundMeaning:
                    del jishoWordMeanings[jishoMeaningLoopIndex]
                    del jishoWordTypesForSheets[jishoMeaningLoopIndex]
                    del jishoWordMeaningTypes[jishoMeaningLoopIndex]

                localMeaningLoopIndex += 1

            # The leftover meanings are not present in the local database, so add them to it and update the Local meanings index list
            for i in range(len(jishoWordMeanings)):

                # If the verb is a suru verb, the meanings are not added since it is assumed that it already has the correct meanings
                if (jishoWordTypesForSheets[i] == "V") \
                        & (not (" suru" in currentLocalRomaji)) \
                        & (not ("ssuru" in currentLocalRomaji)):
                    jishoWordMeaningString = jishoWordMeanings[i]
                    jishoWordMeaningType = jishoWordMeaningTypes[i]

                    # Adding the results to the meanings excel sheet
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_INDEX).value = lastLocalMeaningsIndex + 1
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_TYPE).value = jishoWordMeaningType
                    wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_SOURCE).value = "J"

                    # Updating the meaning indexes in the Local Types
                    wsLocalVerbs.cell(row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_MEANINGS_EN).value = str(wsLocalVerbs.cell(
                        row=rowIndexInLocalWordSheet, column=Globals.VERBS_COL_MEANINGS_EN).value) + ";" + str(lastLocalMeaningsIndex + 1)

                    # Incrementing the meaning index for the next iteration
                    lastLocalMeaningsIndex += 1

            # Once the meanings have been handled, break the while loop in order to process the next Jisho word
            break

        rowIndexInLocalWordSheet += 1
    # endregion

    # region Finding the identical local entry in the Types sheet
    rowIndexInLocalWordSheet = 2
    while True:
        value = wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_KANJI).value
        if not value:
            break

        currentLocalRomaji = str(wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ROMAJI).value)
        currentLocalKanji = str(wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_KANJI).value)
        currentLocalMeaningsIndexesEN = str(wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value)
        currentLocalAltSpellings = str(wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ALTS).value)
        currentLocalCommon = str(wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_COMMON).value)

        # Once the identical local entry is found, find the meanings that are unique to the Jisho word, and add them to the database
        if (currentLocalRomaji == jishoWordRomaji) and (currentLocalKanji == jishoWordKanji):

            wordAlreadyExists = True

            # Updating the altSpellings
            if currentLocalAltSpellings == "" or currentLocalAltSpellings == "None":
                wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ALTS).value = jishoWordAltSpellings
            else:
                currentLocalAltSpellingsList = [element.strip() for element in currentLocalAltSpellings.split(",")]
                for jishoWordAltSpelling in jishoWordAltSpellings.split(","):
                    if jishoWordAltSpelling not in currentLocalAltSpellings:
                        currentLocalAltSpellingsList.append(jishoWordAltSpelling)
                wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_ALTS).value = ', '.join(currentLocalAltSpellingsList)

            # Updating the common status
            if currentLocalCommon == 0 or currentLocalCommon is None or currentLocalCommon == "":
                wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_COMMON).value = jishoWordCommon

            # Updating the meanings
            # Removing identical meaning entries from the jisho meanings list: they correspond to different types (e.g. Noun and Ana)
            # but may cause problems with already existing multiple entries.
            # Namely, meaning_Noun = meaning_Ana in Jisho, but meaning_Noun != meaning_Ana in local, so the next loop would create a new
            # version of meaning_Noun or meaning_Ana
            jishoWordMeaningStringIndex = 0
            while jishoWordMeaningStringIndex < len(jishoWordMeanings) - 1:
                duplicateIndex = jishoWordMeaningStringIndex + 1
                while duplicateIndex < len(jishoWordMeanings):
                    if jishoWordMeanings[duplicateIndex] == jishoWordMeanings[jishoWordMeaningStringIndex]:
                        del jishoWordMeanings[duplicateIndex]
                        del jishoWordTypesForSheets[duplicateIndex]
                        del jishoWordMeaningTypes[duplicateIndex]
                    else:
                        duplicateIndex += 1
                jishoWordMeaningStringIndex += 1

            localMeaningIndexes = currentLocalMeaningsIndexesEN.split(";")
            localMeaningLoopIndex = 0
            while localMeaningLoopIndex < len(localMeaningIndexes):

                localMeaningIndex = localMeaningIndexes[localMeaningLoopIndex]

                # Getting the current local meaning
                rowIndexInLocalMeanings = int(localMeaningIndex)
                currentMeaning = str(wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_MEANING).value).strip().replace(
                    "\u200b", "")  # Fixes Zero Width Space bug

                # Checking is the Local meaning is already equal to one of the Jisho word's meanings
                foundMeaning = False
                jishoMeaningLoopIndex = 0
                for jishoWordMeaningString in jishoWordMeanings:

                    if jishoWordMeaningString == currentMeaning:
                        foundMeaning = True
                        wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "J"
                        break
                    elif jishoWordMeaningString in currentMeaning:
                        foundMeaning = True
                        wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "LOC"
                        break
                    elif currentMeaning in jishoWordMeaningString:
                        foundMeaning = True
                        wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
                        wsLocalMeaningsEN.cell(row=rowIndexInLocalMeanings, column=Globals.MEANINGS_COL_SOURCE).value = "J"
                        break
                    jishoMeaningLoopIndex += 1

                if foundMeaning:
                    del jishoWordMeanings[jishoMeaningLoopIndex]
                    del jishoWordTypesForSheets[jishoMeaningLoopIndex]
                    del jishoWordMeaningTypes[jishoMeaningLoopIndex]

                localMeaningLoopIndex += 1

            # The leftover meanings are not present in the local database, so add them to it and update the Local meanings index list
            for i in range(len(jishoWordMeanings)):
                jishoWordMeaningString = jishoWordMeanings[i]
                jishoWordMeaningType = jishoWordMeaningTypes[i]

                # Adding the results to the meanings excel sheet
                wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_INDEX).value = lastLocalMeaningsIndex + 1
                wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
                wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_TYPE).value = jishoWordMeaningType
                wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_SOURCE).value = "J"

                # Updating the meaning indexes in the Local Types
                wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value = str(wsLocalTypes.cell(
                    row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value) + ";" + str(lastLocalMeaningsIndex + 1)

                # Incrementing the meaning index for the next iteration
                lastLocalMeaningsIndex += 1

            # Once the meanings have been handled, break the while loop in order to process the next Jisho word
            break

        rowIndexInLocalWordSheet += 1
    # endregion

    # region If there are still meanings left unregistered, create a new entry to the Types sheet
    if not wordAlreadyExists:

        # Adding a new entry in the Local Types list
        wsLocalTypes.cell(row=lastLocalTypesIndex + 1, column=Globals.TYPES_COL_ROMAJI).value = jishoWordRomaji
        wsLocalTypes.cell(row=lastLocalTypesIndex + 1, column=Globals.TYPES_COL_KANJI).value = jishoWordKanji
        wsLocalTypes.cell(row=lastLocalTypesIndex + 1, column=Globals.TYPES_COL_ALTS).value = jishoWordAltSpellings
        wsLocalTypes.cell(row=lastLocalTypesIndex + 1, column=Globals.TYPES_COL_COMMON).value = jishoWordCommon

        # Adding the meanings to the Local Meanings list and updating the meanings indexes in the Types list
        for i in range(len(jishoWordMeanings)):
            jishoWordMeaningString = jishoWordMeanings[i]
            jishoWordMeaningType = jishoWordMeaningTypes[i]

            # Adding the results to the meanings excel sheet
            wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_INDEX).value = lastLocalMeaningsIndex + 1
            wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_MEANING).value = jishoWordMeaningString
            wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_TYPE).value = jishoWordMeaningType
            wsLocalMeaningsEN.cell(row=lastLocalMeaningsIndex + 1, column=Globals.MEANINGS_COL_SOURCE).value = "J"

            # Updating the meaning indexes in the Local Types
            if i == 0:
                wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value = lastLocalMeaningsIndex + 1
            else:
                wsLocalTypes.cell(row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value = str(wsLocalTypes.cell(
                    row=rowIndexInLocalWordSheet, column=Globals.TYPES_COL_MEANINGS_EN).value) + ";" + str(lastLocalMeaningsIndex + 1)

            # Incrementing the Meaning index for the next iteration
            lastLocalMeaningsIndex += 1

        # Incrementing the Type index for the next Jisho word
        lastLocalTypesIndex += 1
    # endregion

# endregion

# region Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with firebase results.xlsx')
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with firebase results.xlsx')
# endregion
