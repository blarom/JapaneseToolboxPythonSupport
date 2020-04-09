#!/usr/bin/python -tt

import re
import openpyxl

import Globals
from Converter import Converter

current_character = ''

# region Preparations
kanjiDictWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/KanjiDict - 3000 Kanji.xlsx', data_only=True)
wsKanjiDict2 = kanjiDictWorkbook["KanjiDict2"]
# endregion

# region Filling KanjiDict list
words = []
reached_first_character = False
excel_row = 2
with open("kanjidic2.xml", encoding='utf-8') as infile:
    for line in infile:

        if "<character>" in line:
            reached_first_character = True
        if not reached_first_character:
            continue

        current_character += line

        if "</character>" in line:
            match = re.search(r"<literal>(\w+)</literal>", current_character)
            kanji = ''
            if match:
                kanji = match.group(1)

            match = re.search(r"<radical [^>]+>(\w+)</radical>", current_character)
            radical_num = ''
            if match:
                radical_num = match.group(1)

            readingsON = re.findall(r"reading r_type=\"ja_on\">(.*?)</reading>", current_character, re.MULTILINE | re.DOTALL)
            readingsKUN = re.findall(r"reading r_type=\"ja_kun\">(.*?)</reading>", current_character, re.MULTILINE | re.DOTALL)
            readingsNAME = re.findall(r"nanori>(.*?)</nanori>", current_character, re.MULTILINE | re.DOTALL)

            meaningsEN = re.findall(r"<meaning>(.*?)</meaning>", current_character, re.MULTILINE | re.DOTALL)
            meaningsFR = re.findall(r"<meaning m_lang=\"fr\">(.*?)</meaning>", current_character, re.MULTILINE | re.DOTALL)
            meaningsES = re.findall(r"<meaning m_lang=\"es\">(.*?)</meaning>", current_character, re.MULTILINE | re.DOTALL)

            unicode = str(kanji.encode("utf-8"))
            unicode_components = re.findall(r'x([\w\d][\w\d])', unicode)
            wsKanjiDict2.cell(row=excel_row, column=1).value = '1.' + ''.join(unicode_components).upper()
            wsKanjiDict2.cell(row=excel_row, column=2).value = kanji
            wsKanjiDict2.cell(row=excel_row, column=3).value = ";".join(readingsON) + '#' + ";".join(readingsKUN) + '#' + ";".join(readingsNAME)
            wsKanjiDict2.cell(row=excel_row, column=4).value = ", ".join(meaningsEN)
            wsKanjiDict2.cell(row=excel_row, column=5).value = ", ".join(meaningsFR)
            wsKanjiDict2.cell(row=excel_row, column=6).value = ", ".join(meaningsES)

            current_character = ''
            excel_row += 1

# endregion

# Saving the results
kanjiDictWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/KanjiDict - 3000 Kanji.xlsx')