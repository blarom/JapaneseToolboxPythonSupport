#!/usr/bin/python -tt

import re
import openpyxl

import Globals
from Converter import Converter

localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - before foreign.xlsx', data_only=True)
wsLocalMeaningsEN = localWordsWorkbook["Meanings"]
wsLocalMeaningsFR = localWordsWorkbook["MeaningsFR"]
wsLocalMeaningsES = localWordsWorkbook["MeaningsES"]


def remove_duplicates(input_list):
    output = []
    for x in input_list:
        if x not in output:
            output.append(x)
    return output


meaning_index = 1600
while meaning_index < 5539:

    meaning = wsLocalMeaningsFR.cell(row=meaning_index, column=Globals.MEANINGS_COL_MEANING).value
    if meaning is not None and meaning != '':
        if meaning[-1] == '.': meaning = meaning[:-1]
        meaning = meaning[0].lower() + meaning[1:]
        meaning = meaning.replace('par exemple, ', 'ex. ')
        meaning_list = [meaning.strip() for meaning in re.split(r',(?![^(]*[)])', meaning)]
        meaning_list = [meaning[2:] if len(meaning) > 2 and meaning[:2].lower() == 'à ' else meaning for meaning in meaning_list]
        meaning = ', '.join(remove_duplicates(meaning_list))
        wsLocalMeaningsFR.cell(row=meaning_index, column=Globals.MEANINGS_COL_MEANING).value = meaning

    meaning = wsLocalMeaningsES.cell(row=meaning_index, column=Globals.MEANINGS_COL_MEANING).value
    if meaning is not None and meaning != '':
        if meaning[-1] == '.': meaning = meaning[:-1]
        meaning = meaning[0].lower() + meaning[1:]
        meaning = meaning.replace('por ejemplo, ', 'p.ej. ')
        meaning_list = [meaning.strip() for meaning in re.split(r',(?![^(]*[)])', meaning)]
        meaning_list = [meaning[5:] if len(meaning) > 5 and meaning[:5].lower() == 'a la ' else meaning for meaning in meaning_list]
        meaning_list = [meaning[2:] if len(meaning) > 2 and meaning[:2].lower() == 'a ' else meaning for meaning in meaning_list]
        meaning_list = [meaning[5:] if len(meaning) > 5 and meaning[:5].lower() == 'para ' else meaning for meaning in meaning_list]
        meaning = ', '.join(remove_duplicates(meaning_list))
        wsLocalMeaningsES.cell(row=meaning_index, column=Globals.MEANINGS_COL_MEANING).value = meaning

    meaning_index += 1

localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - before foreign - suru verbs.xlsx')
