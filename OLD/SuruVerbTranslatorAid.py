#!/usr/bin/python -tt

import re
import openpyxl

import Globals
from Converter import Converter

localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - before foreign.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - before foreign.xlsx', data_only=True)
wsLocalMeaningsEN = localWordsWorkbook["Meanings"]
wsLocalMeaningsFR = localWordsWorkbook["MeaningsFR"]
wsLocalMeaningsES = localWordsWorkbook["MeaningsES"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]

fr_exceptions = [2116, 2375, 2427, 2754, 2961, 2984, 3271, 3283]
es_exceptions = [2116, 2152, 2162, 2375, 2427, 2662, 2663, 2752, 2754, 2960, 2961, 2984, 3271, 3283]
index_suru_verbs = 2112
index_meaningsFR = 1600
index_meaningsES = 1600
while wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.TYPES_COL_ROMAJI).value is not None \
        and len(wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.TYPES_COL_ROMAJI).value) > 0:

    if wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.TYPES_COL_ROMAJI).value == '': break

    meaning_indexes_string = wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_EN).value

    meaning_indexes = []
    try:
        meaning_indexes = [int(index.strip()) for index in str(meaning_indexes_string).split(";")]
    except:
        index_suru_verbs += 1
        continue

    for meaning_index in meaning_indexes:
        meaning = wsLocalMeaningsEN.cell(row=meaning_index, column=Globals.MEANINGS_COL_MEANING).value
        pos = wsLocalMeaningsEN.cell(row=meaning_index, column=Globals.MEANINGS_COL_TYPE).value

        meaning = "to " + meaning.replace(", ", ", to ")

        if index_suru_verbs not in fr_exceptions:
            wsLocalMeaningsFR.cell(row=index_meaningsFR, column=Globals.MEANINGS_COL_INDEX).value = str(index_meaningsFR)
            wsLocalMeaningsFR.cell(row=index_meaningsFR, column=Globals.MEANINGS_COL_MEANING).value = meaning
            wsLocalMeaningsFR.cell(row=index_meaningsFR, column=Globals.MEANINGS_COL_TYPE).value = pos
            wsLocalMeaningsFR.cell(row=index_meaningsFR, column=Globals.MEANINGS_COL_EXPL).value = meaning

            meaningsFR = wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_FR).value
            if meaningsFR is not None and len(meaningsFR) > 0:
                meaningsFR_list = meaningsFR.split(";")
                meaningsFR_list.append(str(index_meaningsFR))
                wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_FR).value = ";".join(meaningsFR_list)
            else:
                wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_FR).value = str(index_meaningsFR)


        if index_suru_verbs not in es_exceptions:
            wsLocalMeaningsES.cell(row=index_meaningsES, column=Globals.MEANINGS_COL_INDEX).value = str(index_meaningsES)
            wsLocalMeaningsES.cell(row=index_meaningsES, column=Globals.MEANINGS_COL_MEANING).value = meaning
            wsLocalMeaningsES.cell(row=index_meaningsES, column=Globals.MEANINGS_COL_TYPE).value = pos
            wsLocalMeaningsES.cell(row=index_meaningsES, column=Globals.MEANINGS_COL_EXPL).value = meaning

            meaningsES = wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_ES).value
            if meaningsES is not None and len(meaningsES) > 0:
                meaningsES_list = meaningsES.split(";")
                meaningsES_list.append(str(index_meaningsES))
                wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_ES).value = ";".join(meaningsES_list)
            else:
                wsLocalVerbs.cell(row=index_suru_verbs, column=Globals.VERBS_COL_MEANINGS_ES).value = str(index_meaningsES)


        index_meaningsFR += 1
        index_meaningsES += 1

    index_suru_verbs += 1

localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - before foreign - suru verbs.xlsx')
