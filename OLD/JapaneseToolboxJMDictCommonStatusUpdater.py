#!/usr/bin/python -tt

import re
import openpyxl
from JapaneseToolboxConverter import Converter

current_entry = ''
VERB_PARTS_OF_SPEECH = ['Vrui', 'Vbu', 'Vgu', 'Vku', 'Vmu', 'Vnu', 'Vrug', 'Vsu', 'Vtsu', 'Vu', 'Vi', 'Viku', 'Vus', 'Varu', 'Vkuru', 'Vsuru']
LEGEND_DICT = {
    'MA': 'ZM',
    'X': 'IES',
    'abbr': 'Abr',
    'adj-i': 'Ai',
    'adj-ix': 'Ai',
    'adj-na': 'Ana',
    'adj-no': 'Ano',
    'adj-pn': 'Apn',
    'adj-t': 'Atr',
    'adj-f': 'Af',
    'adv': 'A',
    'adv-to': 'Ato',
    'arch': 'ar',
    'ateji': '',
    'aux': '',
    'aux-v': 'Vx',
    'aux-adj': 'Ax',
    'Buddh': 'ZB',
    'chem': 'ZC',
    'chn': 'ZCL',
    'col': 'coq',
    'comp': 'ZI',
    'conj': 'CO',
    'cop-da': '',
    'ctr': 'C',
    'derog': 'Dr',
    'eK': '',
    'ek': '',
    'exp': 'CE',
    'fam': 'LF',
    'fem': 'LFt',
    'food': 'ZF',
    'geom': 'ZG',
    'gikun': '',
    'hon': 'LHn',
    'hum': 'LHm',
    'iK': '',
    'id': 'CES',
    'ik': '',
    'int': 'OI',
    'io': '',
    'iv': '',
    'ling': 'ZL',
    'm-sl': 'ZMg',
    'male': 'LMt',
    'male-sl': 'LMt',
    'math': 'ZMt',
    'mil': 'ZMl',
    'n': 'N',
    'n-adv': 'NAdv',
    'n-suf': 'N;Sx',
    'n-pref': 'N;Px',
    'n-t': 'T',
    'num': 'num',
    'oK': '',
    'obs': 'Obs',
    'obsc': 'Obs',
    'ok': 'Obs',
    'oik': 'Obs',
    'on-mim': 'OI',
    'pn': 'P',
    'poet': 'ZP',
    'pol': 'LHn',
    'pref': 'Px',
    'proverb': 'CES',
    'prt': 'PP',
    'physics': 'ZPh',
    'quote': '',
    'rare': 'Obs',
    'sens': '',
    'sl': 'Sl',
    'suf': 'Sx',
    'uK': '',
    'uk': '',
    'unc': 'UNC',
    'yoji': '',
    'v1': 'Vrui',
    'v1-s': 'Vrui',
    'v2a-s': '',
    'v4h': '',
    'v4r': '',
    'v5aru': 'Varu',
    'v5b': 'Vbu',
    'v5g': 'Vgu',
    'v5k': 'Vku',
    'v5k-s': 'Viku',
    'v5m': 'Vmu',
    'v5n': 'Vnu',
    'v5r': 'Vrug',
    'v5r-i': 'Vrug',
    'v5s': 'Vsu',
    'v5t': 'Vtsu',
    'v5u': 'Vu',
    'v5u-s': 'Vus',
    'v5uru': '',
    'vz': 'Vrui',
    'vi': 'intransitive',
    'vk': 'Vkuru',
    'vn': '',
    'vr': '',
    'vs': 'Vsuru',
    'vs-c': '',
    'vs-s': 'Vsuru',
    'vs-i': 'Vsuru',
    'kyb': '',
    'osb': '',
    'ksb': '',
    'ktb': '',
    'tsb': '',
    'thb': '',
    'tsug': '',
    'kyu': '',
    'rkb': '',
    'nab': '',
    'hob': '',
    'vt': 'transitive',
    'vulg': 'vul',
    'adj-kari': '',
    'adj-ku': '',
    'adj-shiku': '',
    'adj-nari': '',
    'n-pr': 'Ne',
    'v-unspec': '',
    'v4k': '',
    'v4g': '',
    'v4s': '',
    'v4t': '',
    'v4n': '',
    'v4b': '',
    'v4m': '',
    'v2k-k': '',
    'v2g-k': '',
    'v2t-k': '',
    'v2d-k': '',
    'v2h-k': '',
    'v2b-k': '',
    'v2m-k': '',
    'v2y-k': '',
    'v2r-k': '',
    'v2k-s': '',
    'v2g-s': '',
    'v2s-s': '',
    'v2z-s': '',
    'v2t-s': '',
    'v2d-s': '',
    'v2n-s': '',
    'v2h-s': '',
    'v2b-s': '',
    'v2m-s': '',
    'v2y-s': '',
    'v2r-s': '',
    'v2w-s': '',
    'archit': 'ZAc',
    'astron': 'ZAs',
    'baseb': 'ZBb',
    'biol': 'ZBi',
    'bot': 'ZBt',
    'bus': 'ZBs',
    'econ': 'ZEc',
    'engr': 'ZEg',
    'finc': 'ZFn',
    'geol': 'ZGg',
    'law': 'ZLw',
    'mahj': 'ZMj',
    'med': 'Md',
    'music': 'ZMc',
    'Shinto': 'ZSt',
    'shogi': 'ZSg',
    'sports': 'ZSp',
    'sumo': 'ZSm',
    'zool': 'ZZ',
    'joc': 'ZH',
    'anat': 'ZAn'
}
MAX_NUM_SURU_VERBS_TO_ADD = 300
UNIQUE_DELIMITER = "---"

# region Preparations
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji.xlsx', data_only=True)
wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]


class Word:
    def __init__(self, akanji, aaltSpellings, ahiragana, aromaji, aenglish_meanings, atypes, acommon):
        self.kanji = akanji
        self.altSpellings = aaltSpellings
        self.hiragana = ahiragana
        self.romaji = aromaji
        self.uniqueID = self.romaji + UNIQUE_DELIMITER + self.kanji
        self.english_meanings = aenglish_meanings
        self.types = atypes
        self.common = acommon
        self.hasSuruTwin = False

    def __str__(self):
        return "Word:\n" \
               + "\t" + self.romaji \
               + "\t" + self.kanji


converter = Converter()
# endregion

# region Filling the english words list
words = []
reached_first_entry = False
with open("JMDict_e", encoding='utf-8') as infile:
    for line in infile:

        if "<entry>" in line:
            reached_first_entry = True
        if not reached_first_entry:
            continue

        current_entry += line

        if "</entry>" in line:
            current_entry = current_entry.replace('\n', '').replace('\r', '')

            # region Getting the basic characteristics
            kanjis = re.findall(r"<keb>(.+?)</keb>", current_entry, re.U)
            if len(kanjis) > 0:
                kanji = kanjis[0]
                altSpellings = []
                for i in range(1, len(kanjis)):
                    altSpellings.append(kanjis[i])
            else:
                altSpellings = []
                kanji = ''

            kanas = re.findall(r"<reb>(.+?)</reb>", current_entry, re.U)
            hiragana = ''
            romaji = ''
            if len(kanas) > 0:
                hiragana = kanas[0]
                romaji = converter.get_latin_hiragana_katakana(hiragana)[converter.TYPE_LATIN]
                if kanji == '': kanji = hiragana

                for i in range(1, len(kanas)):
                    altSpellingRomaji = converter.get_latin_hiragana_katakana(kanas[i])[converter.TYPE_LATIN]
                    if altSpellingRomaji != romaji:
                        altSpellings.append(altSpellingRomaji)

            # endregion

            # region Getting the common status - if not common, skip this word
            match = re.search(r"<ke_pri>(news1|ichi1)</ke_pri>", current_entry)
            #match = re.search(r"<ke_pri>(news1|news2|ichi1|ichi2|spec1|spec2|gai1|gai2)</ke_pri>", current_entry)
            if match:
                common = True
            else:
                common = False
            # endregion

            # region Creating the words
            if not kanji == '' and not hiragana == '':

                word = Word(kanji,
                            altSpellings,
                            hiragana,
                            romaji,
                            [],
                            [],
                            common)

                if (not ((word.romaji == "ha" or word.romaji == "wa") and word.kanji == "は")) \
                        and (not ((word.romaji == "he" or word.romaji == "e") and word.kanji == "へ")) \
                        and (not ((word.romaji == "deha" or word.romaji == "dewa") and word.kanji == "では")) \
                        and (not ((word.romaji == "niha" or word.romaji == "niwa") and word.kanji == "には")) \
                        and (not (word.romaji == "kana" and word.kanji == "かな")) \
                        and (not (word.romaji == "to" and word.kanji == "と")) \
                        and (not (word.romaji == "ya" and word.kanji == "や")) \
                        and (not (word.romaji == "mo" and word.kanji == "も")) \
                        and (not (word.romaji == "no" and word.kanji == "の")) \
                        and (not (word.romaji == "n" and word.kanji == "ん")) \
                        and (not (word.romaji == "wo" and word.kanji == "を")) \
                        and (not (word.romaji == "wa" and word.kanji == "わ")) \
                        and (not (word.romaji == "yo" and word.kanji == "よ")) \
                        and (not (word.romaji == "na" and word.kanji == "な")) \
                        and (not (word.romaji == "ka" and word.kanji == "か")) \
                        and (not (word.romaji == "ga" and word.kanji == "が")) \
                        and (not (word.romaji == "ni" and word.kanji == "に")) \
                        and (not (word.kanji == "ケ" and word.kanji == "ヶ")):
                    words.append(word)

                if len(words) % 1000 == 0:
                    print("built word number " + str(len(words)) + ": " + word.kanji)
            # endregion

            current_entry = ''


# endregion

# region Filtering the words list
def binary_search(words_list, text, lo=0, hi=None):
    # https://stackoverflow.com/questions/4161629/search-a-list-of-objects-in-python
    if hi is None:
        hi = len(words_list)
    while lo < hi:
        mid = (lo + hi) // 2

        if text > words_list[mid].uniqueID:
            lo = mid + 1
        elif text < words_list[mid].uniqueID:
            hi = mid
        elif text == words_list[mid].uniqueID:
            return mid
        else:
            return -1

    return -1


workingWordsList = list(words)
workingWordsList.sort(key=lambda x: x.uniqueID)

print("total number of words: " + str(len(workingWordsList)))

typesIndex = 1
while True:
    romaji = wsLocalTypes.cell(row=typesIndex, column=3).value
    kanji = wsLocalTypes.cell(row=typesIndex, column=4).value
    common = wsLocalTypes.cell(row=typesIndex, column=7).value

    if romaji == "" or not romaji:
        break

    if common is not None and common != '':
        typesIndex += 1
        continue

    uniqueID = romaji + UNIQUE_DELIMITER + kanji
    #print(uniqueID + "\n")

    index_of_first_hit = binary_search(workingWordsList, uniqueID)
    if index_of_first_hit != -1:
        common_status = 1 if workingWordsList[index_of_first_hit].common else 0
        wsLocalTypes.cell(row=typesIndex, column=7).value = common_status

    if typesIndex % 50 == 0:
        print("types index " + str(typesIndex), end="\r")
    if typesIndex % 1000 == 0:
        print("types index " + str(typesIndex))

    typesIndex += 1

verbsIndex = 1
while True:
    romaji = wsLocalVerbs.cell(row=verbsIndex, column=7).value
    kanji = wsLocalVerbs.cell(row=verbsIndex, column=6).value
    common = wsLocalTypes.cell(row=typesIndex, column=14).value

    if wsLocalVerbs.cell(row=verbsIndex, column=1).value == "-" and wsLocalVerbs.cell(row=verbsIndex, column=2).value == "-":
        break

    if romaji is None or kanji is None:
        verbsIndex += 1
        continue

    if common is not None and common != '':
        verbsIndex += 1
        continue

    if ' suru' in romaji:
        romaji = wsLocalVerbs.cell(row=verbsIndex, column=9).value.strip()
        kanji = wsLocalVerbs.cell(row=verbsIndex, column=8).value.strip()

    uniqueID = romaji + UNIQUE_DELIMITER + kanji

    #print(uniqueID + "\n")

    index_of_first_hit = binary_search(workingWordsList, uniqueID)
    if index_of_first_hit != -1:
        common_status = 1 if workingWordsList[index_of_first_hit].common else 0
        wsLocalVerbs.cell(row=verbsIndex, column=14).value = common_status

    if verbsIndex % 50 == 0:
        print("verb index " + str(verbsIndex), end="\r")
    if verbsIndex % 1000 == 0:
        print("verb index " + str(verbsIndex))

    verbsIndex += 1


# region Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji.xlsx')
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji.xlsx')
# endregion
