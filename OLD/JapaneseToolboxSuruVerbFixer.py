# Japanese Toolbox suru verb comparator

import openpyxl
from urllib.request import urlopen
import urllib.parse
from urllib.error import HTTPError, URLError
import re

# region Getting the excel sheets
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji.xlsx', data_only=True)

wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]

lastLocalMeaningsIndex = 1
while True:
    value = wsLocalMeanings.cell(row=lastLocalMeaningsIndex, column=3).value
    if not value:
        lastLocalMeaningsIndex -= 1
        break
    lastLocalMeaningsIndex += 1

lastLocalVerbsIndex = 1
while True:
    value = wsLocalVerbs.cell(row=lastLocalVerbsIndex, column=3).value
    value2 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 1, column=3).value
    value3 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 2, column=3).value
    if (not value) & (not value2) & (not value3):
        lastLocalVerbsIndex -= 1
        break
    lastLocalVerbsIndex += 1
# endregion

# region Running through the suru verbs and adding the local and online meanings
verbRowIndex = 1
while verbRowIndex < lastLocalVerbsIndex:

    if wsLocalVerbs.cell(row=verbRowIndex, column=21).value is not None \
            and wsLocalVerbs.cell(row=verbRowIndex, column=21).value != '' \
            and (wsLocalVerbs.cell(row=verbRowIndex, column=16).value == 'T' or wsLocalVerbs.cell(row=verbRowIndex,
                                                                                                  column=16).value == 'I'):

        kanjiRoot = wsLocalVerbs.cell(row=verbRowIndex, column=8).value
        romajiRoot = wsLocalVerbs.cell(row=verbRowIndex, column=9).value
        meaningIndexesAsString = wsLocalVerbs.cell(row=verbRowIndex, column=12).value

        if kanjiRoot is None:
            verbRowIndex += 1
            continue

        meaningIndexes = []
        try:
            meaningIndexes = [int(meaningIndexAsString.strip()) for meaningIndexAsString in
                              str(meaningIndexesAsString).split(";")]
        except:
            verbRowIndex += 1
            continue

        transElements = wsLocalVerbs.cell(row=verbRowIndex, column=15).value
        if len(transElements.split(",")) == 1 and ";" not in transElements[0] and len(meaningIndexes) > 0:

            wsLocalVerbs.cell(row=verbRowIndex, column=2).value = wsLocalVerbs.cell(row=verbRowIndex, column=21).value
            wsLocalMeanings.cell(row=int(meaningIndexes[0]), column=2).value = wsLocalVerbs.cell(row=verbRowIndex,
                                                                                                 column=21).value

            if wsLocalVerbs.cell(row=verbRowIndex, column=16).value == 'T':
                wsLocalMeanings.cell(row=int(meaningIndexes[0]), column=3).value = "VsuruT"
            elif wsLocalVerbs.cell(row=verbRowIndex, column=16).value == 'I':
                wsLocalMeanings.cell(row=int(meaningIndexes[0]), column=3).value = "VsuruI"

            wsLocalVerbs.cell(row=verbRowIndex, column=21).value = None
            wsLocalVerbs.cell(row=verbRowIndex, column=17).value = None
            wsLocalVerbs.cell(row=verbRowIndex, column=16).value = None
            wsLocalVerbs.cell(row=verbRowIndex, column=15).value = 'MATCHES'
            wsLocalVerbs.cell(row=verbRowIndex, column=14).value = '2'

    if wsLocalVerbs.cell(row=verbRowIndex, column=15).value == 'MATCHES':
        wsLocalVerbs.cell(row=verbRowIndex, column=14).value = '2'

    verbRowIndex += 1

# endregion

# region Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with suru verbs data.xlsx')
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with suru verbs data.xlsx')
# endregion
