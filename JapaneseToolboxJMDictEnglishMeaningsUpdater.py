#!/usr/bin/python -tt

import re
import openpyxl
from JapaneseToolboxConverter import Converter

current_entry = ''
VERB_PARTS_OF_SPEECH = ['Vrui', 'Vbu', 'Vgu', 'Vku', 'Vmu', 'Vnu', 'Vrug', 'Vsu', 'Vtsu', 'Vu', 'Vi', 'Viku', 'Vus', 'Varu', 'Vkuru', 'Vsuru']
LEGEND_DICT = {
    'MA': 'ZM',
    'X': 'IES',
    'abbr': 'Abr',
    'adj-i': 'Ai',
    'adj-ix': 'Ai',
    'adj-na': 'Ana',
    'adj-no': 'Ano',
    'adj-pn': 'Apn',
    'adj-t': 'Atr',
    'adj-f': 'Af',
    'adv': 'A',
    'adv-to': 'Ato',
    'arch': 'ar',
    'ateji': '',
    'aux': '',
    'aux-v': 'Vx',
    'aux-adj': 'Ax',
    'Buddh': 'ZB',
    'chem': 'ZC',
    'chn': 'ZCL',
    'col': 'coq',
    'comp': 'ZI',
    'conj': 'CO',
    'cop-da': '',
    'ctr': 'C',
    'derog': 'Dr',
    'eK': '',
    'ek': '',
    'exp': 'CE',
    'fam': 'LF',
    'fem': 'LFt',
    'food': 'ZF',
    'geom': 'ZG',
    'gikun': '',
    'hon': 'LHn',
    'hum': 'LHm',
    'iK': '',
    'id': 'CES',
    'ik': '',
    'int': 'OI',
    'io': '',
    'iv': '',
    'ling': 'ZL',
    'm-sl': 'ZMg',
    'male': 'LMt',
    'male-sl': 'LMt',
    'math': 'ZMt',
    'mil': 'ZMl',
    'n': 'N',
    'n-adv': 'NAdv',
    'n-suf': 'N;Sx',
    'n-pref': 'N;Px',
    'n-t': 'T',
    'num': 'num',
    'oK': '',
    'obs': 'Obs',
    'obsc': 'Obs',
    'ok': 'Obs',
    'oik': 'Obs',
    'on-mim': 'OI',
    'pn': 'P',
    'poet': 'ZP',
    'pol': 'LHn',
    'pref': 'Px',
    'proverb': 'CES',
    'prt': 'PP',
    'physics': 'ZPh',
    'quote': '',
    'rare': 'Obs',
    'sens': '',
    'sl': 'Sl',
    'suf': 'Sx',
    'uK': '',
    'uk': '',
    'unc': 'UNC',
    'yoji': '',
    'v1': 'Vrui',
    'v1-s': 'Vrui',
    'v2a-s': '',
    'v4h': '',
    'v4r': '',
    'v5aru': 'Varu',
    'v5b': 'Vbu',
    'v5g': 'Vgu',
    'v5k': 'Vku',
    'v5k-s': 'Viku',
    'v5m': 'Vmu',
    'v5n': 'Vnu',
    'v5r': 'Vrug',
    'v5r-i': 'Vrug',
    'v5s': 'Vsu',
    'v5t': 'Vtsu',
    'v5u': 'Vu',
    'v5u-s': 'Vus',
    'v5uru': '',
    'vz': 'Vrui',
    'vi': 'intransitive',
    'vk': 'Vkuru',
    'vn': '',
    'vr': '',
    'vs': 'Vsuru',
    'vs-c': '',
    'vs-s': 'Vsuru',
    'vs-i': 'Vsuru',
    'kyb': '',
    'osb': '',
    'ksb': '',
    'ktb': '',
    'tsb': '',
    'thb': '',
    'tsug': '',
    'kyu': '',
    'rkb': '',
    'nab': '',
    'hob': '',
    'vt': 'transitive',
    'vulg': 'vul',
    'adj-kari': '',
    'adj-ku': '',
    'adj-shiku': '',
    'adj-nari': '',
    'n-pr': 'Ne',
    'v-unspec': '',
    'v4k': '',
    'v4g': '',
    'v4s': '',
    'v4t': '',
    'v4n': '',
    'v4b': '',
    'v4m': '',
    'v2k-k': '',
    'v2g-k': '',
    'v2t-k': '',
    'v2d-k': '',
    'v2h-k': '',
    'v2b-k': '',
    'v2m-k': '',
    'v2y-k': '',
    'v2r-k': '',
    'v2k-s': '',
    'v2g-s': '',
    'v2s-s': '',
    'v2z-s': '',
    'v2t-s': '',
    'v2d-s': '',
    'v2n-s': '',
    'v2h-s': '',
    'v2b-s': '',
    'v2m-s': '',
    'v2y-s': '',
    'v2r-s': '',
    'v2w-s': '',
    'archit': 'ZAc',
    'astron': 'ZAs',
    'baseb': 'ZBb',
    'biol': 'ZBi',
    'bot': 'ZBt',
    'bus': 'ZBs',
    'econ': 'ZEc',
    'engr': 'ZEg',
    'finc': 'ZFn',
    'geol': 'ZGg',
    'law': 'ZLw',
    'mahj': 'ZMj',
    'med': 'Md',
    'music': 'ZMc',
    'Shinto': 'ZSt',
    'shogi': 'ZSg',
    'sports': 'ZSp',
    'sumo': 'ZSm',
    'zool': 'ZZ',
    'joc': 'ZH',
    'anat': 'ZAn'
}
MAX_NUM_SURU_VERBS_TO_ADD = 50
UNIQUE_DELIMITER = "---"

# region Preparations
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji.xlsx', data_only=True)
wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]


class Word:
    def __init__(self, akanji, aaltSpellings, ahiragana, aromaji, aenglish_meanings, atypes, acommon):
        self.kanji = akanji
        self.altSpellings = aaltSpellings
        self.hiragana = ahiragana
        self.romaji = aromaji
        self.uniqueID = self.romaji + UNIQUE_DELIMITER + self.kanji
        self.english_meanings = aenglish_meanings
        self.types = atypes
        self.common = acommon
        self.hasSuruTwin = False

    def __str__(self):
        return "Word:\n" \
               + "\t" + self.romaji \
               + "\t" + self.kanji


converter = Converter()
# endregion

# region Filling the english words list
words = []
suruVerbs = []
num_suru_verbs = 0
reached_first_entry = False
with open("JMDict_e", encoding='utf-8') as infile:
    for line in infile:

        if "<entry>" in line:
            reached_first_entry = True
        if not reached_first_entry:
            continue

        current_entry += line

        if "</entry>" in line:
            current_entry = current_entry.replace('\n', '').replace('\r', '')

            # region Getting the basic characteristics
            kanjis = re.findall(r"<keb>(.+?)</keb>", current_entry, re.U)
            if len(kanjis) > 0:
                kanji = kanjis[0]
                altSpellings = []
                for i in range(1, len(kanjis)):
                    altSpellings.append(kanjis[i])
            else:
                altSpellings = []
                kanji = ''

            kanas = re.findall(r"<reb>(.+?)</reb>", current_entry, re.U)
            hiragana = ''
            romaji = ''
            if len(kanas) > 0:
                hiragana = kanas[0]
                romaji = converter.get_latin_hiragana_katakana(hiragana)[converter.TYPE_LATIN]
                if kanji == '': kanji = hiragana

                for i in range(1, len(kanas)):
                    altSpellingRomaji = converter.get_latin_hiragana_katakana(kanas[i])[converter.TYPE_LATIN]
                    if altSpellingRomaji != romaji:
                        altSpellings.append(altSpellingRomaji)

            # endregion

            # region Getting the common status - if not common, skip this word
            match = re.search(r"<ke_pri>(news1|ichi1)</ke_pri>", current_entry)
            #match = re.search(r"<ke_pri>(news1|news2|ichi1|ichi2|spec1|spec2|gai1|gai2)</ke_pri>", current_entry)
            if match:
                common = True
            else:
                common = False
                current_entry = ''
                continue
            # endregion

            # region Fixing the senses with missing parts of speech
            senses = re.findall(r"<sense>(.+?)</sense>", current_entry, re.DOTALL)
            for i in range(len(senses)-1, 0, -1):
                posTypes = re.findall(r"<pos>&(.+?);</pos>", senses[i])
                if len(posTypes) == 0:
                    for j in range(i-1, -1, -1):
                        posTypesPrev = re.findall(r"<pos>&(.+?);</pos>", senses[j])
                        if len(posTypesPrev) > 0:
                            for posTypePrev in posTypesPrev:
                                senses[i] = r"<pos>&"+posTypePrev+";</pos>" + senses[i]
                            break
            # endregion

            # region Getting the types/meanings and suru verb characteristics
            types = []
            english_meanings = []
            types_for_word = []
            english_meanings_for_word = []
            types_for_suru_verb = []
            english_meanings_for_suru_verb = []
            suru_altSpellings = []
            suru_hiragana = ''
            suru_romaji = ''
            suru_kanji = ''
            types_as_string = '-'
            partsOfSpeechFromLastSense = []
            last_sense_is_result_of_sense_split = False
            create_suru_verb = False
            i = 0
            while i < len(senses):

                if len(senses)>30:
                    a=1

                # region Getting the list of JMtypes
                matches = re.findall(r"<(pos|field|misc)>&(.+?);</(pos|field|misc)>", senses[i])
                JMtypes = [match[1] for match in matches]
                partsOfSpeech = ';'.join([LEGEND_DICT[JMtype] for JMtype in JMtypes if LEGEND_DICT[JMtype] != '']).split(';')
                partsOfSpeech = list(set(partsOfSpeech))
                # endregion

                # region Getting the TRANSITIVE status
                trans = 'I'
                if "transitive" in partsOfSpeech and "intransitive" in partsOfSpeech:
                    senses.append(senses[i].replace("<pos>&vt;</pos>", ""))
                    trans = 'T'
                    partsOfSpeech.remove("transitive")
                    partsOfSpeech.remove("intransitive")
                elif "transitive" in partsOfSpeech:
                    trans = 'T'
                    partsOfSpeech.remove("transitive")
                elif "intransitive" in partsOfSpeech:
                    trans = 'I'
                    partsOfSpeech.remove("intransitive")

                # endregion

                # region If the word includes both na-adj and noun forms, split them into two senses
                if "Ana" in partsOfSpeech and "N" in partsOfSpeech:
                    senses.insert(i+1, re.sub(r"<pos>&(n|adj-no|n-suf|n-pref|n-adv);</pos>", "", senses[i]))
                    partsOfSpeech.remove("Ana")
                    if "Vsuru" in partsOfSpeech: partsOfSpeech.remove("Vsuru")
                # endregion

                # region If the word includes a suru verb option, create a new sense for it
                if "Vsuru" in partsOfSpeech \
                        and ("N" in partsOfSpeech or "Ano" in partsOfSpeech or "Ana" in partsOfSpeech
                             or "Ai" in partsOfSpeech or "A" in partsOfSpeech or "Ato" in partsOfSpeech):
                            senses.append(re.sub("<pos>&((?!vs).)*;</pos>", "", senses[i])) # removes any pos that is not vs
                            partsOfSpeech.remove("Vsuru")
                # endregion

                # region Adding T/I to the part of speech if it is a verb
                type_is_verb = False
                is_suru_verb = False
                for j in range(len(partsOfSpeech)):
                    if partsOfSpeech[j] in VERB_PARTS_OF_SPEECH:
                        partsOfSpeech[j] = partsOfSpeech[j] + trans
                        type_is_verb = True
                        if 'Vsuru' in partsOfSpeech[j]: is_suru_verb = True
                # endregion

                # region Removing part of speech duplicates
                partsOfSpeech = list(set(partsOfSpeech))
                # endregion

                # region Reordering parts of speech
                if "Ai" in partsOfSpeech:
                    partsOfSpeech.remove("Ai")
                    partsOfSpeech.insert(0, "Ai")

                if "Ana" in partsOfSpeech:
                    partsOfSpeech.remove("Ana")
                    partsOfSpeech.insert(0, "Ana")

                if "NAdv" in partsOfSpeech:
                    partsOfSpeech.remove("NAdv")
                    partsOfSpeech.insert(0, "NAdv")

                if "Ano" in partsOfSpeech:
                    partsOfSpeech.remove("Ano")
                    partsOfSpeech.insert(0, "Ano")

                if "A" in partsOfSpeech:
                    partsOfSpeech.remove("A")
                    partsOfSpeech.insert(0, "A")

                if "N" in partsOfSpeech:
                    partsOfSpeech.remove("N")
                    partsOfSpeech.insert(0, "N")

                verbPOS = [element for element in partsOfSpeech if element[:-1] in VERB_PARTS_OF_SPEECH]
                if len(verbPOS) > 0:
                    partsOfSpeech.remove(verbPOS[0])
                    partsOfSpeech.insert(0, verbPOS[0])
                # endregion

                # region Getting the meaning
                matches = re.findall(r"<(gloss|gloss g_type=\"expl\")?>(.+?)</gloss>", senses[i], re.U)
                english_meanings = [match[1] for match in matches]
                english_meanings_as_string = ", ".join(english_meanings).replace("&,", "&")
                # endregion

                # region Setting the word characteristics
                if is_suru_verb:
                    types_for_suru_verb.append(';'.join(partsOfSpeech))
                    english_meanings_for_suru_verb.append(english_meanings_as_string)
                    suru_altSpellings = []
                    for altSpelling in altSpellings:
                        match = re.search(r"[a-zA-Z]", altSpelling)
                        if match: suru_altSpellings.append(altSpelling + "suru")
                        else: suru_altSpellings.append(altSpelling + "する")
                    suru_hiragana = hiragana + "する"
                    suru_romaji = romaji + " suru"
                    suru_kanji = kanji + "する"
                    create_suru_verb = True
                else:
                    types_for_word.append(';'.join(partsOfSpeech))
                    english_meanings_for_word.append(english_meanings_as_string)
                # endregion

                i += 1

            # endregion

            # region Cleaning the types after sense repetition
            for i in range(len(english_meanings_for_word)):
                if i > 0 and english_meanings_for_word[i] == english_meanings_for_word[i - 1]:
                    lastTypes = types_for_word[i - 1].split(';')
                    currentTypes = types_for_word[i].split(';')
                    types_for_word[i] = ";".join([currentType for currentType in currentTypes if currentType not in lastTypes])
            # endregion

            # region Creating the words
            if not kanji == '' and not hiragana == '' and not (len(english_meanings_for_word) == 0):

                word = Word(kanji,
                            altSpellings,
                            hiragana,
                            romaji,
                            english_meanings_for_word,
                            types_for_word,
                            common)
                if create_suru_verb: word.hasSuruTwin = True

                if (not ((word.romaji == "ha" or word.romaji == "wa") and word.kanji == "は")) \
                        and (not ((word.romaji == "he" or word.romaji == "e") and word.kanji == "へ")) \
                        and (not ((word.romaji == "deha" or word.romaji == "dewa") and word.kanji == "では")) \
                        and (not ((word.romaji == "niha" or word.romaji == "niwa") and word.kanji == "には")) \
                        and (not (word.romaji == "kana" and word.kanji == "かな")) \
                        and (not (word.romaji == "to" and word.kanji == "と")) \
                        and (not (word.romaji == "ya" and word.kanji == "や")) \
                        and (not (word.romaji == "mo" and word.kanji == "も")) \
                        and (not (word.romaji == "no" and word.kanji == "の")) \
                        and (not (word.romaji == "n" and word.kanji == "ん")) \
                        and (not (word.romaji == "wo" and word.kanji == "を")) \
                        and (not (word.romaji == "wa" and word.kanji == "わ")) \
                        and (not (word.romaji == "yo" and word.kanji == "よ")) \
                        and (not (word.romaji == "na" and word.kanji == "な")) \
                        and (not (word.romaji == "ka" and word.kanji == "か")) \
                        and (not (word.romaji == "ga" and word.kanji == "が")) \
                        and (not (word.romaji == "ni" and word.kanji == "に")) \
                        and (not (word.kanji == "ケ" and word.kanji == "ヶ")):
                    words.append(word)

                    if create_suru_verb:
                        word = Word(suru_kanji,
                                    suru_altSpellings,
                                    suru_hiragana,
                                    suru_romaji,
                                    english_meanings_for_suru_verb,
                                    types_for_suru_verb,
                                    common)
                        suruVerbs.append(word)
                        num_suru_verbs += 1

                if len(words) % 100 == 0:
                    print("built word number " + str(len(words)) + ": " + word.kanji + "\t\tTotal suru verbs:" + str(num_suru_verbs))
            # endregion

            current_entry = ''


# endregion

# region Filtering the words list
def binary_search(words_list, text, lo=0, hi=None):
    # https://stackoverflow.com/questions/4161629/search-a-list-of-objects-in-python
    if hi is None:
        hi = len(words_list)
    while lo < hi:
        mid = (lo + hi) // 2

        if text > words_list[mid].uniqueID:
            lo = mid + 1
        elif text < words_list[mid].uniqueID:
            hi = mid
        elif text == words_list[mid].uniqueID:
            return mid
        else:
            return -1

    return -1


workingWordsList = list(words)
workingWordsList.sort(key=lambda x: x.uniqueID)
workingSuruVerbsList = list(suruVerbs)
workingSuruVerbsList.sort(key=lambda x: x.uniqueID)

print("total number of words: " + str(len(workingWordsList)))
print("total number of suru verbs: " + str(len(workingSuruVerbsList)))

typesIndex = 1
while True:
    romaji = wsLocalTypes.cell(row=typesIndex, column=3).value
    kanji = wsLocalTypes.cell(row=typesIndex, column=4).value

    if romaji == "" or not romaji:
        break

    uniqueID = romaji + UNIQUE_DELIMITER + kanji
    #print(uniqueID + "\n")

    index_of_first_hit = binary_search(workingWordsList, uniqueID)
    if index_of_first_hit != -1:
        del workingWordsList[index_of_first_hit]

    typesIndex += 1

verbsIndex = 1

print("total number of words after words filter: " + str(len(workingWordsList)))

while True:
    romaji = wsLocalVerbsForGrammar.cell(row=verbsIndex, column=3).value
    kanji = wsLocalVerbsForGrammar.cell(row=verbsIndex, column=4).value

    if romaji == "" or not romaji:
        break

    uniqueID = romaji + UNIQUE_DELIMITER + kanji
    #print(uniqueID + "\n")

    index_of_first_hit = binary_search(workingWordsList, uniqueID)
    if index_of_first_hit != -1:
        del workingWordsList[index_of_first_hit]

    index_of_first_hit = binary_search(workingSuruVerbsList, uniqueID)
    if index_of_first_hit != -1:
        del workingSuruVerbsList[index_of_first_hit]

    verbsIndex += 1

print("total number of words after verbs filter: " + str(len(workingWordsList)))

print("total number of suru verbs after filter: " + str(len(workingSuruVerbsList)))
# endregion

# region Adding all the remaining words from JMDict to the end of the types sheet

# Getting the first available row index of the Local Types list
availableTypesIndex = 1
while True:
    value = wsLocalTypes.cell(row=availableTypesIndex, column=3).value
    if not value:
        break
    availableTypesIndex += 1

# Getting the first available row index of the Local Meanings list
availableMeaningsIndex = 1
while True:
    value = wsLocalMeanings.cell(row=availableMeaningsIndex, column=1).value
    if not value:
        break
    availableMeaningsIndex += 1

# Adding the words
for word in workingWordsList:

    if (word.romaji + " suru" == workingSuruVerbsList[MAX_NUM_SURU_VERBS_TO_ADD+1].romaji) \
            and (word.kanji + "する" == workingSuruVerbsList[MAX_NUM_SURU_VERBS_TO_ADD+1].kanji):
        break

    if word.types[0] == '':
        pass

    if not (word.common or word.types[0][0] == 'V'): continue

    wsLocalTypes.cell(row=availableTypesIndex, column=3).value = word.romaji
    wsLocalTypes.cell(row=availableTypesIndex, column=4).value = word.kanji
    wsLocalTypes.cell(row=availableTypesIndex, column=6).value = ", ".join(word.altSpellings)

    meaningIndexes = []
    for i in range(len(word.types)):
        currentPartOfSpeech = word.types[i]
        meaning = word.english_meanings[i]

        if len(currentPartOfSpeech) == 0:
            print("word with missing type: " + str(word))

        elif currentPartOfSpeech[0] == 'V':
            meaningAsList = [element.strip() for element in meaning.split(",")]
            meaningAsList = [element[3:] for element in meaningAsList if element[:3] == "to "]
            meaning = ", ".join(meaningAsList)

        wsLocalMeanings.cell(row=availableMeaningsIndex, column=1).value = availableMeaningsIndex
        wsLocalMeanings.cell(row=availableMeaningsIndex, column=2).value = meaning
        wsLocalMeanings.cell(row=availableMeaningsIndex, column=3).value = currentPartOfSpeech
        wsLocalMeanings.cell(row=availableMeaningsIndex, column=9).value = 'JM'

        meaningIndexes.append(str(availableMeaningsIndex))
        availableMeaningsIndex += 1

    wsLocalTypes.cell(row=availableTypesIndex, column=5).value = ";".join(meaningIndexes)

    availableTypesIndex += 1

# Adding the suru verbs
word_addition_index = 0
for word in workingSuruVerbsList:

    if word_addition_index == MAX_NUM_SURU_VERBS_TO_ADD: break
    word_addition_index += 1

    if word.types[0] == '':
        pass

    if not (word.common or word.types[0][0] == 'V'): continue

    wsLocalTypes.cell(row=availableTypesIndex, column=3).value = word.romaji
    wsLocalTypes.cell(row=availableTypesIndex, column=4).value = word.kanji
    wsLocalTypes.cell(row=availableTypesIndex, column=6).value = ", ".join(word.altSpellings)

    meaningIndexes = []
    for i in range(len(word.types)):
        currentPartOfSpeech = word.types[i]
        meaning = word.english_meanings[i]

        if len(currentPartOfSpeech) == 0:
            print("word with missing type: " + str(word))

        wsLocalMeanings.cell(row=availableMeaningsIndex, column=1).value = availableMeaningsIndex
        wsLocalMeanings.cell(row=availableMeaningsIndex, column=2).value = meaning
        wsLocalMeanings.cell(row=availableMeaningsIndex, column=3).value = currentPartOfSpeech
        wsLocalMeanings.cell(row=availableMeaningsIndex, column=9).value = 'LOC'

        meaningIndexes.append(str(availableMeaningsIndex))
        availableMeaningsIndex += 1

    wsLocalTypes.cell(row=availableTypesIndex, column=5).value = ";".join(meaningIndexes)

    availableTypesIndex += 1


# endregion

# region Saving the results
localWordsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - updated with JMDict_english.xlsx')
# endregion
