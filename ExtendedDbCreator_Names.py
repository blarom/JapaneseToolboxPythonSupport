#!/usr/bin/python -tt

# Windows Powershell grep:
# Select-String -Path .\JMdict.xml -Pattern "CNET" -Context 4,4

import re
import openpyxl

import Globals
from Converter import Converter

def main():
    current_entry = ''

    LEGEND_DICT = {
        "surname": '0',  # "NmSu",
        "place": '1',  # "NmPl",
        "unclass": '2',  # "NmU",
        "company": '3',  # "NmC",
        "product": '4',  # "NmPr",
        "work": '5',  # "NmW",
        "masc": '6',  # "NmM",
        "fem": '7',  # "NmF",
        "person": '8',  # "NmPe",
        "given": '9',  # "NmG",
        "station": '10',  # "NmSt",
        "organization": '11',  # "NmO",
        "ok": '12',  # "NmI",
    }

    # region Preparations
    localWordsWorkbook = openpyxl.load_workbook(filename=f'{Globals.OUTPUT_DIR}/Grammar - 3000 kanji - before foreign.xlsx', data_only=True)
    print("Finished loading Grammar - 3000 kanji - before foreign.xlsx")
    namesWorkbook = openpyxl.load_workbook(filename=f'{Globals.MASTER_DIR}/Names - 3000 kanji - MASTER.xlsx', data_only=True)
    print("Finished loading Extended Words - 3000 kanji - before foreign.xlsx")
    wsLocalMeanings = localWordsWorkbook["Meanings"]
    wsLocalTypes = localWordsWorkbook["Types"]
    wsExtendedWords = namesWorkbook["Words"]
    wsExtendedWordsRomajiIndex = namesWorkbook["RomajiIndex"]
    wsNamesKanjiIndex = namesWorkbook["KanjiIndex"]


    class Word:
        def __init__(self,
                     kanji='',
                     hiragana='',
                     romaji='',
                     classifications=None,
                     englishes=None):
            if classifications is None: classifications = []
            if englishes is None: englishes = []
            self.kanji = kanji
            self.hiragana = hiragana
            self.romaji = romaji
            self.uniqueID = romaji + "zzz" + kanji
            self.classifications = classifications
            self.englishes = englishes


    converter = Converter()


    def add_index_to_dict(index_dict, index, key):
        if key in index_dict.keys():
            indexes = index_dict[key]
            if index not in indexes:
                indexes.append(index)
                index_dict[key] = indexes
        else:
            index_dict[key] = [index]


    def convert_to_utf8(text):
        return '1.' + text.encode('utf-8').hex()


    def convert_from_utf8(text):
        return bytearray.fromhex(text[2:]).decode()


    def binary_search(words_list, text, lo=0, hi=None):
        # https://stackoverflow.com/questions/4161629/search-a-list-of-objects-in-python
        if hi is None:
            hi = len(words_list)
        while lo < hi:
            mid = (lo + hi) // 2

            if text > words_list[mid].uniqueID:
                lo = mid + 1
            elif text < words_list[mid].uniqueID:
                hi = mid
            elif text == words_list[mid].uniqueID:
                return mid
            else:
                return -1

        return -1


    # endregion

    # region Filling the foreign words list
    words = []
    reached_first_entry = False
    print("Starting words list creation from JMneDict.xml")
    with open("JMneDict.xml", encoding='utf-8') as infile:
        print("Loaded JMneDict.xml")
        for line in infile:

            if "<entry>" in line:
                reached_first_entry = True
            if not reached_first_entry:
                continue

            current_entry += line

            if "</entry>" in line:

                # region Getting the basic characteristics
                kanjis = re.findall(r"<keb>(.+?)</keb>", current_entry, re.U)
                if len(kanjis) > 0:
                    kanji = kanjis[0]
                else:
                    kanji = ''

                kanas = re.findall(r"<reb>(.+?)</reb>", current_entry, re.U)
                hiragana = ''
                romaji = ''
                if len(kanas) > 0:
                    hiragana = kanas[0]
                    romaji = converter.get_latin_hiragana_katakana(hiragana)[converter.TYPE_LATIN]
                    if kanji == '': kanji = hiragana
                # endregion

                # region Fixing the senses with missing parts of speech
                trans = re.findall(r"<trans>(.+?)</trans>", current_entry, re.DOTALL)
                # endregion

                # region Getting the classifications and englishes
                classifications = []
                englishes = []
                i = 0
                while i < len(trans):

                    # region Getting the classification
                    match = re.search(r"<name_type>&(.+?);</name_type>", trans[i])
                    if match:
                        if match.group(1) in LEGEND_DICT.keys():
                            classifications.append(LEGEND_DICT[match.group(1)])
                        else:
                            classifications.append(match.group(1))
                    # endregion

                    # region Getting the english
                    match = re.search(r"<trans_det>(.+?)</trans_det>", trans[i])
                    if match: englishes.append(match.group(1).replace('"', '$$$$').replace('#', '@@@@').replace('&amp;', '&'))
                    # endregion

                    i += 1

                # endregion

                # region Creating the words
                if kanji != '' and hiragana != '' and '*' not in romaji:

                    word = Word(kanji,
                                hiragana.replace('・', ''),
                                romaji.replace('・', ''),
                                classifications,
                                [])  # Setting englishes to [] to reduce db size and load time, since the main need is for romaji<->kanji interpretation

                    if not [word.romaji, word.kanji] in Globals.EDICT_EXCEPTIONS and not ["*", word.kanji] in Globals.EDICT_EXCEPTIONS:
                        words.append(word)

                    if len(words) % 500 == 0:
                        print("built word number " + str(len(words)) + ": " + word.kanji)
                # endregion

                current_entry = ''

    # endregion

    # region Creating the new words list
    print("Creating the new words list")
    localWordsList = []
    jmDictWordsList = list(words)
    jmDictWordsList.sort(key=lambda x: x.uniqueID)
    sheetNum_types = 0

    typesIndex = 2
    while True:
        # region Row skip conditions
        romaji = wsLocalTypes.cell(row=typesIndex, column=Globals.TYPES_COL_ROMAJI).value
        kanji = wsLocalTypes.cell(row=typesIndex, column=Globals.TYPES_COL_KANJI).value

        if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
            break
        # endregion

        romaji = romaji.replace(' ', '')
        kanji = kanji.replace('～', '')
        localWordsList.append(Word(kanji=kanji, romaji=romaji))

        typesIndex += 1

    newWords = []
    localWordsList.sort(key=lambda x: x.uniqueID)
    for word in jmDictWordsList:
        index_of_first_hit = binary_search(localWordsList, word.uniqueID)
        if index_of_first_hit == -1:
            newWords.append(word)

    # endregion

    # region Creating the Extended Words sheet and index dicts
    print("Creating the Extended Words sheet and index dicts")
    wsExtendedWords.cell(row=1, column=Globals.EXT_WORD_COL_INDEX).value = "Index"
    wsExtendedWords.cell(row=1, column=Globals.EXT_WORD_COL_ROMAJI).value = "romaji"
    wsExtendedWords.cell(row=1, column=Globals.EXT_WORD_COL_KANJI).value = "kanji"
    wsExtendedWords.cell(row=1, column=Globals.EXT_WORD_COL_POS).value = "classifications"
    wsExtendedWords.cell(row=1, column=Globals.EXT_WORD_COL_MEANINGS_EN).value = "englishes"
    wsExtendedWordsCSV_rows = ['|'.join(["Index", "romaji", "kanji", "classifications"]) + '|']
    row_index = 2
    romaji_index_dict = {}
    english_index_dict = {}
    kanji_index_dict = {}
    for word in newWords:
        if word.romaji == '' or word.kanji == '': continue

        classifications_final = '#'.join(word.classifications)
        englishes_final = '#'.join(word.englishes)

        wsExtendedWords.cell(row=row_index, column=Globals.NAMES_COL_INDEX).value = row_index
        wsExtendedWords.cell(row=row_index, column=Globals.NAMES_COL_ROMAJI).value = word.romaji
        wsExtendedWords.cell(row=row_index, column=Globals.NAMES_COL_KANJI).value = word.kanji
        wsExtendedWords.cell(row=row_index, column=Globals.NAMES_COL_CLASSIFICATION).value = classifications_final
        wsExtendedWords.cell(row=row_index, column=Globals.NAMES_COL_CLASSIFICATION).value = englishes_final

        wsExtendedWordsCSV_rows.append('|'.join([str(row_index), word.romaji, word.kanji, classifications_final]) + '|')

        add_index_to_dict(romaji_index_dict, str(row_index), word.romaji)
        add_index_to_dict(kanji_index_dict, str(row_index), convert_to_utf8(word.kanji))

        row_index += 1

    # endregion

    # region Creating the indexes
    print("Creating the indexes")
    wsExtendedWordsKanjiIndexCSV_rows = ['value|word_ids|']
    wsNamesRomajiIndexCSV_rows = ['value|word_ids|']
    row_index = 1
    index_dict_keys_sorted = list(romaji_index_dict.keys())
    index_dict_keys_sorted.sort()
    for key in index_dict_keys_sorted:
        wsExtendedWordsRomajiIndex.cell(row=row_index, column=1).value = key
        wsExtendedWordsRomajiIndex.cell(row=row_index, column=2).value = ';'.join(romaji_index_dict[key])
        wsNamesRomajiIndexCSV_rows.append(key + '|' + wsExtendedWordsRomajiIndex.cell(row=row_index, column=2).value + '|')
        row_index += 1

    row_index = 1
    index_dict_keys_sorted = list(kanji_index_dict.keys())
    index_dict_keys_sorted.sort()
    for key in index_dict_keys_sorted:
        wsNamesKanjiIndex.cell(row=row_index, column=1).value = convert_from_utf8(key)
        wsNamesKanjiIndex.cell(row=row_index, column=2).value = ';'.join(kanji_index_dict[key])
        wsNamesKanjiIndex.cell(row=row_index, column=3).value = key
        wsExtendedWordsKanjiIndexCSV_rows.append(wsNamesKanjiIndex.cell(row=row_index, column=1).value + '|'
                                                 + wsNamesKanjiIndex.cell(row=row_index, column=2).value + '|')
        row_index += 1
    # endregion

    # Saving the results
    print("Saving the results")
    namesWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Names - 3000 kanji.xlsx')
    base = f'{Globals.JAPAGRAM_ASSETS_DIR}/LineNamesDb - '
    # base = 'C:/Projects/Workspace/Web/JT database'
    with open(base + 'Words.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsCSV_rows))
    with open(base + 'KanjiIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsKanjiIndexCSV_rows))
    with open(base + 'RomajiIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsNamesRomajiIndexCSV_rows))
