#!/usr/bin/python -tt

#Windows Powershell grep:
#Select-String -Path .\JMdict.xml -Pattern "CNET" -Context 4,4

import re
import openpyxl

import Constants
from Converter import Converter

current_entry = ''

VERB_PARTS_OF_SPEECH = ['Vrui', 'Vbu', 'Vgu', 'Vku', 'Vmu', 'Vnu', 'Vrug', 'Vsu', 'Vtsu', 'Vu', 'Vi', 'Viku', 'Vus', 'Varu', 'Vkuru', 'Vsuru']
LEGEND_DICT = {
    'MA': 'ZM',
    'X': 'IES',
    'abbr': 'Abr',
    'adj-i': 'Ai',
    'adj-ix': 'Ai',
    'adj-na': 'Ana',
    'adj-no': 'Ano',
    'adj-pn': 'Apn',
    'adj-t': 'Atr',
    'adj-f': 'Af',
    'adv': 'A',
    'adv-to': 'Ato',
    'arch': 'ar',
    'ateji': '',
    'aux': '',
    'aux-v': 'Vx',
    'aux-adj': 'Ax',
    'Buddh': 'ZB',
    'chem': 'ZC',
    'chn': 'ZCL',
    'col': 'coq',
    'comp': 'ZI',
    'conj': 'CO',
    'cop-da': '',
    'ctr': 'C',
    'derog': 'Dr',
    'eK': '',
    'ek': '',
    'exp': 'CE',
    'fam': 'LF',
    'fem': 'LFt',
    'food': 'ZF',
    'geom': 'ZG',
    'gikun': '',
    'hon': 'LHn',
    'hum': 'LHm',
    'iK': '',
    'id': 'idp',
    'ik': '',
    'int': 'OI',
    'io': '',
    'iv': '',
    'ling': 'ZL',
    'm-sl': 'ZMg',
    'male': 'LMt',
    'male-sl': 'LMt',
    'math': 'ZMt',
    'mil': 'ZMl',
    'n': 'N',
    'n-adv': 'NAdv',
    'n-suf': 'N;Sx',
    'n-pref': 'N;Px',
    'n-t': 'T',
    'num': 'num',
    'oK': '',
    'obs': 'Obs',
    'obsc': 'Obs',
    'ok': 'Obs',
    'oik': 'Obs',
    'on-mim': 'OI',
    'pn': 'P',
    'poet': 'ZP',
    'pol': 'LHn',
    'pref': 'Px',
    'proverb': 'CES',
    'prt': 'PP',
    'physics': 'ZPh',
    'quote': '',
    'rare': 'Obs',
    'sens': '',
    'sl': 'Sl',
    'suf': 'Sx',
    'uK': '',
    'uk': '',
    'unc': 'UNC',
    'yoji': '',
    'v1': 'Vrui',
    'v1-s': 'Vrui',
    'v2a-s': '',
    'v4h': '',
    'v4r': '',
    'v5aru': 'Varu',
    'v5b': 'Vbu',
    'v5g': 'Vgu',
    'v5k': 'Vku',
    'v5k-s': 'Viku',
    'v5m': 'Vmu',
    'v5n': 'Vnu',
    'v5r': 'Vrug',
    'v5r-i': 'Vrug',
    'v5s': 'Vsu',
    'v5t': 'Vtsu',
    'v5u': 'Vu',
    'v5u-s': 'Vus',
    'v5uru': '',
    'vz': 'Vrui',
    'vi': 'intransitive',
    'vk': 'Vkuru',
    'vn': '',
    'vr': '',
    'vs': 'Vsuru',
    'vs-c': '',
    'vs-s': 'Vsuru',
    'vs-i': 'Vsuru',
    'kyb': '',
    'osb': '',
    'ksb': '',
    'ktb': '',
    'tsb': '',
    'thb': '',
    'tsug': '',
    'kyu': '',
    'rkb': '',
    'nab': '',
    'hob': '',
    'vt': 'transitive',
    'vulg': 'vul',
    'adj-kari': '',
    'adj-ku': '',
    'adj-shiku': '',
    'adj-nari': '',
    'n-pr': 'Ne',
    'v-unspec': '',
    'v4k': '',
    'v4g': '',
    'v4s': '',
    'v4t': '',
    'v4n': '',
    'v4b': '',
    'v4m': '',
    'v2k-k': '',
    'v2g-k': '',
    'v2t-k': '',
    'v2d-k': '',
    'v2h-k': '',
    'v2b-k': '',
    'v2m-k': '',
    'v2y-k': '',
    'v2r-k': '',
    'v2k-s': '',
    'v2g-s': '',
    'v2s-s': '',
    'v2z-s': '',
    'v2t-s': '',
    'v2d-s': '',
    'v2n-s': '',
    'v2h-s': '',
    'v2b-s': '',
    'v2m-s': '',
    'v2y-s': '',
    'v2r-s': '',
    'v2w-s': '',
    'archit': 'ZAc',
    'astron': 'ZAs',
    'baseb': 'ZBb',
    'biol': 'ZBi',
    'bot': 'ZBt',
    'bus': 'ZBs',
    'econ': 'ZEc',
    'engr': 'ZEg',
    'finc': 'ZFn',
    'geol': 'ZGg',
    'law': 'ZLw',
    'mahj': 'ZMj',
    'med': 'Md',
    'music': 'ZMc',
    'Shinto': 'ZSt',
    'shogi': 'ZSg',
    'sports': 'ZSp',
    'sumo': 'ZSm',
    'zool': 'ZZ',
    'joc': 'ZH',
    'anat': 'ZAn'
}

# region Preparations
localWordsWorkbook = openpyxl.load_workbook(filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji - before foreign.xlsx', data_only=True)
print("Finished loading Grammar - 3000 kanji - before foreign.xlsx")
localVerbsWorkbook = openpyxl.load_workbook(filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - before foreign.xlsx', data_only=True)
print("Finished loading Verbs - 3000 kanji - before foreign.xlsx")
extendedWordsWorkbook = openpyxl.load_workbook(filename='C:/Users/Bar/Dropbox/Japanese/Extended Words - 3000 kanji.xlsx', data_only=True)
print("Finished loading Extended Words - 3000 kanji - before foreign.xlsx")
wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalMeaningsFR = localWordsWorkbook["MeaningsFR"]
wsLocalMeaningsES = localWordsWorkbook["MeaningsES"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]
wsExtendedWords = extendedWordsWorkbook["Words"]
wsExtendedWordsRomajiIndex = extendedWordsWorkbook["RomajiIndex"]
wsExtendedWordsEnglishIndex = extendedWordsWorkbook["EnglishIndex"]
wsExtendedWordsFrenchIndex = extendedWordsWorkbook["FrenchIndex"]
wsExtendedWordsSpanishIndex = extendedWordsWorkbook["SpanishIndex"]
wsExtendedWordsKanjiIndex = extendedWordsWorkbook["KanjiIndex"]


class Word:
    def __init__(self,
                 kanji='',
                 altSpellings=None,
                 hiragana='',
                 romaji='',
                 english_meanings=None,
                 french_meanings=None,
                 spanish_meanings=None,
                 types_for_word=None,
                 common=False):
        if altSpellings is None: altSpellings = []
        if types_for_word is None: types_for_word = []
        if english_meanings is None: english_meanings = []
        if french_meanings is None: french_meanings = []
        if spanish_meanings is None: spanish_meanings = []
        self.kanji = kanji
        self.altSpellings = altSpellings
        self.hiragana = hiragana
        self.romaji = romaji
        self.uniqueID = romaji + "zzz" + kanji
        self.english_meanings = english_meanings
        self.french_meanings = french_meanings
        self.spanish_meanings = spanish_meanings
        self.types_for_word = types_for_word
        self.common = common


converter = Converter()


def add_index_to_dict(index_dict, index, key):
    if key in index_dict.keys():
        indexes = index_dict[key]
        if index not in indexes:
            indexes.append(index)
            index_dict[key] = indexes
    else:
        index_dict[key] = [index]


def convert_to_utf8(text):
    return '1.' + text.encode('utf-8').hex()


def convert_from_utf8(text):
    return bytearray.fromhex(text[2:]).decode()


def binary_search(words_list, text, lo=0, hi=None):
    # https://stackoverflow.com/questions/4161629/search-a-list-of-objects-in-python
    if hi is None:
        hi = len(words_list)
    while lo < hi:
        mid = (lo + hi) // 2

        if text > words_list[mid].uniqueID:
            lo = mid + 1
        elif text < words_list[mid].uniqueID:
            hi = mid
        elif text == words_list[mid].uniqueID:
            return mid
        else:
            return -1

    return -1


# endregion

# region Filling the foreign words list
words = []
reached_first_entry = False
print("Starting words list creation from JMDict.xml")
with open("JMDict.xml", encoding='utf-8') as infile:
    print("Loaded JMDict.xml")
    for line in infile:

        if "<entry>" in line:
            reached_first_entry = True
        if not reached_first_entry:
            continue

        current_entry += line

        if "</entry>" in line:

            # region Getting the basic characteristics
            kanjis = re.findall(r"<keb>(.+?)</keb>", current_entry, re.U)
            if len(kanjis) > 0:
                kanji = kanjis[0]
                altSpellings = []
                for i in range(1, len(kanjis)):
                    altSpellings.append(kanjis[i].replace('・', ''))
            else:
                altSpellings = []
                kanji = ''

            kanas = re.findall(r"<reb>(.+?)</reb>", current_entry, re.U)
            hiragana = ''
            romaji = ''
            if len(kanas) > 0:
                hiragana = kanas[0]
                romaji = converter.get_latin_hiragana_katakana(hiragana)[converter.TYPE_LATIN]
                if kanji == '': kanji = hiragana

                for i in range(1, len(kanas)):
                    altSpellingRomaji = converter.get_latin_hiragana_katakana(kanas[i])[converter.TYPE_LATIN]
                    if altSpellingRomaji != romaji:
                        altSpellings.append(altSpellingRomaji.replace('・', ''))
            # endregion

            # region Getting the common status - if not common, skip this word
            match = re.search(r"<ke_pri>(news1|ichi1)</ke_pri>", current_entry)
            # match = re.search(r"<ke_pri>(news1|news2|ichi1|ichi2|spec1|spec2|gai1|gai2)</ke_pri>", current_entry)
            if match:
                common = True
                current_entry = ''
                continue
            else:
                common = False
            # endregion

            # region Fixing the senses with missing parts of speech
            senses = re.findall(r"<sense>(.+?)</sense>", current_entry, re.DOTALL)
            for i in range(len(senses) - 1, 0, -1):
                posTypes = re.findall(r"<pos>&(.+?);</pos>", senses[i])
                if len(posTypes) == 0:
                    for j in range(i - 1, -1, -1):
                        posTypesPrev = re.findall(r"<pos>&(.+?);</pos>", senses[j])
                        if len(posTypesPrev) > 0:
                            for posTypePrev in posTypesPrev:
                                senses[i] = r"<pos>&" + posTypePrev + ";</pos>" + senses[i]
                            break
            # endregion

            # region Getting the types/meanings and suru verb characteristics
            types = []
            english_meanings = []
            types_for_word = []
            english_meanings_for_word = []
            types_for_suru_verb = []
            english_meanings_for_suru_verb = []
            suru_altSpellings = []
            suru_hiragana = ''
            suru_romaji = ''
            suru_kanji = ''
            types_as_string = '-'
            partsOfSpeechFromLastSense = []
            last_sense_is_result_of_sense_split = False
            create_suru_verb = False
            i = 0
            while i < len(senses):

                # region Getting the list of JMtypes
                matches = re.findall(r"<(pos|field|misc)>&(.+?);</(pos|field|misc)>", senses[i])
                JMtypes = [match[1] for match in matches]
                partsOfSpeech = ';'.join([LEGEND_DICT[JMtype] for JMtype in JMtypes if LEGEND_DICT[JMtype] != '']).split(';')
                partsOfSpeech = list(set(partsOfSpeech))
                # endregion

                # region Getting the TRANSITIVE status
                trans = 'I'
                if "transitive" in partsOfSpeech and "intransitive" in partsOfSpeech:
                    senses.append(senses[i].replace("<pos>&vt;</pos>", ""))
                    trans = 'T'
                    partsOfSpeech.remove("transitive")
                    partsOfSpeech.remove("intransitive")
                elif "transitive" in partsOfSpeech:
                    trans = 'T'
                    partsOfSpeech.remove("transitive")
                elif "intransitive" in partsOfSpeech:
                    trans = 'I'
                    partsOfSpeech.remove("intransitive")

                # endregion

                # region Adding T/I to the part of speech if it is a verb
                type_is_verb = False
                is_suru_verb = False
                for j in range(len(partsOfSpeech)):
                    if partsOfSpeech[j] in VERB_PARTS_OF_SPEECH:
                        partsOfSpeech[j] = partsOfSpeech[j] + trans
                        type_is_verb = True
                        if 'Vsuru' in partsOfSpeech[j]: is_suru_verb = True
                # endregion

                # region Removing part of speech duplicates
                partsOfSpeech = list(set(partsOfSpeech))
                # endregion

                # region Reordering parts of speech
                if "Ai" in partsOfSpeech:
                    partsOfSpeech.remove("Ai")
                    partsOfSpeech.insert(0, "Ai")

                if "Ana" in partsOfSpeech:
                    partsOfSpeech.remove("Ana")
                    partsOfSpeech.insert(0, "Ana")

                if "NAdv" in partsOfSpeech:
                    partsOfSpeech.remove("NAdv")
                    partsOfSpeech.insert(0, "NAdv")

                if "Ano" in partsOfSpeech:
                    partsOfSpeech.remove("Ano")
                    partsOfSpeech.insert(0, "Ano")

                if "A" in partsOfSpeech:
                    partsOfSpeech.remove("A")
                    partsOfSpeech.insert(0, "A")

                if "N" in partsOfSpeech:
                    partsOfSpeech.remove("N")
                    partsOfSpeech.insert(0, "N")

                verbPOS = [element for element in partsOfSpeech if element[:-1] in VERB_PARTS_OF_SPEECH]
                if len(verbPOS) > 0:
                    partsOfSpeech.remove(verbPOS[0])
                    partsOfSpeech.insert(0, verbPOS[0])
                # endregion

                # region Getting the english meaning
                matches = re.findall(r"<(gloss|gloss g_type=\"expl\")?>(.+?)</gloss>", senses[i], re.U)
                english_meanings = [match[1] for match in matches]
                english_meanings_as_string = ", ".join(english_meanings).replace("&,", "&")
                # endregion

                # region Setting the word characteristics
                types_for_word.append(';'.join(partsOfSpeech))
                english_meanings_for_word.append(english_meanings_as_string)
                # endregion

                i += 1

            # endregion

            # Getting the foreign meanings
            french_meanings = []
            spanish_meanings = []
            for sense in senses:
                french_meanings_in_sense = re.findall(r"<gloss xml:lang=\"fre\">(.+?)</gloss>", sense, re.MULTILINE | re.DOTALL)
                if len(french_meanings_in_sense) > 0: french_meanings.append(", ".join(french_meanings_in_sense))
                spanish_meanings_in_sense = re.findall(r"<gloss xml:lang=\"spa\">(.+?)</gloss>", sense, re.MULTILINE | re.DOTALL)
                if len(spanish_meanings_in_sense) > 0: spanish_meanings.append(", ".join(spanish_meanings_in_sense))
            # endregion

            # Cleaning the POS and meanings
            tempPOS = types_for_word
            tempMeanings = english_meanings_for_word
            types_for_word = []
            english_meanings_for_word = []
            for i in range(0, len(tempMeanings)):
                if tempMeanings[i] != '':
                    types_for_word.append(tempPOS[i])
                    english_meanings_for_word.append(tempMeanings[i]
                                                     .replace('"','$$$$')
                                                     .replace('#','@@@@')
                                                     .replace('&amp;','&')
                                                     .replace('&gt;','>')
                                                     .replace('&lt;','<'))
            tempMeanings = french_meanings_in_sense
            french_meanings_in_sense = []
            for i in range(0, len(tempMeanings)):
                if tempMeanings[i] != '':
                    french_meanings_in_sense.append(tempMeanings[i]
                                                     .replace('"','$$$$')
                                                     .replace('#','@@@@')
                                                     .replace('&amp;','&')
                                                     .replace('&gt;','>')
                                                     .replace('&lt;','<'))
            tempMeanings = spanish_meanings_in_sense
            spanish_meanings_in_sense = []
            for i in range(0, len(tempMeanings)):
                if tempMeanings[i] != '':
                    spanish_meanings_in_sense.append(tempMeanings[i]
                                                     .replace('"','$$$$')
                                                     .replace('#','@@@@')
                                                     .replace('&amp;','&')
                                                     .replace('&gt;','>')
                                                     .replace('&lt;','<'))
            # endregion

            # region Creating the words
            if kanji != '' and hiragana != '' \
                    and len(english_meanings_for_word) > 0 \
                    and '*' not in romaji:

                word = Word(kanji,
                            altSpellings,
                            hiragana.replace('・', ''),
                            romaji.replace('・', ''),
                            english_meanings_for_word,
                            french_meanings,
                            spanish_meanings,
                            types_for_word,
                            common)

                if not [word.romaji, word.kanji] in Constants.EDICT_EXCEPTIONS and not ["*", word.kanji] in Constants.EDICT_EXCEPTIONS:
                    words.append(word)

                if len(words) % 100 == 0:
                    print("built word number " + str(len(words)) + ": " + word.kanji)
            # endregion

            current_entry = ''

# endregion

# region Creating the new words list
print("Creating the new words list")
localWordsList = []
jmDictWordsList = list(words)
jmDictWordsList.sort(key=lambda x: x.uniqueID)
sheetNum_types = 0
sheetNum_grammar = 1
sheetNum_verbs = 2

for sheetNum in range(0, 3):

    if sheetNum == sheetNum_types:
        typesIndex = 2
        workingWorksheet = wsLocalTypes
    elif sheetNum == sheetNum_grammar:
        typesIndex = 2
        workingWorksheet = wsLocalGrammar
    elif sheetNum == sheetNum_verbs:
        typesIndex = 4
        workingWorksheet = wsLocalVerbs
    else:
        typesIndex = 2
        workingWorksheet = wsLocalTypes

    while True:

        # region Row skip conditions
        romaji = ""
        kanji = ""
        if sheetNum == sheetNum_types or sheetNum == sheetNum_grammar:
            romaji = workingWorksheet.cell(row=typesIndex, column=Constants.TYPES_COL_ROMAJI).value
            kanji = workingWorksheet.cell(row=typesIndex, column=Constants.TYPES_COL_KANJI).value

            if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
                break

        elif sheetNum == sheetNum_verbs:
            romaji = workingWorksheet.cell(row=typesIndex, column=Constants.VERBS_COL_ROMAJI).value
            kanji = workingWorksheet.cell(row=typesIndex, column=Constants.VERBS_COL_KANJI).value

            if (workingWorksheet.cell(row=typesIndex, column=1).value == '-'
                    and workingWorksheet.cell(row=typesIndex, column=2).value == '-'
                    and workingWorksheet.cell(row=typesIndex, column=3).value == '-'):
                break

            if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
                typesIndex += 1
                continue
        # endregion

        romaji = romaji.replace(' ', '')
        kanji = kanji.replace('～','')
        localWordsList.append(Word(kanji=kanji, romaji=romaji))

        typesIndex += 1

newWords = []
localWordsList.sort(key=lambda x: x.uniqueID)
for word in jmDictWordsList:
    index_of_first_hit = binary_search(localWordsList, word.uniqueID)
    if index_of_first_hit == -1:
        newWords.append(word)

# endregion

# region Creating the Extended Words sheet and index dicts
print("Creating the Extended Words sheet and index dicts")
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_INDEX).value = "Index"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_ROMAJI).value = "romaji"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_KANJI).value = "kanji"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_POS).value = "POS"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_ALTS).value = "altSpellings"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_MEANINGS_EN).value = "meaningsEN"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_MEANINGS_FR).value = "meaningsFR"
wsExtendedWords.cell(row=1, column=Constants.EXT_WORD_COL_MEANINGS_ES).value = "meaningsES"
wsExtendedWordsCSV_rows = ['|'.join(["Index", "romaji", "kanji", "POS", "altSpellings", "meaningsEN", "meaningsFR", "meaningsES"]) + '|']
row_index = 2
romaji_index_dict = {}
english_index_dict = {}
french_index_dict = {}
spanish_index_dict = {}
kanji_index_dict = {}
for word in newWords:
    if word.romaji == '' or word.kanji == '': continue

    types = '#'.join(word.types_for_word)
    altS = '#'.join(word.altSpellings)
    meaningsEN = '#'.join(word.english_meanings).replace('|','-')
    meaningsFR = '#'.join(word.french_meanings).replace('|','-')
    meaningsES = '#'.join(word.spanish_meanings).replace('|','-')
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_INDEX).value = row_index
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_ROMAJI).value = word.romaji
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_KANJI).value = word.kanji
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_POS).value = types
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_ALTS).value = altS
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_MEANINGS_EN).value = meaningsEN
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_MEANINGS_FR).value = meaningsFR
    wsExtendedWords.cell(row=row_index, column=Constants.EXT_WORD_COL_MEANINGS_ES).value = meaningsES

    wsExtendedWordsCSV_rows.append('|'.join([str(row_index), word.romaji, word.kanji, types, altS, meaningsEN, meaningsFR, meaningsES]) + '|')

    add_index_to_dict(romaji_index_dict, str(row_index), word.romaji)
    add_index_to_dict(kanji_index_dict, str(row_index), convert_to_utf8(word.kanji))

    for altSpelling in word.altSpellings:
        if altSpelling == '': continue
        match = re.search(r'[a-zA-Z\']', altSpelling)
        if match:
            add_index_to_dict(romaji_index_dict, str(row_index), altSpelling.lower())
        else:
            add_index_to_dict(kanji_index_dict, str(row_index), convert_to_utf8(altSpelling))

    for meaning in word.english_meanings:
        clean_meaning = meaning
        clean_meaning = clean_meaning.replace('ie.', '')
        clean_meaning = clean_meaning.replace('i.e.', '')
        clean_meaning = clean_meaning.replace('e.g.', '')
        clean_meaning = re.sub(r'[^\w]', ' ', clean_meaning)
        clean_meaning = re.sub(' +', ' ', clean_meaning)
        for full_word in clean_meaning.split(' '):
            if len(full_word) > 2:
                add_index_to_dict(english_index_dict, str(row_index), full_word.lower())

    for meaning in word.french_meanings:
        clean_meaning = meaning
        clean_meaning = clean_meaning.replace('c.-à-d.', '')
        clean_meaning = clean_meaning.replace('par ex.', '')
        clean_meaning = clean_meaning.replace('ex.', '')
        clean_meaning = clean_meaning.replace(',', ' ')
        clean_meaning = clean_meaning.replace('œ', 'oe')
        clean_meaning = re.sub(r'[^\w]', ' ', clean_meaning)
        clean_meaning = re.sub(' +', ' ', clean_meaning)
        for full_word in clean_meaning.split(' '):
            if len(full_word) > 2:
                add_index_to_dict(french_index_dict, str(row_index), full_word.lower())

    for meaning in word.spanish_meanings:
        clean_meaning = meaning
        clean_meaning = clean_meaning.replace('es decir', '')
        clean_meaning = clean_meaning.replace('por ej.', '')
        clean_meaning = clean_meaning.replace('ej.', '')
        clean_meaning = re.sub(r'[^\ws]', ' ', clean_meaning)
        clean_meaning = re.sub(' +', ' ', clean_meaning)
        for full_word in clean_meaning.split(' '):
            if len(full_word) > 2:
                add_index_to_dict(spanish_index_dict, str(row_index), full_word.lower())

    row_index += 1

# endregion

# region Creating the indexes
print("Creating the indexes")
wsExtendedWordsKanjiIndexCSV_rows = ['value|word_ids|']
wsExtendedWordsRomajiIndexCSV_rows = ['value|word_ids|']
wsExtendedWordsEnglishIndexCSV_rows = ['value|word_ids|']
wsExtendedWordsFrenchIndexCSV_rows = ['value|word_ids|']
wsExtendedWordsSpanishIndexCSV_rows = ['value|word_ids|']
row_index = 1
index_dict_keys_sorted = list(romaji_index_dict.keys())
index_dict_keys_sorted.sort()
for key in index_dict_keys_sorted:
    wsExtendedWordsRomajiIndex.cell(row=row_index, column=1).value = key
    wsExtendedWordsRomajiIndex.cell(row=row_index, column=2).value = ';'.join(romaji_index_dict[key])
    wsExtendedWordsRomajiIndexCSV_rows.append(key + '|' + wsExtendedWordsRomajiIndex.cell(row=row_index, column=2).value + '|')
    row_index += 1

row_index = 1
index_dict_keys_sorted = list(english_index_dict.keys())
index_dict_keys_sorted.sort()
for key in index_dict_keys_sorted:
    wsExtendedWordsEnglishIndex.cell(row=row_index, column=1).value = key
    wsExtendedWordsEnglishIndex.cell(row=row_index, column=2).value = ';'.join(english_index_dict[key])
    wsExtendedWordsEnglishIndexCSV_rows.append(key + '|' + wsExtendedWordsEnglishIndex.cell(row=row_index, column=2).value + '|')
    row_index += 1

row_index = 1
index_dict_keys_sorted = list(french_index_dict.keys())
index_dict_keys_sorted.sort()
for key in index_dict_keys_sorted:
    wsExtendedWordsFrenchIndex.cell(row=row_index, column=1).value = key
    wsExtendedWordsFrenchIndex.cell(row=row_index, column=2).value = ';'.join(french_index_dict[key])
    wsExtendedWordsFrenchIndexCSV_rows.append(key + '|' + wsExtendedWordsFrenchIndex.cell(row=row_index, column=2).value + '|')
    row_index += 1

row_index = 1
index_dict_keys_sorted = list(spanish_index_dict.keys())
index_dict_keys_sorted.sort()
for key in index_dict_keys_sorted:
    wsExtendedWordsSpanishIndex.cell(row=row_index, column=1).value = key
    wsExtendedWordsSpanishIndex.cell(row=row_index, column=2).value = ';'.join(spanish_index_dict[key])
    wsExtendedWordsSpanishIndexCSV_rows.append(key + '|' + wsExtendedWordsSpanishIndex.cell(row=row_index, column=2).value + '|')
    row_index += 1

row_index = 1
index_dict_keys_sorted = list(kanji_index_dict.keys())
index_dict_keys_sorted.sort()
for key in index_dict_keys_sorted:
    wsExtendedWordsKanjiIndex.cell(row=row_index, column=1).value = convert_from_utf8(key)
    wsExtendedWordsKanjiIndex.cell(row=row_index, column=2).value = ';'.join(kanji_index_dict[key])
    wsExtendedWordsKanjiIndex.cell(row=row_index, column=3).value = key
    wsExtendedWordsKanjiIndexCSV_rows.append(wsExtendedWordsKanjiIndex.cell(row=row_index, column=1).value + '|'
                                             + wsExtendedWordsKanjiIndex.cell(row=row_index, column=2).value + '|')
    row_index += 1
# endregion

# Saving the results
print("Saving the results")
extendedWordsWorkbook.save(filename='C:/Users/Bar/Dropbox/Japanese/Extended Words - 3000 kanji.xlsx')
base = 'C:/Projects/Workspace/Japagram/app/src/main/assets/LineExtendedDb - '
with open(base + 'Words.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsCSV_rows))
with open(base + 'KanjiIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsKanjiIndexCSV_rows))
with open(base + 'RomajiIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsRomajiIndexCSV_rows))
with open(base + 'EnglishIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsEnglishIndexCSV_rows))
with open(base + 'FrenchIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsFrenchIndexCSV_rows))
with open(base + 'SpanishIndex.csv', 'w', encoding='utf-8') as f_out: f_out.write('\n'.join(wsExtendedWordsSpanishIndexCSV_rows))
