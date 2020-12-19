#!/usr/bin/python -tt
import os
import re
import openpyxl
import Globals
import ExtendedDbCreator
import ExtendedDbCreator_Names
import JMDictForeignMeaningsUpdater
import Converter
from Globals import *

prepare_foreign_meanings = False
prepare_grammar_db = False
prepare_extended_db = False
prepare_conj_lengths = False
prepare_names_db = False
prepare_kanji_db = False
prepare_conj_db = False  # Not used, resulting db is too big to be useful
prepare_frequency_db = False
update_workbooks = False
prepare_db_for_release = False

prepare_foreign_meanings = True
prepare_grammar_db = True
prepare_extended_db = True
prepare_conj_lengths = True
prepare_names_db = False
prepare_kanji_db = False
prepare_conj_db = False  # Not used, resulting db is too big to be useful
prepare_frequency_db = False
update_workbooks = True

prepare_db_for_release = True
if prepare_db_for_release:
    prepare_foreign_meanings = True
    prepare_grammar_db = True
    prepare_extended_db = True
    prepare_conj_lengths = True
    prepare_names_db = False
    prepare_kanji_db = False
    prepare_frequency_db = False
    update_workbooks = True

if prepare_foreign_meanings:
    JMDictForeignMeaningsUpdater.main()

if prepare_frequency_db:
    # region Reading worksheets
    FrequencyWorkbook = openpyxl.load_workbook(filename=f'{Globals.MASTER_DIR}/Frequencies.xlsx', data_only=True)
    print("Finished loading Frequencies.xlsx")
    wsFrequencyWords = FrequencyWorkbook["Words"]
    Globals.clearSheet(FrequencyWorkbook, "WordsIndexed")
    wsFrequencyWordsIndexed = FrequencyWorkbook["WordsIndexed"]
    # endregion

    # region Sorting the words
    words = {}
    row = 2
    while not Globals.isLastRow(wsFrequencyWords, row):
        kanji = wsFrequencyWords.cell(row=row, column=2).value
        words[kanji] = wsFrequencyWords.cell(row=row, column=1).value
        row += 1

    sortedKanjis = sorted(words.keys(), key=lambda x: convert_to_utf8(x))
    row = 1
    for item in sortedKanjis:
        wsFrequencyWordsIndexed.cell(row=row, column=1).value = item
        wsFrequencyWordsIndexed.cell(row=row, column=2).value = words[item]
        row += 1
    # endregion

    # region Saving the results to xlsx & csv
    if update_workbooks:
        FrequencyWorkbook.save(filename=f'{Globals.MASTER_DIR}/Frequencies.xlsx')

    Globals.create_csv_from_worksheet(wsFrequencyWordsIndexed, name("FrequenciesIndexed"), idx("A"), idx("B"), False, 1)
    # endregion

if prepare_grammar_db:
    # region Reading worksheets
    content_freq = Globals.get_file_contents(os.path.join(Globals.JAPAGRAM_ASSETS_DIR, 'LineFrequencies - 3000 kanji.csv')).split('\n')
    freq_dict = {}
    for i in range(len(content_freq)):
        if content_freq[i].strip() not in freq_dict.keys(): freq_dict[content_freq[i].strip()] = i + 1

    GrammarWorkbook = openpyxl.load_workbook(filename=f'{Globals.OUTPUT_DIR}/Grammar - 3000 kanji - with foreign.xlsx', data_only=True)
    print("Finished loading Grammar - 3000 kanji.xlsx")

    wsMeaningsEN = GrammarWorkbook["Meanings"]
    wsMeaningsFR = GrammarWorkbook["MeaningsFR"]
    wsMeaningsES = GrammarWorkbook["MeaningsES"]
    wsMultExplEN = GrammarWorkbook["MultExplEN"]
    wsMultExplFR = GrammarWorkbook["MultExplFR"]
    wsMultExplES = GrammarWorkbook["MultExplES"]
    wsExamples = GrammarWorkbook["Examples"]
    wsTypes = GrammarWorkbook["Types"]
    wsGrammar = GrammarWorkbook["Grammar"]

    VerbsWorkbook = openpyxl.load_workbook(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - with foreign.xlsx', data_only=True)
    print("Finished loading Verbs - 3000 kanji.xlsx")
    wsVerbsForGrammar = VerbsWorkbook["VerbsForGrammar"]
    wsVerbs = VerbsWorkbook["Verbs"]

    Globals.clearSheet(GrammarWorkbook, "SortedIndexKanji")
    Globals.clearSheet(GrammarWorkbook, "SortedIndexRomaji")
    Globals.clearSheet(GrammarWorkbook, "SortedIndexLatinEN")
    Globals.clearSheet(GrammarWorkbook, "SortedIndexLatinFR")
    Globals.clearSheet(GrammarWorkbook, "SortedIndexLatinES")
    wsSortedIndexKanji = GrammarWorkbook["SortedIndexKanji"]
    wsSortedIndexRomaji = GrammarWorkbook["SortedIndexRomaji"]
    wsSortedIndexLatinEN = GrammarWorkbook["SortedIndexLatinEN"]
    wsSortedIndexLatinFR = GrammarWorkbook["SortedIndexLatinFR"]
    wsSortedIndexLatinES = GrammarWorkbook["SortedIndexLatinES"]
    # endregion

    # region Settings local parameters & functions
    wsMeaningsPerLanguage = {
        "latinEN": wsMeaningsEN,
        "latinFR": wsMeaningsFR,
        "latinES": wsMeaningsES
    }
    wsSortedIndexPerLanguage = {
        "kanji": wsSortedIndexKanji,
        "romaji": wsSortedIndexRomaji,
        "latinEN": wsSortedIndexLatinEN,
        "latinFR": wsSortedIndexLatinFR,
        "latinES": wsSortedIndexLatinES
    }
    meaningColPerLanguage = {
        "latinEN": Globals.TYPES_COL_MEANINGS_EN,
        "latinFR": Globals.TYPES_COL_MEANINGS_FR,
        "latinES": Globals.TYPES_COL_MEANINGS_ES
    }
    kwColPerLanguage = {
        "latinEN": Globals.TYPES_COL_KW_EN,
        "latinFR": Globals.TYPES_COL_KW_FR,
        "latinES": Globals.TYPES_COL_KW_ES
    }
    SECTIONS_DICT = {
        "kanji": {}, "romaji": {}, "latinEN": {}, "latinFR": {}, "latinES": {}
    }


    def cleanKeywords(keywords, lang):
        if lang == "latinEN":
            keywords = re.sub(r'(i\.e\. |e\.g\. |usu. )', '', keywords)
            keywords = re.sub(r'(\(of |\(in )', Globals.DELIMITER, keywords)
        elif lang == "latinFR":
            keywords = re.sub(r'(c\.-à-d\. |ex\. |par ex\. |en part\. |norm\. |surt\. |surtout |spéc\. )', '', keywords)
        elif lang == "latinES":
            keywords = re.sub(r'(p\.ej\. |p \.ej\. |es decir, |en part\. |norm\. )', '', keywords)

        keywords = re.sub(r'('
                          r'esp. '
                          r'|incl. '
                          r'|, etc.'
                          r'|etc.'
                          r'|/'
                          r'|\.\.\.'
                          r'|…'
                          r'|\('
                          r'|\), '
                          r'|\. '
                          r'|\.'
                          r'|\) '
                          r'|\)'
                          r'\n'
                          r')', '', keywords)

        keywords = keywords.replace('\u200c', '').replace('\u200b', '')  # Removes zero-length chars

        return keywords


    def isANumber(text):
        return (text.isdigit() or text.replace("'", '').isdigit()) and not any(character in "０１２３４５６７８９" for character in text)


    def addRunningIndexToSectionsDict(idx_type, text, idx_local):
        if not text: return
        global SECTIONS_DICT
        if text in SECTIONS_DICT[idx_type].keys():
            indexes_local = SECTIONS_DICT[idx_type][text]
            indexes_local.append(idx_local)
            SECTIONS_DICT[idx_type][text] = indexes_local
        else:
            SECTIONS_DICT[idx_type][text] = [idx_local]


    # endregion

    # region Updating VerbsForGrammar sheet values according to Verbs sheet values
    max_row = 2
    while wsVerbsForGrammar.cell(row=max_row, column=Globals.TYPES_COL_ROMAJI).value is not None: max_row += 1
    wsVerbsForGrammar.delete_rows(2, max_row)
    verbsRow = 2
    verbsForGrammarRow = 2
    while not (wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_FAMILY).value == "-" and wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_SUMMARY_EN).value == "-"):
        if wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KANJI).value:
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_ROMAJI).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_ROMAJI).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_KANJI).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KANJI).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_ALTS).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_ALTS).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_COMMON).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_COMMON).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_KW_JAP).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KW_JAP).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_PREP).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_PREP).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_KANJI_ROOT).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KANJI_ROOT).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_ROMAJI_ROOT).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_ROMAJI_ROOT).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_EXCEPTION_INDEX).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_EXCEPTION_INDEX).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_MEANINGS_EN).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_MEANINGS_EN).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_KW_EN).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KW_EN).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_MEANINGS_FR).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_MEANINGS_FR).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_KW_FR).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KW_FR).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_MEANINGS_ES).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_MEANINGS_ES).value
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_KW_ES).value = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KW_ES).value
            romaji = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_ROMAJI).value
            hiragana = Converter.getOfficialKana(romaji)[0]
            wsVerbsForGrammar.cell(row=verbsForGrammarRow, column=Globals.TYPES_COL_HIRAGANA_FIRST_CHAR).value = str(hiragana[0])

            verbsForGrammarRow += 1
        verbsRow += 1
        if verbsRow % 1000 == 0: print(f'Updated verbsForGrammar for {wsVerbs} - row {verbsRow}')
    # endregion

    # region Setting the running index in the types / grammar / verbs sheets
    runningIndex = 1
    for current_worksheet in [wsTypes, wsGrammar, wsVerbsForGrammar]:
        rowIndex = 2
        modulus = (4000 if current_worksheet != wsGrammar else 200)
        while not Globals.isLastRow(current_worksheet, rowIndex):
            if Globals.isIrrelevantRow(current_worksheet, rowIndex): continue
            if current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_ROMAJI):
                runningIndex += 1
                current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_INDEX).value = runningIndex
            rowIndex += 1
            if rowIndex % modulus == 0: print(f'Updated running index for {current_worksheet} - row {rowIndex}')
    # endregion

    # region Updating the Frequency columns
    for current_worksheet in [wsTypes, wsGrammar, wsVerbsForGrammar]:
        rowIndex = 2
        modulus = (4000 if current_worksheet != wsGrammar else 200)
        while not Globals.isLastRow(current_worksheet, rowIndex):
            if Globals.isIrrelevantRow(current_worksheet, rowIndex): continue
            kanji = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_KANJI).value
            if kanji:
                freq = Converter.get_frequency_from_dict(kanji, freq_dict)
                current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_FREQUENCY).value = freq
            rowIndex += 1
            if rowIndex % modulus == 0: print(f'Updated frequency for {current_worksheet} - row {rowIndex}')
    # endregion

    # region Creating the index dicts and updating the worksheets
    for indexType in SECTIONS_DICT.keys():
        for current_worksheet in [wsTypes, wsGrammar, wsVerbsForGrammar]:

            rowIndex = 2
            modulus = (20000 if current_worksheet != wsGrammar else 200)
            while not Globals.isLastRow(current_worksheet, rowIndex):
                runningIndex = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_INDEX).value
                if rowIndex % modulus == 0: print(f"Creating {indexType} Index for {current_worksheet} - row {rowIndex}")
                if Globals.isIrrelevantRow(current_worksheet, rowIndex): continue
                if current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_ROMAJI).value:
                    if indexType == "kanji":
                        phrasesToIndex = []
                        value = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_KANJI).value
                        if value: phrasesToIndex.append(str(value))
                        value = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_KW_JAP).value
                        if value: phrasesToIndex.append(str(value))
                        value = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_ALTS).value
                        if value: phrasesToIndex.append(str(value))
                    elif indexType == "romaji":
                        phrasesToIndex = []
                        value = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_ROMAJI).value
                        if value: phrasesToIndex.append(str(value))
                        value = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_KW_JAP).value
                        if value: phrasesToIndex.append(str(value))
                        value = current_worksheet.cell(row=rowIndex, column=Globals.TYPES_COL_ALTS).value
                        if value: phrasesToIndex.append(str(value))
                    else:
                        meaningsIndexes = str(current_worksheet.cell(rowIndex, meaningColPerLanguage[indexType]).value)
                        if not meaningsIndexes or meaningsIndexes is None: continue
                        indexes = meaningsIndexes.split(Globals.DELIMITER)
                        phrasesToIndex = [wsMeaningsPerLanguage[indexType].cell(row=int(index), column=Globals.MEANINGS_COL_MEANING).value for index in indexes if index.isdigit()]
                        value = current_worksheet.cell(row=rowIndex, column=kwColPerLanguage[indexType]).value
                        if value: phrasesToIndex.append(str(value))

                    currentKeywords = Globals.DELIMITER.join([str(item) for item in phrasesToIndex]).lower()
                    currentKeywordsClean = cleanKeywords(currentKeywords, indexType)

                    # Separating the keywords into different keyword sections to limit the length of index substrings, thereby limiting the total index size
                    keywordSections = []
                    cumulativeWords = []
                    for phrase in currentKeywordsClean.split(Globals.DELIMITER):
                        words = phrase.strip().split(' ')
                        for i in range(len(words)):
                            if not cumulativeWords:
                                cumulativeWords.append(words[i])
                            else:
                                if len(' '.join(cumulativeWords + [words[i]])) <= Globals.SECTION_LENGTH_THRESHOLD or words[i] == 'suru':
                                    cumulativeWords.append(words[i])
                                else:
                                    keywordSections.append(' '.join(cumulativeWords))
                                    cumulativeWords = []
                        if cumulativeWords:
                            keywordSections.append(' '.join(cumulativeWords))
                            cumulativeWords = []

                    for section in keywordSections:
                        if not section: continue

                        # Determining if the section is a pure number
                        sectionIsANumber = isANumber(section)

                        concatenatedSection = ""
                        if indexType == 'kanji':
                            # Leaving only relevant characters in the section
                            for char in section:
                                if not (char in Globals.LATIN_CHAR_ALPHABET or char in Globals.LATIN_CHAR_ALPHABET_CAP or char in Globals.SYMBOLS_ALPHABET):
                                    concatenatedSection += char

                            for i in range(len(concatenatedSection)):
                                section_substring = concatenatedSection[i:]
                                addRunningIndexToSectionsDict(indexType, section_substring, str(runningIndex))
                        else:
                            # Leaving only relevant characters in the section
                            for char in section:
                                if (sectionIsANumber and char in Globals.NUMBER_ALPHABET) \
                                        or (not sectionIsANumber and (char in Globals.LATIN_CHAR_ALPHABET or char in Globals.LATIN_CHAR_ALPHABET_CAP)
                                ):
                                    concatenatedSection += char
                            if not sectionIsANumber: concatenatedSection = re.sub(Globals.NUMBER_ALPHABET, '', concatenatedSection)  # Removing digits from non-number sections

                            if len(concatenatedSection) <= 2 or sectionIsANumber:
                                # Don't write entries that will result in a substring of only one or two characters or consisting only of digits, unless that was the original word
                                if re.search(r'^0+$', concatenatedSection):
                                    section_substring = "@@@@@@@@" + concatenatedSection
                                elif sectionIsANumber:
                                    section_substring = "@@@@@@@@" + re.sub(r'^0+', '', concatenatedSection)
                                else:
                                    section_substring = concatenatedSection
                                addRunningIndexToSectionsDict(indexType, section_substring, str(runningIndex))
                            else:
                                # Creating the word fragments in the index
                                for i in range(len(concatenatedSection) - 2):
                                    section_substring = concatenatedSection[i:]
                                    if re.search(r'^0+$', concatenatedSection):
                                        section_substring = "@@@@@@@@" + concatenatedSection
                                    elif isANumber(section_substring):
                                        section_substring = "@@@@@@@@" + re.sub(r'^0+', '', section_substring)
                                    addRunningIndexToSectionsDict(indexType, section_substring, str(runningIndex))
                rowIndex += 1

        sortedSectionSubstrings = sorted(SECTIONS_DICT[indexType].keys(), key=lambda x: Globals.convert_to_utf8(x))
        print(str(sortedSectionSubstrings[:10]))
        print(f'Number of sortedSectionSubstrings: {len(sortedSectionSubstrings)}')
        if indexType == 'kanji':
            for i in range(len(sortedSectionSubstrings)):
                if i % 20000 == 0: print(f'Updated {indexType} sorted index - item {i}/{len(sortedSectionSubstrings)}')
                wsSortedIndexPerLanguage[indexType].cell(row=i + 1, column=1).value = sortedSectionSubstrings[i]
                wsSortedIndexPerLanguage[indexType].cell(row=i + 1, column=2).value = Globals.DELIMITER.join(sorted(list(set(SECTIONS_DICT[indexType][sortedSectionSubstrings[i]]))))
                wsSortedIndexPerLanguage[indexType].cell(row=i + 1, column=3).value = Globals.convert_to_utf8(sortedSectionSubstrings[i]).upper()
        else:
            for i in range(len(sortedSectionSubstrings)):
                if i % 10000 == 0: print(f'Updated {indexType} sorted index - item {i}/{len(sortedSectionSubstrings)}')
                wsSortedIndexPerLanguage[indexType].cell(row=i + 1, column=1).value = sortedSectionSubstrings[i].replace("@@@@@@@@", '')
                wsSortedIndexPerLanguage[indexType].cell(row=i + 1, column=2).value = Globals.DELIMITER.join(sorted(list(set(SECTIONS_DICT[indexType][sortedSectionSubstrings[i]]))))
    # endregion

    # region Saving the results to xlsx & csv
    if update_workbooks:
        GrammarWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Grammar - 3000 kanji - ready for Japagram.xlsx')
        VerbsWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - ready for Japagram.xlsx')

    Globals.create_csv_from_worksheet(wsSortedIndexKanji, name("GrammarSortedIndexKanji"), idx("A"), idx("C"))
    Globals.create_csv_from_worksheet(wsSortedIndexRomaji, name("GrammarSortedIndexRomaji"), idx("A"), idx("B"))
    Globals.create_csv_from_worksheet(wsSortedIndexLatinEN, name("GrammarSortedIndexLatinEN"), idx("A"), idx("B"))
    Globals.create_csv_from_worksheet(wsSortedIndexLatinFR, name("GrammarSortedIndexLatinFR"), idx("A"), idx("B"))
    Globals.create_csv_from_worksheet(wsSortedIndexLatinES, name("GrammarSortedIndexLatinES"), idx("A"), idx("B"))
    Globals.create_csv_from_worksheet(wsMeaningsEN, name("Meanings"), idx("A"), idx("H"))
    Globals.create_csv_from_worksheet(wsMeaningsFR, name("MeaningsFR"), idx("A"), idx("H"))
    Globals.create_csv_from_worksheet(wsMeaningsES, name("MeaningsES"), idx("A"), idx("H"))
    Globals.create_csv_from_worksheet(wsTypes, name("Types"), idx("A"), idx("R"))
    Globals.create_csv_from_worksheet(wsGrammar, name("Grammar"), idx("A"), idx("R"))
    Globals.create_csv_from_worksheet(wsMultExplEN, name("MultExplEN"), idx("A"), idx("D"))
    Globals.create_csv_from_worksheet(wsMultExplFR, name("MultExplFR"), idx("A"), idx("D"))
    Globals.create_csv_from_worksheet(wsMultExplES, name("MultExplES"), idx("A"), idx("D"))
    Globals.create_csv_from_worksheet(wsExamples, name("Examples"), idx("A"), idx("F"))
    Globals.create_csv_from_worksheet(wsVerbsForGrammar, name("VerbsForGrammar"), idx("A"), idx("R"))
    # endregion

if prepare_conj_lengths:
    # region Reading worksheets
    VerbsWorkbook = openpyxl.load_workbook(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - ready for Japagram.xlsx', data_only=True)
    print("Finished loading Verbs - 3000 kanji - ready for Japagram.xlsx")
    wsVerbs = VerbsWorkbook["Verbs"]
    wsLatinConj = VerbsWorkbook["LatinConj"]
    wsKanjiConj = VerbsWorkbook["KanjiConj"]

    Globals.clearSheet(VerbsWorkbook, "VerbsLengths")
    Globals.clearSheet(VerbsWorkbook, "VerbsKanjiLengths")
    wsVerbsLengths = VerbsWorkbook["VerbsLengths"]
    wsVerbsKanjiLengths = VerbsWorkbook["VerbsKanjiLengths"]
    # endregion

    # region Updating the max verb conjugation lengths
    verbsRow = 2
    verbsForGrammarRow = 2
    latin_lengths = {}
    kanji_lengths = {}
    family_row_index = 0
    family = ''
    while not (wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_FAMILY).value == "-" and wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_SUMMARY_EN).value == "-"):

        is_family_row = False
        if wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_FAMILY).value and not wsVerbs.cell(row=verbsRow - 1, column=Globals.VERBS_COL_FAMILY).value:
            family_row_index = int(wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_EXCEPTION_INDEX).value) + 1
            family = wsLatinConj.cell(row=family_row_index, column=1).value
            wsVerbsLengths.cell(row=family_row_index, column=1).value = family
            wsVerbsKanjiLengths.cell(row=family_row_index, column=1).value = family
            is_family_row = True

        if wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KANJI).value or is_family_row:
            exception_row_index = int(wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_EXCEPTION_INDEX).value) + 1
            if exception_row_index not in latin_lengths.keys(): latin_lengths[exception_row_index] = {}
            if exception_row_index not in kanji_lengths.keys(): kanji_lengths[exception_row_index] = {}

            current_col = VERBS_COL_CONJ_FIRST_CONJ
            while wsLatinConj.cell(row=family_row_index, column=current_col).value \
                    or wsLatinConj.cell(row=family_row_index, column=current_col + 1).value \
                    or wsLatinConj.cell(row=family_row_index, column=current_col + 2).value \
                    or wsLatinConj.cell(row=family_row_index, column=current_col + 3).value \
                    or wsLatinConj.cell(row=family_row_index, column=current_col + 4).value \
                    or wsLatinConj.cell(row=family_row_index, column=current_col + 5).value:
                if is_family_row or exception_row_index != family_row_index and wsLatinConj.cell(row=exception_row_index, column=current_col).value:
                    conjugation = wsLatinConj.cell(row=exception_row_index, column=current_col).value
                    romaji_conjugation = conjugation if conjugation else ""
                    conjugation = wsKanjiConj.cell(row=exception_row_index, column=current_col).value
                    kanji_conjugation = conjugation if conjugation else ""
                else:
                    root = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_ROMAJI_ROOT).value
                    conjugation = wsLatinConj.cell(row=family_row_index, column=current_col).value
                    romaji_conjugation = (root if root else "") + (conjugation if conjugation else "")
                    root = wsVerbs.cell(row=verbsRow, column=Globals.VERBS_COL_KANJI_ROOT).value
                    conjugation = wsKanjiConj.cell(row=family_row_index, column=current_col).value
                    kanji_conjugation = (root if root else "") + (conjugation if conjugation else "")

                romaji_length = len(romaji_conjugation)
                kanji_length = len(kanji_conjugation)
                latin_lengths[exception_row_index][current_col] = max(romaji_length, latin_lengths[exception_row_index][current_col]) \
                    if current_col in latin_lengths[exception_row_index].keys() else romaji_length
                kanji_lengths[exception_row_index][current_col] = max(kanji_length, kanji_lengths[exception_row_index][current_col]) \
                    if current_col in kanji_lengths[exception_row_index].keys() else kanji_length
                current_col += 1
        verbsRow += 1
        if verbsRow % 1000 == 0: print(f'Processed lengths for {wsVerbs} - row {verbsRow}')

    exception_indexes = sorted(list(latin_lengths.keys()))
    family_exception_index = 0
    family_exception_indexes = []
    last_index = 0
    for index in exception_indexes:
        if index == last_index + 1:
            for length in latin_lengths[index].keys():
                latin_lengths[family_exception_index][length] = max(latin_lengths[family_exception_index][length], latin_lengths[index][length])
                kanji_lengths[family_exception_index][length] = max(kanji_lengths[family_exception_index][length], kanji_lengths[index][length])
        else:
            family_exception_index = index
            family_exception_indexes.append(index)
        last_index = index

    for index in family_exception_indexes:
        for column_index in latin_lengths[index].keys():
            wsVerbsLengths.cell(row=index, column=column_index).value = latin_lengths[index][column_index]
            wsVerbsKanjiLengths.cell(row=index, column=column_index).value = kanji_lengths[index][column_index]

            max_latin_value = wsVerbsLengths.cell(row=1, column=column_index).value if wsVerbsLengths.cell(row=1, column=column_index).value else 0
            max_kanji_value = wsVerbsKanjiLengths.cell(row=1, column=column_index).value if wsVerbsKanjiLengths.cell(row=1, column=column_index).value else 0
            wsVerbsLengths.cell(row=1, column=column_index).value = max(latin_lengths[index][column_index], max_latin_value)
            wsVerbsKanjiLengths.cell(row=1, column=column_index).value = max(kanji_lengths[index][column_index], max_kanji_value)

    # endregion

    # region Saving the results to xlsx & csv
    if update_workbooks:
        VerbsWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - ready for Japagram.xlsx')

    Globals.create_csv_from_worksheet(wsLatinConj, name("LatinConj"), idx("A"), idx(Globals.MAX_VERB_CONG_COL))
    Globals.create_csv_from_worksheet(wsKanjiConj, name("KanjiConj"), idx("A"), idx(Globals.MAX_VERB_CONG_COL))
    Globals.create_csv_from_worksheet(wsVerbsLengths, name("VerbsLengths"), idx("A"), idx(Globals.MAX_VERB_CONG_COL), False)
    Globals.create_csv_from_worksheet(wsVerbsKanjiLengths, name("VerbsKanjiLengths"), idx("A"), idx(Globals.MAX_VERB_CONG_COL), False)
    # endregion

if prepare_extended_db: ExtendedDbCreator.main()

if prepare_names_db: ExtendedDbCreator_Names.main()

if prepare_kanji_db:
    # region Reading worksheets
    RootsWorkbook = openpyxl.load_workbook(filename=f'{Globals.MASTER_DIR}/Roots - 3000 kanji - MASTER.xlsx', data_only=True)
    print("Finished loading Roots - 3000 kanji.xlsx")
    Globals.clearSheet(RootsWorkbook, "Components")
    wsRadicals = RootsWorkbook["Radicals"]
    wsRadicalsOnly = RootsWorkbook["RadicalsOnly"]
    wsKanjiDict = RootsWorkbook["Kanji_Dict"]
    wsCJK_Decomposition = RootsWorkbook["CJK_Decomposition"]
    wsSimilars = RootsWorkbook["Similars"]
    wsComponents = RootsWorkbook["Components"]
    # endregion

    # region Creating the sorted decompositions dictionary
    decompositions = {}
    radicals = []
    radicals_not_to_decompose = []
    for radical_index in range(1, 361):  # 360 = First index of katakana in RadicalsOnly sheet
        radicals.append(str(wsRadicalsOnly.cell(row=radical_index, column=1).value))
    for radical_index in range(1, 552):  # 551 = Last index of RadicalsOnly sheet
        if not (radical_index == 96 or radical_index == 124 or radical_index == 152 or radical_index == 154 or radical_index == 172 or radical_index == 177 or radical_index == 190
                or radical_index == 191 or radical_index == 194 or radical_index == 195 or radical_index == 203 or radical_index == 204 or radical_index == 205 or radical_index == 207
                or radical_index == 208 or radical_index == 214 or radical_index == 218 or radical_index == 222 or radical_index == 223 or radical_index == 226 or radical_index == 227
                or radical_index == 229 or radical_index == 233 or radical_index == 236 or radical_index == 237 or radical_index == 240 or radical_index == 244 or radical_index == 250
                or radical_index == 251 or radical_index == 256 or radical_index == 257 or radical_index == 258 or radical_index == 259 or radical_index == 264 or radical_index == 265
                or radical_index == 266 or radical_index == 270 or radical_index == 286 or radical_index == 298 or radical_index == 301 or radical_index == 303 or radical_index == 310
                or radical_index == 311 or radical_index == 312 or radical_index == 315 or radical_index == 316 or radical_index == 317 or radical_index == 320 or radical_index == 321
                or radical_index == 323 or radical_index == 324
                or (329 <= radical_index <= 338)
                or radical_index == 342 or radical_index == 343
                or (346 <= radical_index <= 358 and radical_index != 355)):
            radicals_not_to_decompose.append(str(wsRadicalsOnly.cell(row=radical_index, column=1).value))
    VALUE = 0
    UTF8_VALUE = 1
    STRUCTURE = 2
    DECOMPOSITION = 3
    FULL_DECOMPOSITION = 4
    row = 3
    while not Globals.isLastRow(wsCJK_Decomposition, row):
        value = str(wsCJK_Decomposition.cell(row=row, column=1).value)
        key = [
            value,
            str(Globals.convert_to_utf8(str(wsCJK_Decomposition.cell(row=row, column=1).value)).upper()),
            str(wsCJK_Decomposition.cell(row=row, column=3).value) if wsCJK_Decomposition.cell(row=row, column=4).value else "",
            str(wsCJK_Decomposition.cell(row=row, column=4).value) if wsCJK_Decomposition.cell(row=row, column=4).value else "",
            ""
        ]
        decompositions[value] = key
        row += 1


    def Recursive_Decomposition(line_key):
        if line_key not in decompositions or not decompositions[line_key][DECOMPOSITION]: return []
        line_decompositions = decompositions[line_key][DECOMPOSITION].split(Globals.DELIMITER)
        current_full_decomposition = []
        for item in line_decompositions:
            if not item: continue
            if item.isdigit():
                current_full_decomposition += Recursive_Decomposition(item)
            else:
                current_full_decomposition.append(item)
                if item not in radicals_not_to_decompose:
                    current_full_decomposition += Recursive_Decomposition(item)
        return current_full_decomposition


    for key in decompositions.keys():
        full_decomposition = Recursive_Decomposition(key)
        final_current_value_list = remove_duplicates_keep_order(full_decomposition)
        decompositions[key][FULL_DECOMPOSITION] = "".join(final_current_value_list)

    print('\n'.join([str(decompositions[item]) for item in list(decompositions.keys())[:20]]))

    # endregion

    # region Updating the CJK_Decompositons, KanjiDict and Radicals sheets
    row = 3
    sorted_keys = sorted(decompositions.keys(), key=lambda x: decompositions[x][UTF8_VALUE])
    for i in range(len(sorted_keys)):
        wsCJK_Decomposition.cell(row=i + 3, column=1).value = decompositions[sorted_keys[i]][VALUE]
        wsCJK_Decomposition.cell(row=i + 3, column=2).value = decompositions[sorted_keys[i]][UTF8_VALUE]
        wsCJK_Decomposition.cell(row=i + 3, column=3).value = decompositions[sorted_keys[i]][STRUCTURE]
        wsCJK_Decomposition.cell(row=i + 3, column=4).value = decompositions[sorted_keys[i]][DECOMPOSITION]
        wsCJK_Decomposition.cell(row=i + 3, column=5).value = decompositions[sorted_keys[i]][FULL_DECOMPOSITION]

    row = 2
    radicals = []
    while not Globals.isLastRow(wsRadicals, row):
        item = [
            str(wsRadicals.cell(row=row, column=1).value) if wsRadicals.cell(row=row, column=1).value is not None else "",
            str(Globals.convert_to_utf8(str(wsRadicals.cell(row=row, column=1).value))).upper(),
            str(wsRadicals.cell(row=row, column=3).value) if wsRadicals.cell(row=row, column=3).value is not None else ""
        ]
        radicals.append(item)
        row += 1
    sorted_radicals = sorted(radicals, key=lambda x: x[1])
    for i in range(len(sorted_radicals)):
        wsRadicals.cell(row=i + 2, column=1).value = sorted_radicals[i][0]
        wsRadicals.cell(row=i + 2, column=2).value = sorted_radicals[i][1]
        wsRadicals.cell(row=i + 2, column=3).value = sorted_radicals[i][2]

    row = 2
    kanji_dict_items = []
    while not Globals.isLastRow(wsKanjiDict, row):
        item = [
            str(wsKanjiDict.cell(row=row, column=1).value),
            str(Globals.convert_to_utf8(str(wsKanjiDict.cell(row=row, column=1).value))).upper(),
            str(wsKanjiDict.cell(row=row, column=3).value) if wsKanjiDict.cell(row=row, column=3).value is not None else "",
            str(wsKanjiDict.cell(row=row, column=4).value) if wsKanjiDict.cell(row=row, column=4).value is not None else "",
            str(wsKanjiDict.cell(row=row, column=5).value) if wsKanjiDict.cell(row=row, column=5).value is not None else "",
            str(wsKanjiDict.cell(row=row, column=6).value) if wsKanjiDict.cell(row=row, column=6).value is not None else ""
        ]
        kanji_dict_items.append(item)
        row += 1
    sorted_kanji_dict_items = sorted(kanji_dict_items, key=lambda x: x[1])
    for i in range(len(sorted_kanji_dict_items)):
        wsKanjiDict.cell(row=i + 2, column=1).value = sorted_kanji_dict_items[i][0]
        wsKanjiDict.cell(row=i + 2, column=2).value = sorted_kanji_dict_items[i][1]
        wsKanjiDict.cell(row=i + 2, column=3).value = sorted_kanji_dict_items[i][2]
        wsKanjiDict.cell(row=i + 2, column=4).value = sorted_kanji_dict_items[i][3]
        wsKanjiDict.cell(row=i + 2, column=5).value = sorted_kanji_dict_items[i][4]
        wsKanjiDict.cell(row=i + 2, column=6).value = sorted_kanji_dict_items[i][5]
        row += 1
    # endregion

    # region Classify the components
    struct_map = {
        "a2": "across2",
        "a2m": "across2",
        "a2t": "across2",
        "a3": "across3",
        "a4": "across4",
        "d2": "down2",
        "d2m": "down2",
        "d2t": "down2",
        "d3": "down3",
        "d4": "down4",
        "r3tr": "repeat3special",
        "r3gw": "repeat3special",
        "r3st": "repeat3special",
        "r3stl": "repeat3special",
        "r3str": "repeat3special",
        "r4sq": "repeat4special",
        "4sq": "foursquare",
        "r5": "repeat5special",
        "stl": "topleftout",
        "st": "topout",
        "str": "toprightout",
        "sl": "leftout",
        "s": "fullout",
        "sbl": "bottomleftout",
        "sb": "bottomout",
        "sbr": "bottomrightout",
    }
    compts_by_struct = {
        "full": {},
        "across2": {},
        "across3": {},
        "across4": {},
        "down2": {},
        "down3": {},
        "down4": {},
        "repeat3special": {},
        "repeat4special": {},
        "foursquare": {},
        "repeat5special": {},
        "topleftout": {},
        "topout": {},
        "toprightout": {},
        "leftout": {},
        "fullout": {},
        "bottomleftout": {},
        "bottomout": {},
        "bottomrightout": {},
    }
    for key in decompositions.keys():
        if key.isdigit(): continue
        for element in list(decompositions[key][FULL_DECOMPOSITION]):
            compts_by_struct["full"][element] = (compts_by_struct["full"][element] + key) if element in compts_by_struct["full"].keys() else key

        if decompositions[key][STRUCTURE] not in struct_map: continue
        struct = struct_map[decompositions[key][STRUCTURE]]
        elements = decompositions[key][DECOMPOSITION].split(Globals.DELIMITER)
        if struct[-3:] == 'out':
            element = elements[0]
            compts_by_struct[struct][element] = (compts_by_struct[struct][element] + key) if element in compts_by_struct[struct].keys() else key
        else:
            for element in elements:
                if element.isdigit(): continue
                compts_by_struct[struct][element] = (compts_by_struct[struct][element] + key) if element in compts_by_struct[struct].keys() else key
    # endregion

    # region Update the Components sheet
    row = 1
    for structure in compts_by_struct.keys():
        if structure != "full":
            wsComponents.cell(row=row, column=1).value = structure
            row += 1
        ordered_elements = sorted(compts_by_struct[structure].keys(), key=lambda x: convert_to_utf8(x))
        for element in ordered_elements:
            wsComponents.cell(row=row, column=1).value = element
            wsComponents.cell(row=row, column=2).value = compts_by_struct[structure][element]
            row += 1
    # print('\n'.join([key + " : " + str(compts_by_struct["full"][key]) for key in list(compts_by_struct["full"].keys())[:20]]))
    # endregion

    # region Saving the results to xlsx & csv
    if update_workbooks:
        RootsWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Roots - 3000 kanji - ready for Japagram.xlsx')

    Globals.create_csv_from_worksheet(wsRadicals, name("Radicals"), idx("B"), idx("C"), False, 2)
    Globals.create_csv_from_worksheet(wsKanjiDict, name("KanjiDictionary"), idx("B"), idx("F"), False, 2)
    Globals.create_csv_from_worksheet(wsCJK_Decomposition, name("CJK_Decomposition"), idx("B"), idx("D"), False, 2)
    Globals.create_csv_from_worksheet(wsRadicalsOnly, name("RadicalsOnly"), idx("A"), idx("G"), False, 2)
    Globals.create_csv_from_worksheet(wsSimilars, name("Similars"), idx("A"), idx("E"), False, 2)
    Globals.create_csv_from_worksheet(wsComponents, name("Components"), idx("A"), idx("B"), False, 1)
    # endregion

if prepare_conj_db:
    # region Reading worksheets
    VerbsWorkbook = openpyxl.load_workbook(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - with foreign.xlsx', data_only=True)
    print("Finished loading Verbs - 3000 kanji.xlsx")
    wsVerbsForGrammar = VerbsWorkbook["VerbsForGrammar"]
    wsVerbs = VerbsWorkbook["Verbs"]
    wsLatinConj = VerbsWorkbook["LatinConj"]
    wsKanjiConj = VerbsWorkbook["KanjiConj"]
    wsLatinConjIndex = VerbsWorkbook["LatinConjIndex"]
    wsKanjiConjIndex = VerbsWorkbook["KanjiConjIndex"]
    # endregion

    # region Getting the family conjugations
    conjugations_latin_per_family = {}
    conjugations_kanji_per_family = {}
    last_col = 1
    while wsLatinConj.cell(row=3, column=last_col).value:
        last_col += 1

    row = 2
    while not Globals.isLastRow(wsLatinConj, row):
        family = wsLatinConj.cell(row=row, column=idx("A")).value
        if not family:
            row += 1
            continue
        match = re.search(r'(.*(godan|special class|ichidan|suru|kuru|desu).*)', wsLatinConj.cell(row=row, column=idx("A")).value)
        if match and not wsLatinConj.cell(row=row, column=idx("H")).value:
            conjugations_latin_per_family[match.group(0)] = [wsLatinConj.cell(row=row, column=i).value if wsLatinConj.cell(row=row, column=i).value else "" for i in range(1, last_col + 1)]
            conjugations_kanji_per_family[match.group(0)] = [wsKanjiConj.cell(row=row, column=i).value if wsKanjiConj.cell(row=row, column=i).value else "" for i in range(1, last_col + 1)]
        row += 1
    # endregion

    # region Creating the conjugation dicts and index
    latinConjIndex = {}
    kanjiConjIndex = {}
    row = 2
    while not Globals.isLastRow(wsVerbsForGrammar, row):
        conjIndex = int(wsVerbsForGrammar.cell(row=row, column=Globals.TYPES_COL_EXCEPTION_INDEX).value) + 1
        verbIndex = str(wsVerbsForGrammar.cell(row=row, column=Globals.TYPES_COL_INDEX).value)
        conj_family = wsLatinConj.cell(row=conjIndex, column=idx("A")).value
        is_family_conj = not wsLatinConj.cell(row=conjIndex, column=idx("B")).value
        alt_spellings = wsVerbsForGrammar.cell(row=row, column=Globals.TYPES_COL_ALTS).value
        alt_spellings = alt_spellings.replace(' ', '') if alt_spellings else ""
        root = wsVerbsForGrammar.cell(row=row, column=Globals.TYPES_COL_ROMAJI_ROOT).value
        latinRoots = [root if root else ""]
        root = wsVerbsForGrammar.cell(row=row, column=Globals.TYPES_COL_KANJI_ROOT).value
        kanjiRoots = [root if root else ""]
        for item in alt_spellings.split(Globals.DELIMITER):
            if not item: continue
            if Globals.is_latin(item):
                latinRoots.append(Globals.get_root_from_masu_stem_latin(item, conj_family))
            else:
                kanjiRoots.append(Globals.get_root_from_masu_stem_kanji(item, conj_family))

        for latinRoot in latinRoots:
            current_verb_conjugations_latin = [latinRoot + termination for termination in conjugations_latin_per_family[conj_family][10:]]
            if not is_family_conj:
                for col in range(0, last_col - 10):
                    current_verb_conjugations_latin[col] = wsLatinConj.cell(row=conjIndex, column=col + 10).value
            for item in current_verb_conjugations_latin:
                if item:
                    latinConjIndex[item] = [verbIndex] if item not in latinConjIndex.keys() else latinConjIndex[item] + [verbIndex]
                    latinConjIndex[item] = list(set(latinConjIndex[item]))

        for kanjiRoot in kanjiRoots:
            current_verb_conjugations_kanji = [kanjiRoot + termination for termination in conjugations_kanji_per_family[conj_family][10:]]
            if not is_family_conj:
                for col in range(0, last_col - 10):
                    current_verb_conjugations_kanji[col] = wsKanjiConj.cell(row=conjIndex, column=col + 10).value
            for item in current_verb_conjugations_kanji:
                if item and item != '*':
                    kanjiConjIndex[item] = [verbIndex] if item not in kanjiConjIndex.keys() else kanjiConjIndex[item] + [verbIndex]
                    kanjiConjIndex[item] = list(set(kanjiConjIndex[item]))

        if row % 100 == 0: print(f"VerbsForGrammar - finished row {row}")
        row += 1
    # endregion

    # region Writing to the worksheets
    sorted_keys = sorted(latinConjIndex.keys())
    for i in range(0, len(sorted_keys)):
        wsLatinConjIndex.cell(row=i + 1, column=1).value = sorted_keys[i]
        wsLatinConjIndex.cell(row=i + 1, column=2).value = Globals.DELIMITER.join(latinConjIndex[sorted_keys[i]])
    sorted_keys = sorted(kanjiConjIndex.keys())
    for i in range(0, len(sorted_keys)):
        wsKanjiConjIndex.cell(row=i + 1, column=1).value = sorted_keys[i]
        wsKanjiConjIndex.cell(row=i + 1, column=2).value = Globals.DELIMITER.join(kanjiConjIndex[sorted_keys[i]])
    # endregion

    # region Saving the results to xlsx & csv
    if update_workbooks:
        VerbsWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - ready for Japagram.xlsx')

    Globals.create_csv_from_worksheet(wsLatinConjIndex, name("VerbConjLatinSortedIndex"), idx("A"), idx("B"))
    Globals.create_csv_from_worksheet(wsKanjiConjIndex, name("VerbConjKanjiSortedIndex"), idx("A"), idx("B"))
    # endregion
