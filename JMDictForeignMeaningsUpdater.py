#!/usr/bin/python -tt

import re
import openpyxl

import Globals
import Converter

def main():
    current_entry = ''

    # region Preparations
    localWordsWorkbook = openpyxl.load_workbook(filename=f'{Globals.MASTER_DIR}/Grammar - 3000 kanji - MASTER.xlsx', data_only=True)
    localVerbsWorkbook = openpyxl.load_workbook(filename=f'{Globals.MASTER_DIR}/Verbs - 3000 kanji - MASTER.xlsx', data_only=True)
    wsLocalMeanings = localWordsWorkbook["Meanings"]
    wsLocalMeaningsFR = localWordsWorkbook["MeaningsFR"]
    wsLocalMeaningsES = localWordsWorkbook["MeaningsES"]
    wsLocalTypes = localWordsWorkbook["Types"]
    wsLocalGrammar = localWordsWorkbook["Grammar"]
    wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
    wsLocalVerbs = localVerbsWorkbook["Verbs"]


    class Word:
        def __init__(self, kanji, hiragana, romaji, french_meanings, spanish_meanings):
            self.kanji = kanji
            self.hiragana = hiragana
            self.romaji = romaji
            self.uniqueID = romaji + "zzz" + kanji
            self.french_meanings = french_meanings
            self.spanish_meanings = spanish_meanings

    # endregion

    # region Filling the foreign words list
    words = []
    reached_first_entry = False
    with open("JMDict.xml", encoding='utf-8') as infile:
        for line in infile:

            if "<entry>" in line:
                reached_first_entry = True
            if not reached_first_entry:
                continue

            current_entry += line

            if "</entry>" in line:
                match = re.search("<keb>(\w+)</keb>", current_entry)
                kanji = ''
                if match:
                    kanji = match.group(1)

                match = re.search("<reb>(\w+)</reb>", current_entry)
                hiragana = ''
                if match:
                    hiragana = match.group(1)
                    if kanji == '': kanji = hiragana

                senses = re.findall(r"<sense>(.*?)</sense>", current_entry, re.MULTILINE | re.DOTALL)

                french_meanings = []
                spanish_meanings = []
                for sense in senses:
                    french_meanings_in_sense = re.findall(r"<gloss xml:lang=\"fre\">(.+?)</gloss>", sense, re.MULTILINE | re.DOTALL)
                    if len(french_meanings_in_sense) > 0: french_meanings.append(", ".join(french_meanings_in_sense))
                    spanish_meanings_in_sense = re.findall(r"<gloss xml:lang=\"spa\">(.+?)</gloss>", sense, re.MULTILINE | re.DOTALL)
                    if len(spanish_meanings_in_sense) > 0: spanish_meanings.append(", ".join(spanish_meanings_in_sense))

                if not kanji == '' and not hiragana == '' \
                        and not (len(french_meanings) == 0 and len(spanish_meanings) == 0):
                    word = Word(kanji,
                                hiragana,
                                Converter.getOfficialWaapuroOnly(hiragana),
                                french_meanings,
                                spanish_meanings)
                    words.append(word)

                current_entry = ''


    # endregion

    # region Updating the Types and Meaning sheets with the values
    def binary_search(words_list, text, lo=0, hi=None):
        # https://stackoverflow.com/questions/4161629/search-a-list-of-objects-in-python
        if hi is None:
            hi = len(words_list)
        while lo < hi:
            mid = (lo + hi) // 2

            if text > words_list[mid].uniqueID:
                lo = mid + 1
            elif text < words_list[mid].uniqueID:
                hi = mid
            elif text == words_list[mid].uniqueID:
                return mid
            else:
                return -1

        return -1


    localWordsList = []
    jmDictWordsList = list(words)
    jmDictWordsList.sort(key=lambda x: x.uniqueID)
    meaningFRindex = 2
    meaningESindex = 2
    sheetNum_types = 0
    sheetNum_grammar = 1
    sheetNum_verbs = 2

    for sheetNum in range(0, 3):

        if sheetNum == sheetNum_types:
            typesIndex = 2
            workingWorksheet = wsLocalTypes
        elif sheetNum == sheetNum_grammar:
            typesIndex = 2
            workingWorksheet = wsLocalGrammar
        elif sheetNum == sheetNum_verbs:
            typesIndex = 4
            workingWorksheet = wsLocalVerbs
        else:
            typesIndex = 2
            workingWorksheet = wsLocalTypes

        # region Row skip conditions
        while True:

            romaji = ""
            kanji = ""
            if sheetNum == sheetNum_types or sheetNum == sheetNum_grammar:
                romaji = workingWorksheet.cell(row=typesIndex, column=Globals.TYPES_COL_ROMAJI).value
                kanji = workingWorksheet.cell(row=typesIndex, column=Globals.TYPES_COL_KANJI).value

                if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
                    break
            elif sheetNum == sheetNum_verbs:
                romaji = workingWorksheet.cell(row=typesIndex, column=Globals.VERBS_COL_ROMAJI).value
                kanji = workingWorksheet.cell(row=typesIndex, column=Globals.VERBS_COL_KANJI).value

                if (workingWorksheet.cell(row=typesIndex, column=1).value == '-'
                        and workingWorksheet.cell(row=typesIndex, column=2).value == '-'
                        and workingWorksheet.cell(row=typesIndex, column=3).value == '-'):
                    break

                if (romaji == "" or not romaji or romaji is None) or (kanji == "" or not kanji or kanji is None):
                    typesIndex += 1
                    continue
            # endregion

            # region Retrieving the type from the english dB
            meaningsEN = workingWorksheet.cell(row=typesIndex,
                                               column=Globals.TYPES_COL_MEANINGS_EN if sheetNum != sheetNum_verbs else Globals.VERBS_COL_MEANINGS_EN
                                               ).value
            meaning_types = [wsLocalMeanings.cell(row=int(meaning_index), column=Globals.MEANINGS_COL_TYPE).value for meaning_index in str(meaningsEN).split(";")]

            first_meaning_type = meaning_types[0]
            use_english_type = False
            if sheetNum == sheetNum_verbs or len(set(meaning_types)) == 1:  # ie. all meanings are the same
                use_english_type = True
            # endregion

            romaji = romaji.replace(' ', '')

            localWordsList.append(Word(kanji, "", romaji, "", ""))

            uniqueID = romaji + "zzz" + kanji
            # print(uniqueID + "\n")

            index_of_first_hit = binary_search(jmDictWordsList, uniqueID)
            if index_of_first_hit != -1:

                word = jmDictWordsList[index_of_first_hit]

                # print(word.romaji + "-" + word.kanji + "\n")

                currentFRmeanings = workingWorksheet.cell(row=typesIndex,
                                                          column=Globals.TYPES_COL_MEANINGS_FR if sheetNum != sheetNum_verbs else Globals.VERBS_COL_MEANINGS_FR
                                                          ).value
                if not currentFRmeanings or currentFRmeanings == "":
                    meaningIndexesFR = []
                    while wsLocalMeaningsFR.cell(row=meaningFRindex, column=Globals.MEANINGS_COL_SOURCE).value == 'LOC':
                        meaningFRindex += 1
                    for j in range(len(word.french_meanings)):
                        fixed_meanings = word.french_meanings[j].replace('œ', 'oe')
                        fixed_meanings = re.sub(r'(\d),(\d)', r'\g<1>.\g<2>', fixed_meanings)

                        wsLocalMeaningsFR.cell(row=meaningFRindex, column=Globals.MEANINGS_COL_INDEX).value = meaningFRindex
                        wsLocalMeaningsFR.cell(row=meaningFRindex, column=Globals.MEANINGS_COL_MEANING).value = fixed_meanings
                        wsLocalMeaningsFR.cell(row=meaningFRindex, column=Globals.MEANINGS_COL_TYPE).value = first_meaning_type if use_english_type else ""
                        wsLocalMeaningsFR.cell(row=meaningFRindex, column=Globals.MEANINGS_COL_SOURCE).value = "JM"
                        meaningIndexesFR.append(str(meaningFRindex))
                        meaningFRindex += 1
                    workingWorksheet.cell(row=typesIndex,
                                          column=Globals.TYPES_COL_MEANINGS_FR if sheetNum != sheetNum_verbs else Globals.VERBS_COL_MEANINGS_FR
                                          ).value = ";".join(meaningIndexesFR)

                currentESmeanings = workingWorksheet.cell(row=typesIndex,
                                                          column=Globals.TYPES_COL_MEANINGS_ES if sheetNum < 2 else Globals.VERBS_COL_MEANINGS_ES
                                                          ).value
                if not currentESmeanings or currentESmeanings == "":
                    meaningIndexesES = []
                    while wsLocalMeaningsES.cell(row=meaningESindex, column=Globals.MEANINGS_COL_SOURCE).value == 'LOC':
                        meaningESindex += 1
                    for j in range(len(word.spanish_meanings)):
                        fixed_meanings = re.sub(r'(\d),(\d)', r'\g<1>.\g<2>', word.spanish_meanings[j])
                        wsLocalMeaningsES.cell(row=meaningESindex, column=Globals.MEANINGS_COL_INDEX).value = meaningESindex
                        wsLocalMeaningsES.cell(row=meaningESindex, column=Globals.MEANINGS_COL_MEANING).value = fixed_meanings
                        wsLocalMeaningsES.cell(row=meaningESindex, column=Globals.MEANINGS_COL_TYPE).value = first_meaning_type if use_english_type else ""
                        wsLocalMeaningsES.cell(row=meaningESindex, column=Globals.MEANINGS_COL_SOURCE).value = "JM"
                        meaningIndexesES.append(str(meaningESindex))
                        meaningESindex += 1
                    workingWorksheet.cell(row=typesIndex,
                                          column=Globals.TYPES_COL_MEANINGS_ES if sheetNum < 2 else Globals.VERBS_COL_MEANINGS_ES
                                          ).value = ";".join(meaningIndexesES)

            typesIndex += 1
    # endregion

    # region Creating a list of words that are not in the local database
    newWords = []
    localWordsList.sort(key=lambda x: x.uniqueID)
    for word in jmDictWordsList:
        index_of_first_hit = binary_search(localWordsList, word.uniqueID)
        if index_of_first_hit == -1:
            newWords.append(word)

    print("Number of words left out: " + str(len(newWords)))

    # endregion

    # Saving the results
    localWordsWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Grammar - 3000 kanji - with foreign.xlsx')
    localVerbsWorkbook.save(filename=f'{Globals.OUTPUT_DIR}/Verbs - 3000 kanji - with foreign.xlsx')
