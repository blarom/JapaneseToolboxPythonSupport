# Japanese Toolbox suru verb comparator

import openpyxl
from urllib.request import urlopen
import urllib.parse
from urllib.error import HTTPError, URLError
import re

# region Getting the excel sheets
localWordsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Grammar - 3000 kanji.xlsx', data_only=True)
localVerbsWorkbook = openpyxl.load_workbook(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji.xlsx', data_only=True)

wsLocalMeanings = localWordsWorkbook["Meanings"]
wsLocalTypes = localWordsWorkbook["Types"]
wsLocalGrammar = localWordsWorkbook["Grammar"]
wsLocalVerbsForGrammar = localVerbsWorkbook["VerbsForGrammar"]
wsLocalVerbs = localVerbsWorkbook["Verbs"]

lastLocalMeaningsIndex = 1
while True:
    value = wsLocalMeanings.cell(row=lastLocalMeaningsIndex, column=3).value
    if not value:
        lastLocalMeaningsIndex -= 1
        break
    lastLocalMeaningsIndex += 1

lastLocalVerbsIndex = 1
while True:
    value = wsLocalVerbs.cell(row=lastLocalVerbsIndex, column=3).value
    value2 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 1, column=3).value
    value3 = wsLocalVerbs.cell(row=lastLocalVerbsIndex + 2, column=3).value
    if (not value) & (not value2) & (not value3):
        lastLocalVerbsIndex -= 1
        break
    lastLocalVerbsIndex += 1
# endregion

# region Running through the suru verbs and adding the local and online meanings
verbRowIndex = 1
while verbRowIndex < lastLocalVerbsIndex:

    if wsLocalVerbs.cell(row=verbRowIndex, column=1).value == 'suru verb'\
            and (wsLocalVerbs.cell(row=verbRowIndex, column=15).value is None
                     or wsLocalVerbs.cell(row=verbRowIndex, column=15).value == ''):

        kanjiRoot = wsLocalVerbs.cell(row=verbRowIndex, column=8).value
        romajiRoot = wsLocalVerbs.cell(row=verbRowIndex, column=9).value
        meaningIndexesAsString = wsLocalVerbs.cell(row=verbRowIndex, column=12).value
        try:
            meaningIndexes = [int(meaningIndexAsString.strip()) for meaningIndexAsString in
                              str(meaningIndexesAsString).split(";")]
        except:
            verbRowIndex += 1
            continue

        # Updating the english column with the local meanings
        meanings = []
        types = []
        for meaningIndex in meaningIndexes:
            meanings.append(wsLocalMeanings.cell(row=int(meaningIndex), column=2).value)
            types.append(wsLocalMeanings.cell(row=meaningIndex, column=3).value)

        wsLocalVerbs.cell(row=verbRowIndex, column=9).value = ', '.join(meanings)
        wsLocalVerbs.cell(row=verbRowIndex, column=15).value = ', '.join(types)

        # Updating the verbs sheet with the values extracted from wiktionary
        if kanjiRoot is None:
            verbRowIndex += 1
            continue

        url = "https://en.wiktionary.org/wiki/" + kanjiRoot
        url = urllib.parse.urlsplit(url)
        url = list(url)
        url[2] = urllib.parse.quote(url[2])
        url = urllib.parse.urlunsplit(url)

        try:
            # content = urlopen(url, timeout=10).read().decode('utf-8')
            content = urlopen(url).read().decode('utf-8')

            hasJapaneseVerb = re.search(r"Conjugation of \"<span class=\"Jpan\" lang=\"ja\">", content)
            if hasJapaneseVerb:
                match = re.search(r"title=\"する\"(.+?)id=\"(Conjugation|Usage)", content, re.M | re.S)
                extendedMeaningsContainerCenter = match.group(0)

                if '>transitive<' in extendedMeaningsContainerCenter and '>intransitive<' in extendedMeaningsContainerCenter:
                    wsLocalVerbs.cell(row=verbRowIndex, column=16).value = 'TI'
                elif '>transitive<' in extendedMeaningsContainerCenter:
                    wsLocalVerbs.cell(row=verbRowIndex, column=16).value = 'T'
                elif '>intransitive<' in extendedMeaningsContainerCenter:
                    wsLocalVerbs.cell(row=verbRowIndex, column=16).value = 'I'
                else:
                    wsLocalVerbs.cell(row=verbRowIndex, column=16).value = 'N/A'

                match = re.search(r"title=\"suru\"(.+?)id=\"(Conjugation|Usage)", content, re.M | re.S)
                meaningsContainerCenter = match.group(0)
                meanings = re.findall(r">(.*?)<", meaningsContainerCenter)

                meanings = [meaning for meaning in meanings if (meaning != ''
                                                                and meaning != '\n'
                                                                and meaning != 'historical hiragana'
                                                                and meaning != 'suru'
                                                                and meaning != 'する')]

                wsLocalVerbs.cell(row=verbRowIndex, column=17).value = ', '.join(meanings)
        except URLError as error:
            print("Request for " + kanjiRoot + " timed out.")
            pass
        except:
            print("Request for " + kanjiRoot + " failed.")
            pass

        print("\rCompleted request No. " + str(verbRowIndex) + ": " + kanjiRoot)
        if verbRowIndex % 25 == 0:
            localVerbsWorkbook.save(
                filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with suru verbs data.xlsx')

    verbRowIndex += 1

# endregion

# region Saving the results
localVerbsWorkbook.save(
    filename='C:/Users/Bar/Dropbox/Japanese/Verbs - 3000 kanji - updated with suru verbs data.xlsx')
# endregion
