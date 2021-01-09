from openpyxl.utils.cell import column_index_from_string


def get_file_contents(filename):
    with open(filename, encoding="utf8") as fh:
        return fh.read()


def write_to_file(rwa, filename, content):
    fh = open(filename, rwa, encoding="utf8")
    fh.write(content)
    fh.close()


MASTER_DIR = 'C:/Users/Bar/Dropbox/Japanese'
OUTPUT_DIR = 'C:/Projects/Workspace/JapaneseToolboxPythonSupport'
JAPAGRAM_ASSETS_DIR = 'C:/Projects/Workspace/Japagram/app/src/main/assets'

MAX_VERB_CONG_COL = "GI"
DELIMITER = ";"

TYPES_COL_INDEX = 1
TYPES_COL_ROMAJI = 2
TYPES_COL_KANJI = 3
TYPES_COL_ALTS = 4
TYPES_COL_COMMON = 5
TYPES_COL_KW_JAP = 6
TYPES_COL_PREP = 7
TYPES_COL_KANJI_ROOT = 8
TYPES_COL_ROMAJI_ROOT = 9
TYPES_COL_EXCEPTION_INDEX = 10
TYPES_COL_MEANINGS_EN = 11
TYPES_COL_KW_EN = 12
TYPES_COL_MEANINGS_FR = 13
TYPES_COL_KW_FR = 14
TYPES_COL_MEANINGS_ES = 15
TYPES_COL_KW_ES = 16
TYPES_COL_HIRAGANA_FIRST_CHAR = 17
TYPES_COL_FREQUENCY = 18

MEANINGS_COL_INDEX = 1
MEANINGS_COL_MEANING = 2
MEANINGS_COL_TYPE = 3
MEANINGS_COL_EXPL = 4
MEANINGS_COL_RULES = 5
MEANINGS_COL_EXAMPLES = 6
MEANINGS_COL_ANTONYM = 7
MEANINGS_COL_SYNONYM = 8
MEANINGS_COL_SOURCE = 9

VERBS_COL_FAMILY = 1
VERBS_COL_SUMMARY_EN = 2
VERBS_COL_TI = 3
VERBS_COL_PREP = 4
VERBS_COL_KANJI = 5
VERBS_COL_ROMAJI = 6
VERBS_COL_KANJI_ROOT = 7
VERBS_COL_ROMAJI_ROOT = 8
VERBS_COL_EXCEPTION_INDEX = 9
VERBS_COL_ALTS = 10
VERBS_COL_KW_JAP = 11
VERBS_COL_COMMON = 12
VERBS_COL_MEANINGS_EN = 13
VERBS_COL_KW_EN = 14
VERBS_COL_MEANINGS_FR = 15
VERBS_COL_KW_FR = 16
VERBS_COL_MEANINGS_ES = 17
VERBS_COL_KW_ES = 18

VERBS_COL_CONJ_FAMILY = 1
VERBS_COL_CONJ_SUMMARY_EN = 2
VERBS_COL_CONJ_TI = 3
VERBS_COL_CONJ_PREP = 4
VERBS_COL_CONJ_KANA = 5
VERBS_COL_CONJ_KANJI = 6
VERBS_COL_CONJ_ROMAJI = 7
VERBS_COL_CONJ_KANJI_ROOT = 8
VERBS_COL_CONJ_ROMAJI_ROOT = 9
VERBS_COL_CONJ_FIRST_CONJ = 10

EXT_WORD_COL_INDEX = 1
EXT_WORD_COL_ROMAJI = 2
EXT_WORD_COL_KANJI = 3
EXT_WORD_COL_POS = 4
EXT_WORD_COL_ALTS = 5
EXT_WORD_COL_MEANINGS_EN = 6
EXT_WORD_COL_MEANINGS_FR = 7
EXT_WORD_COL_MEANINGS_ES = 8
EXT_WORD_COL_FREQUENCY = 9

NAMES_COL_INDEX = 1
NAMES_COL_ROMAJI = 2
NAMES_COL_KANJI = 3
NAMES_COL_CLASSIFICATION = 4
NAMES_COL_ENGLISH = 5

ROMANIZATIONS_HIRA = 0
ROMANIZATIONS_KATA = 1
ROMANIZATIONS_WAPU = 2
ROMANIZATIONS_HEPB = 3
ROMANIZATIONS_NISH = 4
ROMANIZATIONS_KUSH = 5

EDICT_EXCEPTIONS = [
    ["ha", "は"],
    ["wa", "は"],
    ["he", "へ"],
    ["e", "へ"],
    ["deha", "では"],
    ["dewa", "では"],
    ["niha", "には"],
    ["niwa", "には"],
    ["kana", "かな"],
    ["node", "ので"],
    ["nanode", "なので"],
    ["to", "と"],
    ["ya", "や"],
    ["mo", "も"],
    ["no", "の"],
    ["noga", "のが"],
    ["nowo", "のを"],
    ["n", "ん"],
    ["wo", "を"],
    ["wa", "わ"],
    ["yo", "よ"],
    ["na", "な"],
    ["ka", "か"],
    ["ga", "が"],
    ["ni", "に"],
    ["*", "ケ"],
    ["*", "ヶ"],
    ["noha", "のは"],
    ["nowa", "のは"],
    ["demo", "でも"],
    ["tte", "って"],
    ["datte", "だって"],
    ["temo", "ても"],
    ["ba", "ば"],
    ["nakereba", "なければ"],
    ["nakereba", "無ければ"],
    ["nakya", "なきゃ"],
    ["nakya", "無きゃ"],
    ["shi", "し"],
    ["kara", "から"],
    ["dakara", "だから"],
    ["tara", "たら"],
    ["datara", "だたら"],
    ["nakattara", "なかったら"],
    ["soshitara", "そしたら"],
    ["node", "ので"],
    ["nde", "んで"],
    ["te", "て"],
    ["noni", "のに"],
    ["nagara", "ながら"],
    ["nagara", "乍ら"],
    ["nara", "なら"],
    ["dano", "だの"],
    ["oyobi", "及び"],
    ["goto", "如"],
    ["nozoi", "除い"],
    ["made", "まで"],
    ["kara", "から"],
    ["toka", "とか"],
    ["yueni", "故に"],
    ["soko de", "其処で"],
    ["sore de", "それで"],
    ["ni yotte", "に因って"],
    ["you de", "ようで"],
    ["no made", "間で"],
    ["bakarini", "許りに"],
    ["kakawarazu", "拘らず"],
    ["soredemo", "それでも"],
    ["soreyori", "それより"],
    ["tadashi", "但し"],
    ["kedo", "けど"],
    ["keredomo", "けれども"],
    ["tokorode", "所で"],
    ["shikashi", "然し"],
    ["soreni", "其れに"],
    ["tari", "たり"],
    ["igai", "以外"],
    ["ato", "あと"],
    ["tameni", "為に"],
    ["tame", "為"],
    ["hazu", "筈"],
    ["nitsuite", "に就いて"],
    ["naru", "なる"],
    ["onaji", "同じ"],
    ["youni", "ように"],
    ["souna", "そうな"],
    ["yori", "より"],
    ["ato de", "後で"],
    ["maeni", "前に"],
    ["sorekara", "それから"],
    ["soshite", "然して"],

    ["e", "へ"],
    ["mo", "も"],
    ["to", "と"],
    ["ya", "や"],
    ["nado", "等"],
    ["no", "の"],
    ["n", "ん"],
    ["wo", "を"],
    ["wa", "は"],
    ["wa", "わ"],
    ["de", "で"],
    ["datte", "だって"],
    ["yo", "よ"],
    ["yone", "よね"],
    ["na", "な"],
    ["ka", "か"],
    ["dewa", "では"],
    ["kana", "かな"],
    ["ga", "が"],
    ["ka", "ヶ"],
    ["ne", "ね"],
    ["ni", "に"],
    ["demo", "でも"],
    ["tte", "って"],

    ["kedo", "けど"],
    ["tokorode", "所で"],
    ["tame", "為"],
    ["yori", "より"],
    ["toiu", "と言う"],

    ["mataha", "又は"],
    ["konnichiha", "今日は"],
    ["konnichiwa", "こんにちわ"],
    ["konnichiwa", "今日わ"],
    ["niyotte", "に因って"],

    ["soredeha", "其れでは"]
]

LATIN_CHAR_ALPHABET = "etaoinsrhdlcumwfgpybvkjxqzéóàüíáäêèãúôçâöñßùûîõìœëïòðåæþýøžš'"
LATIN_CHAR_ALPHABET_CAP = LATIN_CHAR_ALPHABET.upper()
NUMBER_ALPHABET = "1234567890'^"
SYMBOLS_ALPHABET = ". ,()/1234567890'^[];…!?-+*&:%$«»¿\"？"
WORD_LENGTH_THRESHOLD = 3
SECTION_LENGTH_THRESHOLD = 15

FREQ_DICT = {}


class Font:
    END = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    LIGHT_GREY = '\033[37m'
    DEFAULT = '\033[39m'
    DARK_GREY = '\033[90m'
    LIGHT_RED = '\033[91m'
    LIGHT_GREEN = '\033[92m'
    LIGHT_YELLOW = '\033[93m'
    LIGHT_BLUE = '\033[94m'
    LIGHT_MAGENTA = '\033[95m'
    LIGHT_CYAN = '\033[96m'
    WHITE = '\033[97m'


def is_latin(text):
    for char in list(text):
        if char in LATIN_CHAR_ALPHABET or char in LATIN_CHAR_ALPHABET_CAP:
            return True


def remove_duplicates_keep_order(seq):
    new_list = []
    for item in seq:
        if item not in new_list:
            new_list.append(item)
    return new_list


def convert_to_utf8(text):
    return '1.' + text.encode('utf-8').hex()


def convert_from_utf8(text):
    return bytearray.fromhex(text[2:]).decode()


def clearSheet(wb, sheet_name):
    index = wb.sheetnames.index(sheet_name)
    ws = wb[sheet_name]
    wb.remove(ws)
    wb.create_sheet(sheet_name, index)


def create_csv_from_worksheet(ws, csv_name, start_col, end_col, only_first_row=False, start_row=1):
    fh = open(f'{JAPAGRAM_ASSETS_DIR}/{csv_name}.csv', 'w+', encoding='utf-8')
    content_lines = []
    index = start_row
    while not isLastRow(ws, index):
        line = "|".join([str(ws.cell(row=index, column=col).value) if ws.cell(row=index, column=col).value is not None else "" for col in range(start_col, end_col + 1)])
        if line: content_lines.append(line)
        index += 1
        if only_first_row: break
    fh.write('\n'.join(content_lines) + '\n')
    fh.close()


def isLastRow(sheet, index):
    this_row_is_empty = all(not item for item in [sheet.cell(row=index, column=col).value for col in list(range(1, 10)) + list(range(41, 50))])
    next_row_is_empty = all(not item for item in [sheet.cell(row=index + 1, column=col).value for col in list(range(1, 10)) + list(range(41, 50))])
    col10_is_empty_from_next_row = all(not item for item in [sheet.cell(row=row, column=11).value for row in range(index + 1, index + 31)])
    return this_row_is_empty and next_row_is_empty and col10_is_empty_from_next_row


def get_cell_value(value):
    return str(value) if (value == 1 or value == 0 or value) else ''


def isIrrelevantRow(sheet, index):
    return all(not item for item in [sheet.cell(row=index, column=col).value for col in [TYPES_COL_ROMAJI, TYPES_COL_KANJI, TYPES_COL_ALTS, TYPES_COL_COMMON, TYPES_COL_KW_JAP]])


def name(text):
    return f"Line{text} - 3000 kanji"


def idx(text):
    return column_index_from_string(text)


def get_root_from_masu_stem_latin(latinStem, family):
    if ' godan' in family:
        stem = family.split(' ')[0]
        return latinStem[:-len(stem)]
    elif family == 'iku special class' or family == 'yuku special class' or family == 'aru special class' or family == 'ru ichidan':
        return latinStem[:-2]
    elif family == 'u special class':
        return latinStem[:-1]
    elif ' verb' in family:
        return latinStem[:-4]
    else:
        return latinStem


def get_root_from_masu_stem_kanji(kanjiStem, family):
    if ' godan' in family or 'special class' in family or family == 'ru ichidan':
        return kanjiStem[:-1]
    elif ' verb' in family:
        return kanjiStem[:-2]
    else:
        return kanjiStem


class Word:
    def __init__(self,
                 kanji='',
                 altSpellings=None,
                 hiragana='',
                 romaji='',
                 english_meanings=None,
                 french_meanings=None,
                 spanish_meanings=None,
                 types_for_word=None,
                 common=False):
        if altSpellings is None: altSpellings = []
        if types_for_word is None: types_for_word = []
        if english_meanings is None: english_meanings = []
        if french_meanings is None: french_meanings = []
        if spanish_meanings is None: spanish_meanings = []
        self.kanji = kanji
        self.altSpellings = altSpellings
        self.hiragana = hiragana
        self.romaji = romaji
        self.uniqueID = romaji + "zzz" + kanji
        self.english_meanings = english_meanings
        self.french_meanings = french_meanings
        self.spanish_meanings = spanish_meanings
        self.types_for_word = types_for_word
        self.common = common
